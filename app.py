import streamlit as st
import pandas as pd
import plotly.express as px
from datetime import datetime, timedelta
import calendar as py_calendar
import os
import secrets
import time
import bcrypt
import psycopg2
import psycopg2.extras
from psycopg2 import pool
from zoneinfo import ZoneInfo
from html import escape as html_escape

FUSO_BR = ZoneInfo('America/Sao_Paulo')

def hoje_projeto() -> datetime:
    """Retorna meia-noite de hoje no fuso de São Paulo — sempre atualizado."""
    return datetime.now(FUSO_BR).replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=None)

# Alias para compatibilidade com usos existentes como constante
# hoje_projeto() não é mais constante — use hoje_projeto() nos contextos de renderização

# ── Constantes de segurança ─────────────────────────────
MAX_TENTATIVAS_LOGIN  = 3
BLOQUEIO_MINUTOS      = 15
TIMEOUT_SESSAO_HORAS  = 8
LIMITE_REGISTROS_LOAD = 2000   # máx de linhas carregadas de uma vez

st.set_page_config(page_title="Passold Sistemas de Fachadas", layout="wide", page_icon="🏭", initial_sidebar_state="expanded")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap');
:root {
    --primary:        #1E3A5F;
    --primary-light:  #2A4F7C;
    --primary-hover:  #355F90;
    --accent:         #1A56DB;
    --accent-hover:   #1447C0;
    --accent-light:   #EFF6FF;
    --success:        #059669;
    --success-light:  #ECFDF5;
    --warning:        #D97706;
    --warning-light:  #FFFBEB;
    --danger:         #DC2626;
    --danger-light:   #FEF2F2;
    --info:           #0EA5E9;
    --info-light:     #F0F9FF;
    --bg:             #F1F5F9;
    --bg-card:        #FFFFFF;
    --border:         #E2E8F0;
    --border-hover:   #CBD5E1;
    --text:           #1E293B;
    --text-muted:     #64748B;
    --text-light:     #94A3B8;
    --sidebar-bg:     #1E3A5F;
    --sidebar-text:   #CBD5E1;
    --sidebar-active: #3B82F6;
    --shadow-xs:  0 1px 2px rgba(0,0,0,0.05);
    --shadow-sm:  0 1px 3px rgba(0,0,0,0.1), 0 1px 2px rgba(0,0,0,0.06);
    --shadow-md:  0 4px 6px -1px rgba(0,0,0,0.1), 0 2px 4px -1px rgba(0,0,0,0.06);
    --shadow-lg:  0 10px 15px -3px rgba(0,0,0,0.1), 0 4px 6px -2px rgba(0,0,0,0.05);
    --radius-sm:  6px;
    --radius:     10px;
    --radius-lg:  14px;
}

/* ── Base ───────────────────────────────────── */
.stApp { background: var(--bg); font-family: 'Inter', sans-serif; color: var(--text); }
#MainMenu, footer, header { visibility: hidden; }

/* ── Sidebar ────────────────────────────────── */
section[data-testid="stSidebar"] {
    background: var(--sidebar-bg) !important;
    border-right: none !important;
    box-shadow: 4px 0 24px rgba(0,0,0,0.15) !important;
}
/* Esconde o botão «» de fechar/abrir a sidebar (cobre versões antigas e novas do Streamlit) */
button[data-testid="collapsedControl"],
section[data-testid="stSidebar"] > div > div > button[kind="header"],
[data-testid*="ollapse" i] button,
button[data-testid*="ollapse" i],
[data-testid="stSidebarHeader"] button,
[data-testid="stSidebarCollapseButton"] {
    display: none !important;
    pointer-events: none !important;
}
/* ── Sidebar geral ──────────────────────────────────────── */
section[data-testid="stSidebar"] * { color: var(--sidebar-text) !important; }
section[data-testid="stSidebar"] hr { border-color: rgba(255,255,255,0.12) !important; margin: 8px 0 !important; }
section[data-testid="stSidebar"] .stCaption p { color: rgba(255,255,255,0.45) !important; font-size: 0.72rem !important; }
section[data-testid="stSidebar"] .stSelectbox > div > div {
    background: rgba(255,255,255,0.08) !important;
    border-color: rgba(255,255,255,0.15) !important;
    border-radius: var(--radius-sm) !important;
}

/* ── Esconde o label do radio de nav ──────────────────── */
section[data-testid="stSidebar"] .stRadio > div > label:first-child { display: none !important; }

/* ── Itens do menu ────────────────────────────────────── */
section[data-testid="stSidebar"] .stRadio > div > div[role="radiogroup"] {
    display: flex !important;
    flex-direction: column !important;
    gap: 2px !important;
}
section[data-testid="stSidebar"] .stRadio > div > div[role="radiogroup"] > label {
    display: flex !important;
    align-items: center !important;
    gap: 8px !important;
    padding: 9px 12px !important;
    border-radius: 6px !important;
    cursor: pointer !important;
    font-size: 0.87rem !important;
    font-weight: 500 !important;
    color: #94A3B8 !important;
    border: 1px solid transparent !important;
    transition: all 0.15s ease !important;
    background: transparent !important;
}
section[data-testid="stSidebar"] .stRadio > div > div[role="radiogroup"] > label:hover {
    background: rgba(255,255,255,0.07) !important;
    color: #fff !important;
}
section[data-testid="stSidebar"] .stRadio > div > div[role="radiogroup"] > label:has(input:checked) {
    background: rgba(59,130,246,0.2) !important;
    color: #fff !important;
    border-color: rgba(59,130,246,0.5) !important;
    font-weight: 700 !important;
}
/* esconde o círculo nativo do radio */
section[data-testid="stSidebar"] .stRadio input[type="radio"] {
    display: none !important;
}

/* ── Topbar ─────────────────────────────────── */
.topbar {
    display: flex; align-items: center; justify-content: space-between;
    background: var(--bg-card);
    border-bottom: 1px solid var(--border);
    padding: 14px 28px;
    margin: -1rem -1rem 1.5rem -1rem;
    box-shadow: var(--shadow-xs);
}
.topbar-title { font-size: 1.2rem; font-weight: 800; color: var(--primary); letter-spacing: -0.03em; }
.topbar-title span { color: var(--accent); }
.topbar-right { display: flex; align-items: center; gap: 20px; }
.topbar-user { font-size: 0.82rem; color: var(--text-muted); }
.topbar-user strong { color: var(--text); font-weight: 600; }
.topbar-badge {
    background: var(--accent-light); color: var(--accent);
    padding: 3px 10px; border-radius: 20px;
    font-size: 0.75rem; font-weight: 700;
    border: 1px solid rgba(26,86,219,0.2);
}
.topbar-time { font-size: 0.78rem; color: var(--text-muted); font-family: 'JetBrains Mono', monospace; }

/* ── Page Header ────────────────────────────── */
.page-header {
    display: flex; align-items: flex-start; justify-content: space-between;
    margin-bottom: 1.5rem;
    padding-bottom: 1rem;
    border-bottom: 1px solid var(--border);
}
.page-header-left h2 {
    font-size: 1.5rem !important; font-weight: 800 !important;
    color: var(--primary) !important; letter-spacing: -0.03em;
    margin: 0 0 4px 0 !important;
}
.page-header-left p { font-size: 0.85rem; color: var(--text-muted); margin: 0; }
.page-icon { font-size: 2rem; }

/* ── Cards ──────────────────────────────────── */
div[data-testid="metric-container"] {
    background: var(--bg-card) !important;
    border: 1px solid var(--border) !important;
    border-top: 4px solid var(--accent) !important;
    border-radius: var(--radius) !important;
    padding: 20px !important;
    box-shadow: var(--shadow-sm) !important;
    transition: transform 0.2s ease, box-shadow 0.2s ease !important;
}
div[data-testid="metric-container"]:hover {
    transform: translateY(-2px);
    box-shadow: var(--shadow-md) !important;
}
div[data-testid="stMetricValue"] {
    font-size: 2rem !important; font-weight: 800 !important;
    color: var(--primary) !important;
}
div[data-testid="stMetricLabel"] {
    color: var(--text-muted) !important; font-weight: 600 !important;
    text-transform: uppercase; font-size: 0.72rem !important;
    letter-spacing: 0.06em;
}
div[data-testid="stMetricDelta"] { font-size: 0.82rem !important; }

/* ── Botões ─────────────────────────────────── */
.stButton > button {
    background: var(--primary) !important;
    color: #fff !important;
    font-weight: 600 !important;
    border-radius: var(--radius-sm) !important;
    border: none !important;
    padding: 10px 22px !important;
    font-size: 0.88rem !important;
    box-shadow: var(--shadow-xs) !important;
    transition: all 0.18s ease !important;
    letter-spacing: 0.01em;
}
.stButton > button p, .stButton > button span, .stButton > button div { color: #fff !important; }
.stButton > button:hover {
    background: var(--primary-hover) !important;
    transform: translateY(-1px);
    box-shadow: var(--shadow-sm) !important;
}
.stButton > button[kind="primary"] { background: var(--accent) !important; }
.stButton > button[kind="primary"]:hover { background: var(--accent-hover) !important; }
.stButton > button[kind="secondary"] {
    background: transparent !important;
    color: var(--text) !important;
    border: 1px solid var(--border) !important;
    box-shadow: none !important;
}
.stButton > button[kind="secondary"] p,
.stButton > button[kind="secondary"] span { color: var(--text) !important; }
.stButton > button[kind="secondary"]:hover {
    background: var(--bg) !important;
    border-color: var(--border-hover) !important;
}

/* ── Inputs ─────────────────────────────────── */
.stTextInput > div > div > input,
.stNumberInput > div > div > input,
.stTextArea > div > div > textarea {
    background: var(--bg-card) !important;
    border: 1px solid var(--border) !important;
    border-radius: var(--radius-sm) !important;
    color: var(--text) !important;
    font-size: 0.9rem !important;
    box-shadow: none !important;
    transition: border-color 0.15s ease !important;
}
.stTextInput > div > div > input:focus,
.stNumberInput > div > div > input:focus,
.stTextArea > div > div > textarea:focus {
    border-color: var(--accent) !important;
    box-shadow: 0 0 0 3px rgba(26,86,219,0.1) !important;
}

/* ── Tabelas ────────────────────────────────── */
div[data-testid="stDataFrame"] {
    border-radius: var(--radius) !important;
    border: 1px solid var(--border) !important;
    overflow: hidden;
    box-shadow: var(--shadow-xs);
}

/* ── Expander ───────────────────────────────── */
div[data-testid="stExpander"] {
    border: 1px solid var(--border) !important;
    border-radius: var(--radius) !important;
    box-shadow: var(--shadow-xs) !important;
    overflow: hidden;
    background: var(--bg-card) !important;
}
div[data-testid="stExpander"]:hover { border-color: var(--border-hover) !important; }

/* ── Tabs internas ──────────────────────────── */
.stTabs [data-baseweb="tab-list"] {
    gap: 0; border-bottom: 2px solid var(--border);
    background: transparent;
}
.stTabs [data-baseweb="tab"] {
    height: 44px; background: transparent;
    color: var(--text-muted); font-weight: 500;
    font-size: 0.88rem; padding: 0 20px;
    transition: all 0.15s ease;
    border-bottom: 2px solid transparent;
    margin-bottom: -2px;
}
.stTabs [data-baseweb="tab"]:hover { color: var(--text); background: var(--bg); }
.stTabs [aria-selected="true"] {
    color: var(--accent) !important; font-weight: 700 !important;
    border-bottom: 2px solid var(--accent) !important;
    background: transparent !important;
}

/* ── Alerts ─────────────────────────────────── */
div[data-testid="stAlert"] { border-radius: var(--radius-sm) !important; }

/* ── Badges & Barras ────────────────────────── */
.badge-obra  { background:var(--accent-light); color:#1447C0; padding:3px 10px; border-radius:20px; font-weight:700; font-size:11px; text-transform:uppercase; letter-spacing:0.05em; border:1px solid rgba(20,71,192,0.2); }
.badge-edt   { background:#F1F5F9; color:#334155; padding:3px 10px; border-radius:20px; font-weight:600; font-size:11px; border:1px solid var(--border); }
.badge-lote  { background:var(--success-light); color:#047857; padding:3px 10px; border-radius:20px; font-weight:700; font-size:11px; border:1px solid rgba(4,120,87,0.2); }
.badge-op    { background:var(--info-light); color:#0369A1; padding:3px 10px; border-radius:20px; font-weight:700; font-size:11px; border:1px solid rgba(3,105,161,0.2); }
.bar-ok      { border-left:4px solid var(--success); background:var(--success-light); padding:12px 16px; border-radius:var(--radius-sm); margin-bottom:8px; box-shadow:var(--shadow-xs); }
.bar-warn    { border-left:4px solid var(--warning); background:var(--warning-light); padding:12px 16px; border-radius:var(--radius-sm); margin-bottom:8px; box-shadow:var(--shadow-xs); }
.bar-danger  { border-left:4px solid var(--danger);  background:var(--danger-light);  padding:12px 16px; border-radius:var(--radius-sm); margin-bottom:8px; box-shadow:var(--shadow-xs); }
.bar-neutral { border-left:4px solid var(--text-muted); background:var(--bg); padding:12px 16px; border-radius:var(--radius-sm); margin-bottom:8px; box-shadow:var(--shadow-xs); }

/* ── Pipeline de OP ─────────────────────────── */
.pipeline { display:flex; align-items:center; gap:0; margin:16px 0; }
.pipeline-step {
    display:flex; align-items:center; gap:8px;
    padding:8px 16px; border-radius:var(--radius-sm);
    font-size:0.78rem; font-weight:600;
    background:var(--bg); border:1px solid var(--border);
    color:var(--text-muted);
    position:relative;
}
.pipeline-step.done   { background:var(--success-light); border-color:rgba(5,150,105,0.3); color:var(--success); }
.pipeline-step.active { background:var(--accent-light);  border-color:rgba(26,86,219,0.3);  color:var(--accent); font-weight:700; }
.pipeline-arrow { color:var(--border); font-size:1rem; padding:0 4px; }

/* ── Seção cards do dashboard ───────────────── */
.dash-card {
    background:var(--bg-card); border:1px solid var(--border);
    border-radius:var(--radius); padding:20px 24px;
    box-shadow:var(--shadow-sm); height:100%;
    transition: box-shadow 0.2s ease;
}
.dash-card:hover { box-shadow:var(--shadow-md); }
.dash-card-title { font-size:0.72rem; font-weight:700; text-transform:uppercase; letter-spacing:0.08em; color:var(--text-muted); margin-bottom:8px; }
.dash-card-value { font-size:2.4rem; font-weight:800; color:var(--primary); line-height:1; }
.dash-card-value.red   { color:var(--danger); }
.dash-card-value.orange{ color:var(--accent); }
.dash-card-value.green { color:var(--success); }
.dash-card-sub  { font-size:0.78rem; color:var(--text-muted); margin-top:6px; }

/* ── Sidebar nav label ──────────────────────── */
.nav-section-label {
    font-size:0.65rem; font-weight:700; text-transform:uppercase;
    letter-spacing:0.12em; color:rgba(255,255,255,0.3);
    padding:16px 14px 6px; margin:0;
}

/* ── Empty state ────────────────────────────── */
.empty-state {
    text-align:center; padding:48px 24px;
    color:var(--text-muted);
}
.empty-state .empty-icon { font-size:2.5rem; margin-bottom:12px; }
.empty-state h4 { font-size:1rem; font-weight:600; color:var(--text); margin-bottom:6px; }
.empty-state p  { font-size:0.85rem; }

/* ── Sidebar logo area ──────────────────────── */
.sidebar-logo {
    padding:20px 16px 16px;
    border-bottom:1px solid rgba(255,255,255,0.08);
    margin-bottom:8px;
}
.sidebar-logo-name { font-size:0.95rem; font-weight:800; color:#fff; letter-spacing:-0.02em; }
.sidebar-logo-sub  { font-size:0.7rem; color:rgba(255,255,255,0.4); margin-top:2px; }
</style>
""", unsafe_allow_html=True)

# ========================================================
# CONNECTION POOL — criado uma única vez para todos os usuários
# ========================================================
@st.cache_resource
def get_connection_pool():
    return pool.ThreadedConnectionPool(
        minconn=2,
        maxconn=20,
        dsn=st.secrets["supabase"]["url"]
    )

def conectar_banco():
    return get_connection_pool().getconn()

def liberar_conexao(conn):
    try:
        get_connection_pool().putconn(conn)
    except Exception:
        pass

# ========================================================
# SENHAS — bcrypt com salt
# ========================================================
def hash_senha(senha: str) -> str:
    return bcrypt.hashpw(senha.encode(), bcrypt.gensalt()).decode()

def verificar_senha(senha: str, hash_salvo: str) -> bool:
    try:
        return bool(bcrypt.checkpw(senha.encode(), hash_salvo.encode()))
    except Exception:
        return False


# ── Rate limiting de login — persistido no banco ─────────
# Resiste a refresh de página e novas sessões do mesmo usuário.

def verificar_bloqueio(usuario: str) -> tuple:
    """Retorna (bloqueado: bool, segundos_restantes: int)."""
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT tentativas, bloqueado_ate FROM login_tentativas WHERE usuario=%s",
            (usuario,)
        )
        row = cursor.fetchone()
        if not row:
            return False, 0
        tentativas, bloqueado_ate_str = row
        if bloqueado_ate_str:
            bloqueado_ate = datetime.fromisoformat(bloqueado_ate_str)
            if datetime.now() < bloqueado_ate:
                return True, int((bloqueado_ate - datetime.now()).total_seconds())
        return False, 0
    except Exception:
        return False, 0
    finally:
        liberar_conexao(conn)

def registrar_tentativa_falha(usuario: str):
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT tentativas FROM login_tentativas WHERE usuario=%s", (usuario,)
        )
        row = cursor.fetchone()
        tentativas = (row[0] if row else 0) + 1
        bloqueado_ate = None
        if tentativas >= MAX_TENTATIVAS_LOGIN:
            bloqueado_ate = (datetime.now() + timedelta(minutes=BLOQUEIO_MINUTOS)).isoformat()
            tentativas = 0
        cursor.execute(
            """INSERT INTO login_tentativas (usuario, tentativas, bloqueado_ate)
               VALUES (%s, %s, %s)
               ON CONFLICT (usuario) DO UPDATE
               SET tentativas=%s, bloqueado_ate=%s""",
            (usuario, tentativas, bloqueado_ate, tentativas, bloqueado_ate)
        )
        conn.commit()
    except Exception:
        pass
    finally:
        liberar_conexao(conn)

def resetar_tentativas(usuario: str):
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute(
            "DELETE FROM login_tentativas WHERE usuario=%s", (usuario,)
        )
        conn.commit()
    except Exception:
        pass
    finally:
        liberar_conexao(conn)

def tentativas_restantes(usuario: str) -> int:
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT tentativas FROM login_tentativas WHERE usuario=%s", (usuario,))
        row = cursor.fetchone()
        return MAX_TENTATIVAS_LOGIN - (row[0] if row else 0)
    except Exception:
        return MAX_TENTATIVAS_LOGIN
    finally:
        liberar_conexao(conn)

# ── Timeout de sessão ────────────────────────────────────
def registrar_atividade():
    st.session_state["ultima_atividade"] = datetime.now()

def verificar_timeout_sessao() -> bool:
    """Retorna True se a sessão expirou por inatividade."""
    ultima = st.session_state.get("ultima_atividade")
    if ultima is None:
        return False
    return (datetime.now() - ultima).total_seconds() > TIMEOUT_SESSAO_HORAS * 3600

# ── Log de auditoria persistente ────────────────────────
def registrar_auditoria(usuario: str, acao: str, detalhes: str = ""):
    """Grava ação crítica no banco. Nunca quebra o app se falhar."""
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO auditoria_log (usuario, acao, detalhes, criado_em) VALUES (%s, %s, %s, %s)",
            (usuario, acao, detalhes, datetime.now(FUSO_BR).strftime('%d/%m/%Y %H:%M:%S'))
        )
        conn.commit()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
    finally:
        liberar_conexao(conn)



# ========================================================
# INICIALIZAÇÃO DAS TABELAS
# ========================================================
def inicializar_banco_de_dados():
    conn = conectar_banco()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS cronograma_macro (
                id SERIAL PRIMARY KEY,
                Obra TEXT,
                EDT TEXT UNIQUE,
                Tipo_Escopo TEXT,
                Etapa_Macro TEXT,
                Subdivisao TEXT,
                Tarefa TEXT,
                M2_Total_Tarefa REAL,
                Inicio_Previsto DATE,
                Termino_Obra DATE,
                Status TEXT DEFAULT 'Pendente',
                Status_Engenharia TEXT DEFAULT 'Aguardando Medicao In Loco',
                Prazo_Engenharia DATE,
                Data_Limite_Despacho DATE,
                Primeiro_Dia_Producao DATE
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS itens_detalhado (
                id SERIAL PRIMARY KEY,
                Obra_Vinculada TEXT,
                EDT_Vinculado TEXT,
                Cod_Lote TEXT,
                Num_OP TEXT,
                Tipo_Material TEXT,
                Qtd_Caixas INTEGER,
                M2_Item REAL,
                Data_Producao_Programada DATE,
                Data_Limite_Obra DATE,
                Data_Despacho DATE,
                Romaneio_Chapas TEXT,
                Status_Item TEXT DEFAULT 'Pendente',
                Dificuldade INTEGER DEFAULT 3,
                Fase_Produtiva TEXT,
                Enviado_Logistica INTEGER DEFAULT 0,
                updated_at TIMESTAMP DEFAULT NOW()
            )
        """)
        cursor.execute("ALTER TABLE itens_detalhado ADD COLUMN IF NOT EXISTS Data_Despacho DATE")
        cursor.execute("ALTER TABLE itens_detalhado ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP DEFAULT NOW()")
        cursor.execute("ALTER TABLE itens_detalhado ADD COLUMN IF NOT EXISTS Peso_Kg REAL DEFAULT 0.0")
        cursor.execute("ALTER TABLE itens_detalhado ADD COLUMN IF NOT EXISTS em_parada BOOLEAN DEFAULT FALSE")
        cursor.execute("ALTER TABLE itens_detalhado ADD COLUMN IF NOT EXISTS motivo_parada TEXT")
        cursor.execute("ALTER TABLE itens_detalhado ADD COLUMN IF NOT EXISTS Escopo TEXT")
        cursor.execute("ALTER TABLE itens_detalhado ADD COLUMN IF NOT EXISTS Numero_Projeto TEXT")
        cursor.execute("ALTER TABLE itens_detalhado ADD COLUMN IF NOT EXISTS Uso_Interno BOOLEAN DEFAULT FALSE")
        cursor.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_itens_detalhado_num_op ON itens_detalhado(Num_OP) WHERE Num_OP IS NOT NULL")
        # Escopo oficial (enum unico usado em toda a aplicacao): ACM / Esquadria-Vidro / Terceirizada.
        # Itens ligados a EDT herdam o Tipo_Escopo do cronograma (normalizado pro enum);
        # OP Avulsa recupera o escopo do texto legado gravado em Fase_Produtiva na hora do cadastro.
        cursor.execute("""
            UPDATE itens_detalhado i
            SET Escopo = CASE
                WHEN m.Tipo_Escopo ~* 'ESQUADRIA|VIDRO' THEN 'Esquadria-Vidro'
                WHEN m.Tipo_Escopo ~* 'TERCEIRIZAD'      THEN 'Terceirizada'
                WHEN m.Tipo_Escopo ~* 'ACM'              THEN 'ACM'
                ELSE NULL
            END
            FROM cronograma_macro m
            WHERE i.Escopo IS NULL
              AND i.EDT_Vinculado = m.EDT
        """)
        cursor.execute("""
            UPDATE itens_detalhado
            SET Escopo = CASE
                WHEN Fase_Produtiva LIKE 'OP AVULSA — ACM%'       THEN 'ACM'
                WHEN Fase_Produtiva LIKE 'OP AVULSA — Esquadria%' THEN 'Esquadria-Vidro'
                WHEN Fase_Produtiva LIKE 'OP AVULSA — Vidro%'     THEN 'Esquadria-Vidro'
                WHEN Fase_Produtiva LIKE 'OP AVULSA — Outro%'     THEN 'Terceirizada'
            END
            WHERE Escopo IS NULL AND EDT_Vinculado = 'AVULSO' AND Fase_Produtiva LIKE 'OP AVULSA —%'
        """)
        cursor.execute("""
            UPDATE itens_detalhado
            SET Escopo = CASE
                WHEN Tipo_Material ~* 'ESQUADRIA|VIDRO|ALUMINIO|PERFIL' THEN 'Esquadria-Vidro'
                ELSE 'ACM'
            END
            WHERE Escopo IS NULL
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS arquivos_op (
                id SERIAL PRIMARY KEY,
                item_id INTEGER REFERENCES itens_detalhado(id) ON DELETE CASCADE,
                nome_arquivo TEXT NOT NULL,
                tipo_arquivo TEXT,
                conteudo BYTEA NOT NULL,
                enviado_por TEXT,
                enviado_em TIMESTAMP DEFAULT NOW()
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS romaneios_manuais (
                id SERIAL PRIMARY KEY,
                obra_vinculada TEXT NOT NULL,
                data_recebimento DATE NOT NULL,
                criado_por TEXT,
                criado_em TIMESTAMP DEFAULT NOW()
            )
        """)
        cursor.execute("""
            DO $$
            BEGIN
                IF EXISTS (SELECT 1 FROM information_schema.columns WHERE table_name='romaneios_manuais' AND column_name='descricao') THEN
                    ALTER TABLE romaneios_manuais ALTER COLUMN descricao DROP NOT NULL;
                END IF;
                IF EXISTS (SELECT 1 FROM information_schema.columns WHERE table_name='romaneios_manuais' AND column_name='quantidade') THEN
                    ALTER TABLE romaneios_manuais ALTER COLUMN quantidade DROP NOT NULL;
                END IF;
            END $$;
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS romaneios_manuais_itens (
                id SERIAL PRIMARY KEY,
                romaneio_id INTEGER REFERENCES romaneios_manuais(id) ON DELETE CASCADE,
                descricao TEXT NOT NULL,
                quantidade NUMERIC NOT NULL
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS usuarios (
                id SERIAL PRIMARY KEY,
                usuario TEXT UNIQUE,
                nome TEXT,
                setor TEXT,
                senha TEXT
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS solicitacoes_prazo (
                id SERIAL PRIMARY KEY,
                edt TEXT,
                tarefa TEXT,
                prazo_atual TEXT,
                prazo_solicitado TEXT,
                justificativa TEXT,
                criado_por TEXT,
                status TEXT DEFAULT 'Pendente de Aprovacao',
                criado_em TEXT
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS logistica_envios (
                id SERIAL PRIMARY KEY,
                item_id INTEGER,
                Obra_Vinculada TEXT,
                EDT_Vinculado TEXT,
                Cod_Lote TEXT,
                Num_OP TEXT,
                Tipo_Material TEXT,
                Qtd_Caixas INTEGER,
                M2_Item REAL,
                Romaneio_Chapas TEXT,
                Data_Limite_Despacho DATE,
                Data_Envio_Agendado DATE,
                Transportadora TEXT,
                Veiculo TEXT,
                Observacoes TEXT,
                Status_Logistica TEXT DEFAULT 'Aguardando Agendamento',
                Confirmado_Por TEXT,
                Confirmado_Em TEXT
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS componentes_op (
                id SERIAL PRIMARY KEY,
                item_id INTEGER,
                Obra_Vinculada TEXT,
                Cod_Lote TEXT,
                Num_OP TEXT,
                Nome_Componente TEXT,
                Quantidade REAL,
                Unidade TEXT,
                Status_Item TEXT DEFAULT 'Aguardando Conferencia',
                Observacao TEXT,
                Conferido_Por TEXT,
                Conferido_Em TEXT
            )
        """)
        # ── TABELAS DO SISTEMA DE MEDIÇÃO ──────────────────────────
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS medicao_obras (
                id SERIAL PRIMARY KEY,
                nome TEXT NOT NULL,
                valor_m2_global REAL NOT NULL,
                metragem_geral REAL NOT NULL,
                criado_em TIMESTAMP DEFAULT NOW()
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS medicao_servicos (
                id SERIAL PRIMARY KEY,
                obra_id INTEGER NOT NULL REFERENCES medicao_obras(id) ON DELETE CASCADE,
                nome TEXT NOT NULL,
                valor_m2_servico REAL NOT NULL
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS medicao_subdivisoes (
                id SERIAL PRIMARY KEY,
                servico_id INTEGER NOT NULL REFERENCES medicao_servicos(id) ON DELETE CASCADE,
                nome TEXT NOT NULL,
                m2 REAL NOT NULL DEFAULT 0,
                percentual REAL NOT NULL DEFAULT 0
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS medicao_historico (
                id SERIAL PRIMARY KEY,
                obra_id INTEGER NOT NULL REFERENCES medicao_obras(id) ON DELETE CASCADE,
                periodo TEXT NOT NULL,
                total_medido REAL NOT NULL DEFAULT 0,
                snapshot JSONB,
                criado_em TIMESTAMP DEFAULT NOW(),
                UNIQUE(obra_id, periodo)
            )
        """)
        # ── TABELA DE PEÇAS DAS OPs ────────────────────────────────
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS op_pecas (
                id SERIAL PRIMARY KEY,
                lote_id INTEGER REFERENCES itens_detalhado(id) ON DELETE CASCADE,
                obra TEXT,
                cod_lote TEXT,
                num_op TEXT,
                codigo TEXT NOT NULL,
                localizacao TEXT,
                medida TEXT,
                qtd_total INTEGER DEFAULT 0,
                qtd_enviada INTEGER DEFAULT 0,
                saldo INTEGER DEFAULT 0,
                m2_op_real REAL DEFAULT 0,
                componentes_status TEXT DEFAULT 'Aguardando Projetista',
                criado_em TIMESTAMP DEFAULT NOW()
            )
        """)
        # Coluna de m2 executado na etapa
        cursor.execute("ALTER TABLE cronograma_macro ADD COLUMN IF NOT EXISTS m2_executado REAL DEFAULT 0")
        cursor.execute("ALTER TABLE cronograma_macro ADD COLUMN IF NOT EXISTS Numero_Projeto TEXT DEFAULT ''")
        cursor.execute("ALTER TABLE cronograma_macro ADD COLUMN IF NOT EXISTS Detalhado BOOLEAN DEFAULT TRUE")
        cursor.execute("ALTER TABLE op_pecas ADD COLUMN IF NOT EXISTS m2_op_real REAL DEFAULT 0")
        cursor.execute("ALTER TABLE op_pecas ADD COLUMN IF NOT EXISTS descricao TEXT DEFAULT ''")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_op_pecas_lote ON op_pecas(lote_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_op_pecas_obra ON op_pecas(obra)")

        # Orcamento de peso (kg) da frente, espelhando M2_Total_Tarefa/m2_executado — necessario
        # porque frentes de Esquadria-Vidro sao medidas em kg, nao em m2.
        cursor.execute("ALTER TABLE cronograma_macro ADD COLUMN IF NOT EXISTS Peso_Total_Tarefa REAL")
        cursor.execute("ALTER TABLE cronograma_macro ADD COLUMN IF NOT EXISTS peso_executado REAL DEFAULT 0")
        cursor.execute("ALTER TABLE op_pecas ADD COLUMN IF NOT EXISTS peso_op_real REAL DEFAULT 0")

        # Numero_Projeto em itens_detalhado: itens ligados a EDT herdam da frente; OP Avulsa recupera
        # do padrao "PRJ-{n}" ja gravado em Romaneio_Chapas na hora do cadastro avulso.
        cursor.execute("""
            UPDATE itens_detalhado i
            SET Numero_Projeto = m.Numero_Projeto
            FROM cronograma_macro m
            WHERE (i.Numero_Projeto IS NULL OR i.Numero_Projeto = '')
              AND i.EDT_Vinculado = m.EDT
        """)
        cursor.execute("""
            UPDATE itens_detalhado
            SET Numero_Projeto = substring(Romaneio_Chapas FROM 'PRJ-([0-9A-Za-z]+)')
            WHERE (Numero_Projeto IS NULL OR Numero_Projeto = '')
              AND EDT_Vinculado = 'AVULSO' AND Romaneio_Chapas ~ 'PRJ-'
        """)
        # Reserva de m2/kg na criacao do lote (Mudanca 2): dai pra frente, m2_executado (ACM) e
        # peso_executado (Esquadria-Vidro) passam a ser debitados no momento da criacao do lote
        # (salvar_lotes_micro), nao mais so no lancamento de pecas. Bootstrap UNICO pra somar os
        # lotes ja existentes que nunca tiveram peca lancada (logo nunca contribuiram pro saldo ate
        # hoje), travado por sentinela pra rodar exatamente uma vez mesmo com varios restarts.
        cursor.execute("ALTER TABLE cronograma_macro ADD COLUMN IF NOT EXISTS m2_reserva_migrada BOOLEAN DEFAULT FALSE")
        cursor.execute("""
            UPDATE cronograma_macro m
            SET m2_executado = COALESCE(m.m2_executado, 0) + COALESCE((
                SELECT SUM(i.M2_Item) FROM itens_detalhado i
                WHERE i.EDT_Vinculado = m.EDT
                  AND NOT EXISTS (SELECT 1 FROM op_pecas p WHERE p.lote_id = i.id)
            ), 0),
                peso_executado = COALESCE(m.peso_executado, 0) + COALESCE((
                SELECT SUM(i.Peso_Kg) FROM itens_detalhado i
                WHERE i.EDT_Vinculado = m.EDT
                  AND NOT EXISTS (SELECT 1 FROM op_pecas p WHERE p.lote_id = i.id)
            ), 0),
                m2_reserva_migrada = TRUE
            WHERE COALESCE(m.m2_reserva_migrada, FALSE) = FALSE
        """)

        # ── TABELA DE PROJETOS (agrupador dentro da Obra — uma obra pode ter varios) ──
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS projetos (
                id SERIAL PRIMARY KEY,
                Obra TEXT NOT NULL,
                Numero_Projeto TEXT NOT NULL,
                criado_em TIMESTAMP DEFAULT NOW(),
                UNIQUE(Obra, Numero_Projeto)
            )
        """)
        # Migracao: registra como projeto qualquer Numero_Projeto ja usado em frentes ou lotes,
        # pra ninguem perder o historico que ja existe hoje.
        cursor.execute("""
            INSERT INTO projetos (Obra, Numero_Projeto)
            SELECT DISTINCT Obra, Numero_Projeto FROM cronograma_macro
            WHERE Numero_Projeto IS NOT NULL AND Numero_Projeto != ''
            ON CONFLICT (Obra, Numero_Projeto) DO NOTHING
        """)
        cursor.execute("""
            INSERT INTO projetos (Obra, Numero_Projeto)
            SELECT DISTINCT Obra_Vinculada, Numero_Projeto FROM itens_detalhado
            WHERE Numero_Projeto IS NOT NULL AND Numero_Projeto != ''
            ON CONFLICT (Obra, Numero_Projeto) DO NOTHING
        """)

        # ── TABELA DE SOLICITAÇÕES DE OP (PCP → Master) ────────────
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS solicitacoes_op (
                id SERIAL PRIMARY KEY,
                obra TEXT,
                numero_projeto TEXT,
                tipo_material TEXT,
                descricao TEXT,
                quantidade REAL DEFAULT 0,
                observacao TEXT,
                status TEXT DEFAULT 'Aguardando Vinculacao',
                solicitado_por TEXT,
                criado_em TIMESTAMP DEFAULT NOW(),
                item_id INTEGER REFERENCES itens_detalhado(id),
                vinculado_por TEXT,
                vinculado_em TIMESTAMP
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS arquivos_solicitacao_op (
                id SERIAL PRIMARY KEY,
                solicitacao_id INTEGER REFERENCES solicitacoes_op(id) ON DELETE CASCADE,
                nome_arquivo TEXT NOT NULL,
                tipo_arquivo TEXT,
                conteudo BYTEA NOT NULL,
                enviado_por TEXT,
                enviado_em TIMESTAMP DEFAULT NOW()
            )
        """)
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_solicitacoes_op_status ON solicitacoes_op(status)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_arquivos_solicitacao_solicitacao ON arquivos_solicitacao_op(solicitacao_id)")

        # ── ÍNDICES DE PERFORMANCE ──────────────────────────────────
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_itens_obra ON itens_detalhado(Obra_Vinculada)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_itens_status ON itens_detalhado(Status_Item)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_itens_edt ON itens_detalhado(EDT_Vinculado)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_macro_obra ON cronograma_macro(Obra)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_logistica_status ON logistica_envios(Status_Logistica)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_medicao_hist_obra ON medicao_historico(obra_id)")

# ── Tabela de auditoria (nova) ──────────────────────
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS auditoria_log (
                id        SERIAL PRIMARY KEY,
                usuario   TEXT,
                acao      TEXT,
                detalhes  TEXT,
                criado_em TEXT
            )
        """)
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_auditoria_usuario ON auditoria_log(usuario)")

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS login_tentativas (
                usuario      TEXT PRIMARY KEY,
                tentativas   INTEGER DEFAULT 0,
                bloqueado_ate TEXT
            )
        """)

        cursor.execute("SELECT COUNT(*) FROM usuarios")
        if cursor.fetchone()[0] == 0:
            _senha_master = st.secrets.get("MASTER_PASS") or os.environ.get("MASTER_PASS")
            if not _senha_master:
                _senha_master = secrets.token_urlsafe(16)
                st.warning(
                    f"⚠️ Conta Master criada com senha temporária gerada automaticamente: **{_senha_master}** "
                    "— anote agora, ela não será mostrada de novo. Configure MASTER_PASS em secrets.toml "
                    "para definir a senha inicial da próxima vez."
                )
            cursor.execute(
                "INSERT INTO usuarios (usuario, nome, setor, senha) VALUES (%s, %s, %s, %s)",
                ('master', 'Joice Master', 'Master', hash_senha(_senha_master))
            )
        conn.commit()
    except Exception as e:
        conn.rollback()
        # Não exibe detalhes técnicos (pode conter DSN com credenciais)
        st.error("Erro na inicialização do banco. Verifique as configurações de conexão.")
    finally:
        liberar_conexao(conn)

inicializar_banco_de_dados()

# ========================================================
# UTILITÁRIOS DE DATAS
# ========================================================
def subtrair_dias_uteis(data_base: datetime, n: int) -> datetime:
    data = data_base
    contados = 0
    while contados < n:
        data -= timedelta(days=1)
        if data.weekday() < 5:
            contados += 1
    return data

def calcular_cronograma_reverso(inicio_previsto, dias_logistica: int, dias_uteis_fabricacao: int, dias_antecedencia_eng: int = 3):
    dt = inicio_previsto if isinstance(inicio_previsto, datetime) else datetime.combine(inicio_previsto, datetime.min.time())
    data_limite_despacho  = dt - timedelta(days=int(dias_logistica))
    primeiro_dia_producao = subtrair_dias_uteis(data_limite_despacho, int(dias_uteis_fabricacao))
    prazo_engenharia      = subtrair_dias_uteis(primeiro_dia_producao, int(dias_antecedencia_eng))
    return prazo_engenharia, primeiro_dia_producao, data_limite_despacho

def gerar_lote_unico(data_limite_obra, dias_logistica, dias_uteis_fab,
                     total_cx, total_m2, obra, edt, cod_lote,
                     especificacao, txt_pav, dificuldade, total_kg=0.0,
                     escopo=None, numero_projeto=''):
    if isinstance(data_limite_obra, datetime):
        dt_limite = data_limite_obra
    else:
        dt_limite = datetime.combine(data_limite_obra, datetime.min.time())
    dt_despacho = dt_limite - timedelta(days=int(dias_logistica))
    dt_inicio   = subtrair_dias_uteis(dt_despacho, int(dias_uteis_fab))
    return [{
        "Obra_Vinculada":           obra,
        "EDT_Vinculado":            edt,
        "Cod_Lote":                 cod_lote,
        "Num_OP":                   None,
        "Tipo_Material":            especificacao,
        "Qtd_Caixas":               int(total_cx),
        "M2_Item":                  float(round(total_m2, 2)),
        "Peso_Kg":                  float(round(total_kg, 3)),
        "Data_Producao_Programada": dt_inicio.strftime('%Y-%m-%d'),
        "Data_Limite_Obra":         dt_limite.strftime('%Y-%m-%d'),
        "Data_Despacho":            dt_despacho.strftime('%Y-%m-%d'),
        "Romaneio_Chapas":          txt_pav,
        "Status_Item":              "Pendente",
        "Dificuldade":              int(dificuldade),
        "Fase_Produtiva":           f"CORTE→MONTAGEM ({dias_uteis_fab} dias uteis)",
        "Enviado_Logistica":        0,
        "Escopo":                   escopo,
        "Numero_Projeto":           numero_projeto or '',
    }]

def prazo_valido(valor) -> bool:
    if valor is None:
        return False
    try:
        return not pd.isnull(valor)
    except Exception:
        return False

# ========================================================
# FUNÇÕES DE BANCO — com cache, pool e try/finally
# ========================================================
@st.cache_data(ttl=30)
def carregar_projetos():
    conn = conectar_banco()
    try:
        df = pd.read_sql_query("SELECT * FROM projetos ORDER BY Obra, Numero_Projeto", conn)
    finally:
        liberar_conexao(conn)
    df.columns = [c.replace('_', ' ').title().replace(' ', '_') if c != 'id' else c for c in df.columns]
    return df

@st.cache_data(ttl=30)
def carregar_macro():
    conn = conectar_banco()
    try:
        with st.spinner("Carregando cronograma..."):
            df = pd.read_sql_query("SELECT * FROM cronograma_macro ORDER BY id", conn)
    finally:
        liberar_conexao(conn)
    for col in ['inicio_previsto', 'termino_obra', 'prazo_engenharia', 'data_limite_despacho', 'primeiro_dia_producao']:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')
    df.columns = [c.replace('_', ' ').title().replace(' ', '_') if c != 'id' else c for c in df.columns]
    rename = {
        'Inicio_Previsto': 'Inicio_Previsto', 'Termino_Obra': 'Termino_Obra',
        'Prazo_Engenharia': 'Prazo_Engenharia', 'Data_Limite_Despacho': 'Data_Limite_Despacho',
        'Primeiro_Dia_Producao': 'Primeiro_Dia_Producao', 'Edt': 'EDT',
        'M2_Total_Tarefa': 'M2_Total_Tarefa', 'Status_Engenharia': 'Status_Engenharia',
        'Tipo_Escopo': 'Tipo_Escopo', 'Etapa_Macro': 'Etapa_Macro',
    }
    return df.rename(columns=rename)

@st.cache_data(ttl=30)
def carregar_micro():
    """Lotes ativos + últimos 90 dias — evita carregar histórico inteiro."""
    conn = conectar_banco()
    try:
        with st.spinner("Carregando lotes..."):
            df = pd.read_sql_query(
                """SELECT * FROM itens_detalhado
                   WHERE Status_Item != 'Concluido'
                      OR Data_Producao_Programada >= NOW() - INTERVAL '90 days'
                   ORDER BY Data_Producao_Programada ASC
                   LIMIT %s""",
                conn, params=(LIMITE_REGISTROS_LOAD,)
            )
    finally:
        liberar_conexao(conn)
    for col in ['data_producao_programada', 'data_limite_obra', 'data_despacho']:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')
    df.columns = [c.replace('_', ' ').title().replace(' ', '_') if c != 'id' else c for c in df.columns]
    rename = {
        'Data_Producao_Programada': 'Data_Producao_Programada', 'Data_Limite_Obra': 'Data_Limite_Obra',
        'Data_Despacho': 'Data_Despacho', 'Edt_Vinculado': 'EDT_Vinculado',
        'Obra_Vinculada': 'Obra_Vinculada', 'Cod_Lote': 'Cod_Lote', 'Num_Op': 'Num_OP',
        'Tipo_Material': 'Tipo_Material', 'Qtd_Caixas': 'Qtd_Caixas', 'M2_Item': 'M2_Item',
        'Romaneio_Chapas': 'Romaneio_Chapas', 'Status_Item': 'Status_Item',
        'Fase_Produtiva': 'Fase_Produtiva', 'Enviado_Logistica': 'Enviado_Logistica',
        'Updated_At': 'Updated_At', 'Peso_Kg': 'Peso_Kg',
    }
    return df.rename(columns=rename)

@st.cache_data(ttl=60)
def carregar_micro_completo():
    """Histórico completo — usado no Relatório Geral. Limitado a 5000 registros mais recentes."""
    conn = conectar_banco()
    try:
        df = pd.read_sql_query(
            "SELECT * FROM itens_detalhado ORDER BY Data_Producao_Programada DESC LIMIT 5000", conn
        )
    finally:
        liberar_conexao(conn)
    for col in ['data_producao_programada', 'data_limite_obra', 'data_despacho']:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')
    df.columns = [c.replace('_', ' ').title().replace(' ', '_') if c != 'id' else c for c in df.columns]
    rename = {
        'Data_Producao_Programada': 'Data_Producao_Programada', 'Data_Limite_Obra': 'Data_Limite_Obra',
        'Data_Despacho': 'Data_Despacho', 'Edt_Vinculado': 'EDT_Vinculado',
        'Obra_Vinculada': 'Obra_Vinculada', 'Cod_Lote': 'Cod_Lote', 'Num_Op': 'Num_OP',
        'Tipo_Material': 'Tipo_Material', 'Qtd_Caixas': 'Qtd_Caixas', 'M2_Item': 'M2_Item',
        'Romaneio_Chapas': 'Romaneio_Chapas', 'Status_Item': 'Status_Item',
        'Fase_Produtiva': 'Fase_Produtiva', 'Enviado_Logistica': 'Enviado_Logistica',
        'Updated_At': 'Updated_At', 'Peso_Kg': 'Peso_Kg',
    }
    return df.rename(columns=rename)

@st.cache_data(ttl=30)
def carregar_micro_por_obra(obra: str):
    conn = conectar_banco()
    try:
        df = pd.read_sql_query(
            "SELECT * FROM itens_detalhado WHERE Obra_Vinculada = %s ORDER BY Data_Producao_Programada ASC",
            conn, params=(obra,)
        )
    finally:
        liberar_conexao(conn)
    for col in ['data_producao_programada', 'data_limite_obra', 'data_despacho']:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')
    df.columns = [c.replace('_', ' ').title().replace(' ', '_') if c != 'id' else c for c in df.columns]
    rename = {
        'Data_Producao_Programada': 'Data_Producao_Programada', 'Data_Limite_Obra': 'Data_Limite_Obra',
        'Data_Despacho': 'Data_Despacho', 'Edt_Vinculado': 'EDT_Vinculado',
        'Obra_Vinculada': 'Obra_Vinculada', 'Cod_Lote': 'Cod_Lote', 'Num_Op': 'Num_OP',
        'Tipo_Material': 'Tipo_Material', 'Qtd_Caixas': 'Qtd_Caixas', 'M2_Item': 'M2_Item',
        'Romaneio_Chapas': 'Romaneio_Chapas', 'Status_Item': 'Status_Item',
        'Fase_Produtiva': 'Fase_Produtiva', 'Enviado_Logistica': 'Enviado_Logistica',
        'Peso_Kg': 'Peso_Kg',
    }
    return df.rename(columns=rename)

@st.cache_data(ttl=30)
def carregar_macro_por_obra(obra: str):
    conn = conectar_banco()
    try:
        df = pd.read_sql_query(
            "SELECT * FROM cronograma_macro WHERE Obra = %s ORDER BY id", conn, params=(obra,)
        )
    finally:
        liberar_conexao(conn)
    for col in ['inicio_previsto', 'termino_obra', 'prazo_engenharia', 'data_limite_despacho', 'primeiro_dia_producao']:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')
    df.columns = [c.replace('_', ' ').title().replace(' ', '_') if c != 'id' else c for c in df.columns]
    rename = {
        'Inicio_Previsto': 'Inicio_Previsto', 'Termino_Obra': 'Termino_Obra',
        'Prazo_Engenharia': 'Prazo_Engenharia', 'Data_Limite_Despacho': 'Data_Limite_Despacho',
        'Primeiro_Dia_Producao': 'Primeiro_Dia_Producao', 'Edt': 'EDT',
        'M2_Total_Tarefa': 'M2_Total_Tarefa', 'Status_Engenharia': 'Status_Engenharia',
        'Tipo_Escopo': 'Tipo_Escopo', 'Etapa_Macro': 'Etapa_Macro',
    }
    return df.rename(columns=rename)

@st.cache_data(ttl=30)
def carregar_fila_logistica():
    conn = conectar_banco()
    try:
        df = pd.read_sql_query(
            "SELECT * FROM logistica_envios ORDER BY data_limite_despacho ASC NULLS LAST", conn
        )
    finally:
        liberar_conexao(conn)
    for col in ['data_limite_despacho', 'data_envio_agendado']:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')
    df.columns = [c.replace('_', ' ').title().replace(' ', '_') if c != 'id' else c for c in df.columns]
    rename = {
        'Data_Limite_Despacho': 'Data_Limite_Despacho', 'Data_Envio_Agendado': 'Data_Envio_Agendado',
        'Obra_Vinculada': 'Obra_Vinculada', 'Cod_Lote': 'Cod_Lote', 'Num_Op': 'Num_OP',
        'Tipo_Material': 'Tipo_Material', 'Qtd_Caixas': 'Qtd_Caixas', 'M2_Item': 'M2_Item',
        'Romaneio_Chapas': 'Romaneio_Chapas', 'Status_Logistica': 'Status_Logistica',
        'Transportadora': 'Transportadora', 'Veiculo': 'Veiculo', 'Observacoes': 'Observacoes',
        'Confirmado_Por': 'Confirmado_Por', 'Confirmado_Em': 'Confirmado_Em',
        'Item_Id': 'item_id', 'Edt_Vinculado': 'EDT_Vinculado',
    }
    return df.rename(columns=rename)

@st.cache_data(ttl=30)
def carregar_solicitacoes():
    conn = conectar_banco()
    try:
        df = pd.read_sql_query("SELECT * FROM solicitacoes_prazo ORDER BY id DESC", conn)
    finally:
        liberar_conexao(conn)
    return df

@st.cache_data(ttl=15)
def carregar_solicitacoes_op():
    conn = conectar_banco()
    try:
        df = pd.read_sql_query("SELECT * FROM solicitacoes_op ORDER BY criado_em DESC", conn)
    finally:
        liberar_conexao(conn)
    return df

def _limpar_cache_geral():
    for fn in [
        carregar_macro, carregar_micro, carregar_micro_completo,
        carregar_fila_logistica, carregar_solicitacoes, carregar_solicitacoes_op,
        carregar_macro_por_obra, carregar_micro_por_obra, carregar_projetos,
        carregar_medicao_obras, carregar_medicao_servicos,
        carregar_medicao_subdivisoes, carregar_medicao_historico,
    ]:
        try:
            fn.clear()
        except Exception:
            pass

def salvar_lotes_micro(lotes: list):
    """Insere lotes e reserva o m2 (ACM) ou kg (Esquadria-Vidro) na frente vinculada (Mudanca 2).
    Retorna (ok, mensagem) — ok=False quando um lote estouraria o saldo da frente; nesse
    caso nada e commitado (transacao inteira desfeita)."""
    if not lotes:
        return True, ""
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        for l in lotes:
            edt     = l.get('EDT_Vinculado')
            escopo  = l.get('Escopo')
            m2_item   = float(l.get('M2_Item') or 0)
            peso_item = float(l.get('Peso_Kg') or 0)
            if edt and edt != 'AVULSO':
                campo_total, campo_exec, valor_item, unidade = (
                    ('Peso_Total_Tarefa', 'peso_executado', peso_item, 'kg')
                    if escopo == 'Esquadria-Vidro' else
                    ('M2_Total_Tarefa', 'm2_executado', m2_item, 'm²')
                )
                if valor_item > 0:
                    cursor.execute(
                        f"SELECT {campo_total}, {campo_exec}, Detalhado FROM cronograma_macro WHERE EDT=%s",
                        (edt,)
                    )
                    frente_row = cursor.fetchone()
                    if frente_row:
                        total_frente, exec_frente, detalhado = frente_row
                        if detalhado and total_frente is not None:
                            saldo = float(total_frente) - float(exec_frente or 0)
                            if valor_item > saldo:
                                conn.rollback()
                                return False, (
                                    f"Lote {l.get('Cod_Lote','')} ({valor_item:.2f} {unidade}) estoura o saldo da frente "
                                    f"{edt} (saldo disponível: {saldo:.2f} {unidade}). Ajuste a quantidade ou edite a frente."
                                )
            cursor.execute("""
                INSERT INTO itens_detalhado
                (Obra_Vinculada, EDT_Vinculado, Cod_Lote, Num_OP, Tipo_Material,
                 Qtd_Caixas, M2_Item, Peso_Kg, Data_Producao_Programada, Data_Limite_Obra,
                 Data_Despacho, Romaneio_Chapas, Status_Item, Dificuldade, Fase_Produtiva, Enviado_Logistica,
                 Escopo, Numero_Projeto)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                l['Obra_Vinculada'], l['EDT_Vinculado'], l['Cod_Lote'], l['Num_OP'],
                l['Tipo_Material'], l['Qtd_Caixas'], l['M2_Item'], l.get('Peso_Kg', 0.0),
                l['Data_Producao_Programada'], l['Data_Limite_Obra'], l['Data_Despacho'],
                l['Romaneio_Chapas'], l['Status_Item'], l['Dificuldade'],
                l['Fase_Produtiva'], l['Enviado_Logistica'],
                l.get('Escopo'), l.get('Numero_Projeto', '')
            ))
            if edt and edt != 'AVULSO':
                if escopo == 'Esquadria-Vidro' and peso_item > 0:
                    cursor.execute(
                        "UPDATE cronograma_macro SET peso_executado = COALESCE(peso_executado,0) + %s WHERE EDT=%s",
                        (peso_item, edt)
                    )
                elif escopo != 'Esquadria-Vidro' and m2_item > 0:
                    cursor.execute(
                        "UPDATE cronograma_macro SET m2_executado = COALESCE(m2_executado,0) + %s WHERE EDT=%s",
                        (m2_item, edt)
                    )
        conn.commit()
        _limpar_cache_geral()
        return True, ""
    except Exception as e:
        conn.rollback()
        st.error(f"Erro ao salvar lotes: {e}")
        return False, str(e)
    finally:
        liberar_conexao(conn)

def deletar_lotes_por_edt_lote(obra, edt, cod_lote):
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        if edt and edt != 'AVULSO':
            cursor.execute(
                "SELECT COALESCE(SUM(M2_Item), 0), COALESCE(SUM(Peso_Kg), 0) FROM itens_detalhado "
                "WHERE Obra_Vinculada=%s AND EDT_Vinculado=%s AND Cod_Lote=%s",
                (obra, edt, cod_lote)
            )
            m2_liberado, peso_liberado = cursor.fetchone()
            m2_liberado   = float(m2_liberado or 0)
            peso_liberado = float(peso_liberado or 0)
            if m2_liberado > 0 or peso_liberado > 0:
                cursor.execute(
                    "UPDATE cronograma_macro SET "
                    "m2_executado = GREATEST(COALESCE(m2_executado,0) - %s, 0), "
                    "peso_executado = GREATEST(COALESCE(peso_executado,0) - %s, 0) "
                    "WHERE EDT=%s",
                    (m2_liberado, peso_liberado, edt)
                )
            cursor.execute(
                "DELETE FROM itens_detalhado WHERE Obra_Vinculada=%s AND EDT_Vinculado=%s AND Cod_Lote=%s",
                (obra, edt, cod_lote)
            )
        else:
            cursor.execute(
                "DELETE FROM itens_detalhado WHERE Obra_Vinculada=%s AND Cod_Lote=%s",
                (obra, cod_lote)
            )
        conn.commit()
        _limpar_cache_geral()
    except Exception as e:
        conn.rollback()
        st.error(f"Erro ao deletar lote: {e}")
    finally:
        liberar_conexao(conn)

def atualizar_cronograma_macro_datas(edt, prazo_eng, primeiro_prod, despacho):
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT prazo_engenharia, primeiro_dia_producao, data_limite_despacho FROM cronograma_macro WHERE EDT=%s",
            (edt,)
        )
        row = cursor.fetchone()
        def to_dt(v):
            try:
                return pd.to_datetime(v) if v else None
            except Exception:
                return None
        if row:
            pa, pp, pd_ = to_dt(row[0]), to_dt(row[1]), to_dt(row[2])
            novo_prazo    = min(pa,  pd.to_datetime(prazo_eng))     if pa  else pd.to_datetime(prazo_eng)
            novo_primeiro = min(pp,  pd.to_datetime(primeiro_prod)) if pp  else pd.to_datetime(primeiro_prod)
            novo_despacho = min(pd_, pd.to_datetime(despacho))      if pd_ else pd.to_datetime(despacho)
        else:
            novo_prazo    = pd.to_datetime(prazo_eng)
            novo_primeiro = pd.to_datetime(primeiro_prod)
            novo_despacho = pd.to_datetime(despacho)
        cursor.execute("""
            UPDATE cronograma_macro
            SET Prazo_Engenharia=%s, Primeiro_Dia_Producao=%s, Data_Limite_Despacho=%s
            WHERE EDT=%s
        """, (
            novo_prazo.strftime('%Y-%m-%d'),
            novo_primeiro.strftime('%Y-%m-%d'),
            novo_despacho.strftime('%Y-%m-%d'),
            edt
        ))
        conn.commit()
        _limpar_cache_geral()
    except Exception as e:
        conn.rollback()
        st.error(f"Erro ao atualizar datas: {e}")
    finally:
        liberar_conexao(conn)

def atualizar_status_engenharia(edt_id, novo_status):
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute("UPDATE cronograma_macro SET Status_Engenharia=%s WHERE id=%s", (novo_status, edt_id))
        conn.commit()
        _limpar_cache_geral()
    except Exception as e:
        conn.rollback()
        st.error(f"Erro ao atualizar status: {e}")
    finally:
        liberar_conexao(conn)

def salvar_solicitacao(edt, tarefa, prazo_atual, prazo_sol, justif, criado_por):
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO solicitacoes_prazo (edt, tarefa, prazo_atual, prazo_solicitado, justificativa, criado_por, status, criado_em)
            VALUES (%s,%s,%s,%s,%s,%s,'Pendente de Aprovacao',%s)
        """, (edt, tarefa, prazo_atual, prazo_sol, justif, criado_por, datetime.now(FUSO_BR).strftime('%d/%m/%Y %H:%M')))
        conn.commit()
        carregar_solicitacoes.clear()
    except Exception as e:
        conn.rollback()
        st.error(f"Erro ao salvar solicitação: {e}")
    finally:
        liberar_conexao(conn)

def atualizar_status_solicitacao(sol_id, novo_status):
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute("UPDATE solicitacoes_prazo SET status=%s WHERE id=%s", (novo_status, sol_id))
        # Se aprovado, aplica o prazo solicitado no cronograma macro
        if novo_status == "Aprovado":
            cursor.execute(
                "SELECT edt, prazo_solicitado FROM solicitacoes_prazo WHERE id=%s", (sol_id,)
            )
            row = cursor.fetchone()
            if row:
                edt_alvo, prazo_novo = row
                cursor.execute(
                    "UPDATE cronograma_macro SET Prazo_Engenharia=%s WHERE EDT=%s",
                    (prazo_novo, edt_alvo)
                )
        conn.commit()
        carregar_solicitacoes.clear()
        try:
            carregar_macro.clear()
            carregar_macro_por_obra.clear()
        except Exception:
            pass
    except Exception as e:
        conn.rollback()
        st.error(f"Erro ao atualizar solicitação: {e}")
    finally:
        liberar_conexao(conn)

def enviar_para_logistica(row, limite_despacho):
    if bool(row.get('Uso_Interno', False)):
        # Uso Interno nunca precisa de romaneio de despacho — a OP se resolve sozinha,
        # sem entrar na fila de logistica.
        return
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM logistica_envios WHERE item_id=%s", (int(row['id']),))
        if cursor.fetchone():
            return
        cursor.execute("""
            INSERT INTO logistica_envios
            (item_id, Obra_Vinculada, EDT_Vinculado, Cod_Lote, Num_OP, Tipo_Material,
             Qtd_Caixas, M2_Item, Romaneio_Chapas, Data_Limite_Despacho, Status_Logistica)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'Aguardando Agendamento')
        """, (
            int(row['id']), row['Obra_Vinculada'], row['EDT_Vinculado'], row['Cod_Lote'],
            row.get('Num_OP') or '', row['Tipo_Material'], int(row['Qtd_Caixas']), float(row['M2_Item']),
            row['Romaneio_Chapas'],
            limite_despacho.strftime('%Y-%m-%d') if prazo_valido(limite_despacho) else None
        ))
        cursor.execute("UPDATE itens_detalhado SET Enviado_Logistica=1 WHERE id=%s", (int(row['id']),))
        conn.commit()
        _limpar_cache_geral()
    except Exception as e:
        conn.rollback()
        st.error(f"Erro ao enviar para logística: {e}")
    finally:
        liberar_conexao(conn)

def agendar_envio(log_id, data_envio, transportadora, veiculo, obs, usuario):
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE logistica_envios
            SET Data_Envio_Agendado=%s, Transportadora=%s, Veiculo=%s, Observacoes=%s, Status_Logistica='Envio Agendado'
            WHERE id=%s
        """, (
            data_envio.strftime('%Y-%m-%d') if data_envio else None,
            transportadora, veiculo, obs, log_id
        ))
        conn.commit()
        carregar_fila_logistica.clear()
    except Exception as e:
        conn.rollback()
        st.error(f"Erro ao agendar envio: {e}")
    finally:
        liberar_conexao(conn)

def confirmar_despacho(log_id, usuario):
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE logistica_envios
            SET Status_Logistica='Despachado', Confirmado_Por=%s, Confirmado_Em=%s
            WHERE id=%s
        """, (usuario, datetime.now(FUSO_BR).strftime('%d/%m/%Y %H:%M'), log_id))
        conn.commit()
        carregar_fila_logistica.clear()
    except Exception as e:
        conn.rollback()
        st.error(f"Erro ao confirmar despacho: {e}")
    finally:
        liberar_conexao(conn)

def verificar_login(usuario, senha):
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT nome, setor, senha FROM usuarios WHERE usuario=%s", (usuario,))
        resultado = cursor.fetchone()
        if resultado and verificar_senha(senha, resultado[2]):
            return resultado[0], resultado[1]
        return None
    except Exception:
        return None
    finally:
        liberar_conexao(conn)

def trocar_propria_senha(usuario: str, senha_atual: str, senha_nova: str) -> tuple[bool, str]:
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT senha FROM usuarios WHERE usuario=%s", (usuario,))
        resultado = cursor.fetchone()
        if not resultado or not verificar_senha(senha_atual, resultado[0]):
            return False, "Senha atual incorreta."
        cursor.execute(
            "UPDATE usuarios SET senha=%s WHERE usuario=%s",
            (hash_senha(senha_nova), usuario)
        )
        conn.commit()
        return True, "Senha alterada com sucesso!"
    except Exception as e:
        conn.rollback()
        return False, f"Erro ao trocar senha: {e}"
    finally:
        liberar_conexao(conn)

def redefinir_senha_usuario(usuario_alvo: str, senha_nova: str) -> tuple[bool, str]:
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE usuarios SET senha=%s WHERE usuario=%s",
            (hash_senha(senha_nova), usuario_alvo)
        )
        if cursor.rowcount == 0:
            conn.rollback()
            return False, "Usuário não encontrado."
        conn.commit()
        return True, "Senha redefinida com sucesso!"
    except Exception as e:
        conn.rollback()
        return False, f"Erro ao redefinir senha: {e}"
    finally:
        liberar_conexao(conn)

def resetar_banco_dados_completo(usuario=None):
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        for tabela in ['arquivos_op', 'arquivos_solicitacao_op', 'solicitacoes_op', 'op_pecas', 'componentes_op',
                       'cronograma_macro', 'itens_detalhado', 'solicitacoes_prazo', 'logistica_envios',
                       'medicao_historico', 'medicao_subdivisoes', 'medicao_servicos', 'medicao_obras', 'auditoria_log']:
            cursor.execute(f"DELETE FROM {tabela}")
        conn.commit()
        _limpar_cache_geral()
        carregar_medicao_obras.clear()
        carregar_componentes_op.clear()
        carregar_todas_ops_com_componentes.clear()
        carregar_pecas_lote.clear()
        carregar_todas_pecas_obra.clear()
        return True
    except Exception as e:
        conn.rollback()
        st.error(f"Erro no reset: {e}")
        return False
    finally:
        liberar_conexao(conn)

def salvar_arquivo_op(item_id: int, nome: str, tipo: str, conteudo: bytes, usuario: str):
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO arquivos_op (item_id, nome_arquivo, tipo_arquivo, conteudo, enviado_por) VALUES (%s,%s,%s,%s,%s)",
            (item_id, nome, tipo, conteudo, usuario)
        )
        conn.commit()
        return True
    except Exception as e:
        conn.rollback()
        st.error(f"Erro ao salvar arquivo: {e}")
        return False
    finally:
        liberar_conexao(conn)

def carregar_arquivos_op(item_id: int):
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, nome_arquivo, tipo_arquivo, enviado_por, enviado_em FROM arquivos_op WHERE item_id=%s ORDER BY enviado_em DESC",
            (item_id,)
        )
        rows = cursor.fetchall()
        return rows  # lista de (id, nome, tipo, enviado_por, enviado_em)
    except Exception:
        return []
    finally:
        liberar_conexao(conn)

def carregar_conteudo_arquivo(arquivo_id: int):
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT nome_arquivo, tipo_arquivo, conteudo FROM arquivos_op WHERE id=%s", (arquivo_id,))
        row = cursor.fetchone()
        return row  # (nome, tipo, bytes)
    except Exception:
        return None
    finally:
        liberar_conexao(conn)

def deletar_arquivo_op(arquivo_id: int):
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM arquivos_op WHERE id=%s", (arquivo_id,))
        conn.commit()
        return True
    except Exception as e:
        conn.rollback()
        st.error(f"Erro ao deletar arquivo: {e}")
        return False
    finally:
        liberar_conexao(conn)

def salvar_romaneio_manual(obra: str, data_recebimento, itens: list, usuario: str):
    """itens: lista de dicts {'descricao':..., 'quantidade':...}"""
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO romaneios_manuais (obra_vinculada, data_recebimento, criado_por) "
            "VALUES (%s,%s,%s) RETURNING id",
            (obra, data_recebimento, usuario)
        )
        romaneio_id = cursor.fetchone()[0]
        for item in itens:
            cursor.execute(
                "INSERT INTO romaneios_manuais_itens (romaneio_id, descricao, quantidade) VALUES (%s,%s,%s)",
                (romaneio_id, item['descricao'], item['quantidade'])
            )
        conn.commit()
        return romaneio_id
    except Exception as e:
        conn.rollback()
        st.error(f"Erro ao salvar romaneio manual: {e}")
        return None
    finally:
        liberar_conexao(conn)

def carregar_romaneios_manuais():
    conn = conectar_banco()
    try:
        return pd.read_sql_query(
            "SELECT r.id, r.obra_vinculada, r.data_recebimento, r.criado_por, r.criado_em, "
            "COUNT(i.id) AS qtd_itens "
            "FROM romaneios_manuais r LEFT JOIN romaneios_manuais_itens i ON i.romaneio_id = r.id "
            "GROUP BY r.id ORDER BY r.data_recebimento DESC, r.id DESC",
            conn
        )
    except Exception:
        return pd.DataFrame()
    finally:
        liberar_conexao(conn)

def carregar_itens_romaneio_manual(romaneio_id: int):
    conn = conectar_banco()
    try:
        return pd.read_sql_query(
            "SELECT id, descricao, quantidade FROM romaneios_manuais_itens WHERE romaneio_id=%s ORDER BY id",
            conn, params=(romaneio_id,)
        )
    except Exception:
        return pd.DataFrame()
    finally:
        liberar_conexao(conn)

def excluir_romaneio_manual(romaneio_id: int):
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM romaneios_manuais WHERE id=%s", (romaneio_id,))
        conn.commit()
        return True
    except Exception as e:
        conn.rollback()
        st.error(f"Erro ao excluir romaneio manual: {e}")
        return False
    finally:
        liberar_conexao(conn)

def gerar_romaneio_manual_xlsx(obra: str, data_recebimento, itens_df, criado_por: str) -> bytes:
    """Gera o romaneio manual .xlsx com a lista de itens (sem OP vinculada)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Romaneio"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "portrait"

    bd = Side(style='thin', color="000000")
    borda = Border(left=bd, right=bd, top=bd, bottom=bd)
    fill_cab = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    fill_sub = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16

    ws.merge_cells("A1:D1")
    ws["A1"] = "PASSOLD SISTEMAS DE FACHADAS"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = fill_cab
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:D2")
    ws["A2"] = "ROMANEIO MANUAL (SEM OP VINCULADA)"
    ws["A2"].font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    ws["A2"].fill = fill_cab
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    infos = [
        ("Obra:", str(obra)),
        ("Data de recebimento:", pd.to_datetime(data_recebimento).strftime('%d/%m/%Y')),
        ("Cadastrado por:", str(criado_por)),
    ]
    linha = 3
    for label, valor in infos:
        ws.cell(linha, 1, label).font = Font(name="Arial", size=11, bold=True)
        ws.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=4)
        ws.cell(linha, 2, valor).font = Font(name="Arial", size=11)
        for c in range(1, 5):
            ws.cell(linha, c).border = borda
        linha += 1

    linha += 1
    titulos = ["#", "DESCRIÇÃO DO MATERIAL", "QUANTIDADE", "CONFERIDO"]
    for col, titulo in enumerate(titulos, 1):
        cel = ws.cell(linha, col, titulo)
        cel.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        cel.fill = fill_cab
        cel.alignment = Alignment(horizontal="center", vertical="center")
        cel.border = borda
    linha += 1

    for i, (_, item) in enumerate(itens_df.iterrows(), 1):
        dados = [i, item.get('descricao', ''), item.get('quantidade', 0), ""]
        for col, val in enumerate(dados, 1):
            cel = ws.cell(linha, col, val)
            cel.font = Font(name="Arial", size=11)
            cel.alignment = Alignment(horizontal="center", vertical="center")
            cel.border = borda
            if i % 2 == 0:
                cel.fill = fill_sub
        linha += 1

    linha += 1
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=2)
    ws.cell(linha, 1, "TOTAL DE ITENS:").font = Font(name="Arial", size=11, bold=True)
    ws.cell(linha, 3, len(itens_df)).font = Font(name="Arial", size=11, bold=True)
    for c in range(1, 5):
        ws.cell(linha, c).border = borda

    linha += 3
    assinaturas = ["Recebedor na Obra / Data", "Conferência Almoxarifado"]
    for ass in assinaturas:
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=2)
        ws.cell(linha, 1).border = Border(bottom=bd)
        ws.cell(linha + 1, 1, ass).font = Font(name="Arial", size=10, bold=True)
        ws.cell(linha, 3, "____/____/______").font = Font(name="Arial", size=11)
        linha += 3

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def salvar_solicitacao_op(obra: str, numero_projeto: str, tipo_material: str, descricao: str,
                           quantidade: float, observacao: str, usuario: str):
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute(
            """INSERT INTO solicitacoes_op
               (obra, numero_projeto, tipo_material, descricao, quantidade, observacao, solicitado_por)
               VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
            (obra, numero_projeto, tipo_material, descricao, quantidade, observacao, usuario)
        )
        novo_id = cursor.fetchone()[0]
        conn.commit()
        return novo_id
    except Exception as e:
        conn.rollback()
        st.error(f"Erro ao salvar solicitação: {e}")
        return None
    finally:
        liberar_conexao(conn)

def salvar_arquivo_solicitacao(solicitacao_id: int, nome: str, tipo: str, conteudo: bytes, usuario: str):
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO arquivos_solicitacao_op (solicitacao_id, nome_arquivo, tipo_arquivo, conteudo, enviado_por) VALUES (%s,%s,%s,%s,%s)",
            (solicitacao_id, nome, tipo, conteudo, usuario)
        )
        conn.commit()
        return True
    except Exception as e:
        conn.rollback()
        st.error(f"Erro ao salvar arquivo: {e}")
        return False
    finally:
        liberar_conexao(conn)

def carregar_arquivos_solicitacao(solicitacao_id: int):
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, nome_arquivo, tipo_arquivo, enviado_por, enviado_em FROM arquivos_solicitacao_op WHERE solicitacao_id=%s ORDER BY enviado_em DESC",
            (solicitacao_id,)
        )
        return cursor.fetchall()
    except Exception:
        return []
    finally:
        liberar_conexao(conn)

def carregar_conteudo_arquivo_solicitacao(arquivo_id: int):
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT nome_arquivo, tipo_arquivo, conteudo FROM arquivos_solicitacao_op WHERE id=%s", (arquivo_id,))
        return cursor.fetchone()
    except Exception:
        return None
    finally:
        liberar_conexao(conn)

def confirmar_vinculacao_solicitacao(solicitacao_id: int, item_id: int, usuario: str):
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute(
            """UPDATE solicitacoes_op
               SET status='Vinculada', item_id=%s, vinculado_por=%s, vinculado_em=NOW()
               WHERE id=%s""",
            (item_id, usuario, solicitacao_id)
        )
        conn.commit()
        return True
    except Exception as e:
        conn.rollback()
        st.error(f"Erro ao confirmar vinculação: {e}")
        return False
    finally:
        liberar_conexao(conn)

def salvar_parada_op(item_id, em_parada: bool, motivo: str, usuario: str):
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE itens_detalhado SET em_parada=%s, motivo_parada=%s WHERE id=%s",
            (em_parada, motivo if em_parada else None, item_id)
        )
        conn.commit()
        carregar_micro.clear()
        return True
    except Exception as e:
        conn.rollback()
        st.error(f"Erro ao registrar parada: {e}")
        return False
    finally:
        liberar_conexao(conn)

def _parse_qtd_unidade(parte: str):
    """Extrai quantidade e unidade de strings como '1pç', '2 pç', '3un'."""
    import re
    parte = parte.strip()
    m = re.match(r'^(\d+(?:[.,]\d+)?)\s*([a-zA-Zç²]+)', parte)
    if m:
        qtd_str = m.group(1).replace(',', '.')
        und = m.group(2).lower().strip()
        return float(qtd_str), und
    return 1.0, "pç"

def parse_romaneio(texto: str) -> list:
    """Converte texto de romaneio colado em lista de peças para op_pecas.

    Formato esperado por linha:
      REF - COD_PERFIL - DESCRIÇÃO - MEDIDA - ÂNGULOS - QTD
    ou variações com menos campos (mínimo 3 partes separadas por ' - ').

    Retorna lista de dicts com: codigo, descricao, medida, localizacao, qtd.
    """
    pecas = []
    for linha in texto.splitlines():
        linha = linha.strip()
        if not linha:
            continue
        partes = [p.strip() for p in linha.split(" - ")]
        if len(partes) < 3:
            continue
        qtd_raw = partes[-1]
        qtd, _ = _parse_qtd_unidade(qtd_raw)
        codigo = partes[0]
        if len(partes) >= 6:
            # REF - COD - DESC - MEDIDA - ANGULOS - QTD
            descricao = f"{partes[1]} - {partes[2]}"
            medida    = f"{partes[3]} {partes[4]}"
        elif len(partes) == 5:
            # REF - COD - DESC - MEDIDA - QTD
            descricao = f"{partes[1]} - {partes[2]}"
            medida    = partes[3]
        elif len(partes) == 4:
            # REF - DESC - MEDIDA - QTD
            descricao = partes[1]
            medida    = partes[2]
        else:
            # REF - DESC - QTD
            descricao = partes[1]
            medida    = ""
        pecas.append({
            "codigo":     codigo,
            "descricao":  descricao,
            "medida":     medida,
            "localizacao": "",
            "qtd":        int(qtd),
        })
    return pecas

def salvar_componentes(item_id, obra, cod_lote, num_op, componentes: list):
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM componentes_op WHERE item_id=%s", (item_id,))
        for c in componentes:
            cursor.execute("""
                INSERT INTO componentes_op
                (item_id, Obra_Vinculada, Cod_Lote, Num_OP, Nome_Componente, Quantidade, Unidade,
                 Status_Item, Observacao)
                VALUES (%s,%s,%s,%s,%s,%s,%s,'Aguardando Conferencia',NULL)
            """, (item_id, obra, cod_lote, num_op, c['nome'], c['qtd'], c['unidade']))
        conn.commit()
    except Exception as e:
        conn.rollback()
        st.error(f"Erro ao salvar componentes: {e}")
    finally:
        liberar_conexao(conn)

@st.cache_data(ttl=30)
def carregar_componentes_op(item_id):
    conn = conectar_banco()
    try:
        df = pd.read_sql_query(
            "SELECT * FROM componentes_op WHERE item_id=%s ORDER BY id", conn, params=(item_id,)
        )
    finally:
        liberar_conexao(conn)
    return df

@st.cache_data(ttl=30)
def carregar_todas_ops_com_componentes():
    conn = conectar_banco()
    try:
        df = pd.read_sql_query("""
            SELECT DISTINCT item_id, Obra_Vinculada, Cod_Lote, Num_OP
            FROM componentes_op ORDER BY Obra_Vinculada, Cod_Lote
        """, conn)
    finally:
        liberar_conexao(conn)
    return df

def atualizar_componente(comp_id, status, obs, usuario):
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE componentes_op
            SET Status_Item=%s, Observacao=%s, Conferido_Por=%s, Conferido_Em=%s
            WHERE id=%s
        """, (status, obs, usuario, datetime.now(FUSO_BR).strftime('%d/%m/%Y %H:%M'), comp_id))
        conn.commit()
        carregar_componentes_op.clear()
        carregar_todas_ops_com_componentes.clear()
    except Exception as e:
        conn.rollback()
        st.error(f"Erro ao atualizar componente: {e}")
    finally:
        liberar_conexao(conn)

# ========================================================
# FUNÇÕES DE OP_PECAS
# ========================================================
@st.cache_data(ttl=30)
def carregar_pecas_lote(lote_id: int):
    conn = conectar_banco()
    try:
        df = pd.read_sql_query(
            "SELECT * FROM op_pecas WHERE lote_id=%s ORDER BY id", conn, params=(lote_id,)
        )
    finally:
        liberar_conexao(conn)
    return df

@st.cache_data(ttl=30)
def carregar_todas_pecas_obra(obra: str):
    conn = conectar_banco()
    try:
        df = pd.read_sql_query(
            "SELECT * FROM op_pecas WHERE obra=%s ORDER BY cod_lote, id", conn, params=(obra,)
        )
    finally:
        liberar_conexao(conn)
    return df

def salvar_pecas_lote(lote_id: int, obra: str, cod_lote: str, num_op: str,
                      pecas: list, componentes_status: str, m2_op_real: float, peso_op_real: float = 0.0):
    """Salva peças com lock otimista via updated_at. Reconcilia tanto m2_executado (ACM) quanto
    peso_executado (Esquadria-Vidro) — cada lote so contribui de verdade pra um dos dois, o outro
    fica em zero e a conta correspondente vira um no-op."""
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        # Lock otimista: bloqueia a linha antes de escrever
        cursor.execute("SELECT updated_at FROM itens_detalhado WHERE id=%s FOR UPDATE", (lote_id,))
        if not cursor.fetchone():
            st.error("Lote não encontrado. Recarregue a página.")
            conn.rollback()
            return
        cursor.execute("SELECT COALESCE(m2_op_real, 0), COALESCE(peso_op_real, 0) FROM op_pecas WHERE lote_id=%s LIMIT 1", (lote_id,))
        row_antigo = cursor.fetchone()
        if row_antigo:
            m2_antigo, peso_antigo = float(row_antigo[0]), float(row_antigo[1])
        else:
            # Primeiro lancamento de pecas deste lote: o que ja esta reservado em m2_executado/
            # peso_executado e o M2_Item/Peso_Kg planejado do lote (reservado na criacao — Mudanca 2),
            # nao zero.
            cursor.execute("SELECT COALESCE(M2_Item,0), COALESCE(Peso_Kg,0) FROM itens_detalhado WHERE id=%s", (lote_id,))
            row_item = cursor.fetchone()
            m2_antigo, peso_antigo = (float(row_item[0]), float(row_item[1])) if row_item else (0.0, 0.0)
        cursor.execute("DELETE FROM op_pecas WHERE lote_id=%s", (lote_id,))
        for p in pecas:
            qtd = int(p.get('qtd', 0))
            cursor.execute("""
                INSERT INTO op_pecas
                (lote_id, obra, cod_lote, num_op, codigo, descricao, localizacao, medida,
                 qtd_total, qtd_enviada, saldo, m2_op_real, peso_op_real, componentes_status)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,0,%s,%s,%s,%s)
            """, (
                lote_id, obra, cod_lote, num_op,
                p.get('codigo', '').strip().upper(),
                p.get('descricao', '').strip(),
                p.get('localizacao', '').strip().upper(),
                p.get('medida', '').strip().upper(),
                qtd, qtd, float(m2_op_real), float(peso_op_real), componentes_status
            ))
        # Atualiza timestamp — invalida edições concorrentes
        cursor.execute("UPDATE itens_detalhado SET updated_at=NOW() WHERE id=%s", (lote_id,))
        cursor.execute("SELECT EDT_Vinculado FROM itens_detalhado WHERE id=%s", (lote_id,))
        row_lote = cursor.fetchone()
        if row_lote:
            cursor.execute(
                "UPDATE cronograma_macro SET "
                "m2_executado = GREATEST(COALESCE(m2_executado, 0) - %s + %s, 0), "
                "peso_executado = GREATEST(COALESCE(peso_executado, 0) - %s + %s, 0) "
                "WHERE EDT = %s",
                (m2_antigo, float(m2_op_real), peso_antigo, float(peso_op_real), row_lote[0])
            )
        conn.commit()
        carregar_pecas_lote.clear()
        carregar_todas_pecas_obra.clear()
        _limpar_cache_geral()
    except Exception as e:
        conn.rollback()
        st.error(f"Erro ao salvar peças: {e}")
    finally:
        liberar_conexao(conn)

def atualizar_componentes_status_pecas(lote_id: int, novo_status: str):
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE op_pecas SET componentes_status=%s WHERE lote_id=%s",
            (novo_status, lote_id)
        )
        conn.commit()
        carregar_pecas_lote.clear()
    except Exception as e:
        conn.rollback()
        st.error(f"Erro ao atualizar status: {e}")
    finally:
        liberar_conexao(conn)

def gerar_op_xlsx(lote_row, pecas_df, macro_row, campos_extras: dict) -> bytes:
    """Gera o documento de Ordem de Produção em xlsx."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Ordem de Producao"
    ws.sheet_view.showGridLines = False

    bd      = Side(style='thin', color="000000")
    borda   = Border(left=bd, right=bd, top=bd, bottom=bd)
    bd_bot  = Border(bottom=bd)
    fill_az = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    fill_lz = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    fill_gz = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10

    tipo_escopo = str(macro_row.get('Tipo_Escopo', 'ACM') or 'ACM')
    num_projeto  = str(macro_row.get('Numero_Projeto', '') or '')

    # ── CABEÇALHO ──────────────────────────────────────────
    ws.merge_cells("A1:F1")
    ws["A1"] = "PASSOLD SISTEMAS DE FACHADAS LTDA"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = fill_az
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")
    ws["A2"] = f"ORDEM DE PRODUÇÃO — {tipo_escopo.upper()}"
    ws["A2"].font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    ws["A2"].fill = fill_az
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # ── INFO PRINCIPAL ─────────────────────────────────────
    def info_row(ws, linha, label, valor, fill=None):
        ws.cell(linha, 1, label).font = Font(name="Arial", size=11, bold=True)
        ws.cell(linha, 1).fill = fill_gz
        ws.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=6)
        ws.cell(linha, 2, str(valor).upper()).font = Font(name="Arial", size=11)
        for c in range(1, 7):
            ws.cell(linha, c).border = borda
            ws.cell(linha, c).alignment = Alignment(vertical="center")
        ws.row_dimensions[linha].height = 18

    linha = 3
    info_row(ws, linha,   "Nº OP:",          lote_row.get('Num_OP', '—'))
    info_row(ws, linha+1, "DATA:",            datetime.now(FUSO_BR).strftime('%d/%m/%Y'))
    info_row(ws, linha+2, "OBRA:",            lote_row.get('Obra_Vinculada', '—'))
    info_row(ws, linha+3, "PROJETO:",         f"{num_projeto} — {macro_row.get('Tarefa', '—')}" if num_projeto else macro_row.get('Tarefa', '—'))
    info_row(ws, linha+4, "LOTE:",            lote_row.get('Cod_Lote', '—'))
    info_row(ws, linha+5, "ETAPA/PAVIMENTOS:",lote_row.get('Romaneio_Chapas', '—'))
    info_row(ws, linha+6, "MATERIAL:",        campos_extras.get('material', lote_row.get('Tipo_Material', '—')))

    linha = 10

    # ── CAMPOS ESPECÍFICOS POR TIPO ────────────────────────
    if tipo_escopo.upper() == "ACM":
        info_row(ws, linha,   "ÁREA TOTAL (m²):", f"{campos_extras.get('area_total', lote_row.get('M2_Item', 0)):.2f} m²")
        info_row(ws, linha+1, "DIFICULDADE:",     campos_extras.get('dificuldade', lote_row.get('Dificuldade', '—')))
        info_row(ws, linha+2, "QTD CHAPAS:",      lote_row.get('Qtd_Caixas', '—'))
        info_row(ws, linha+3, "QTD FOLHAS PROJ:", campos_extras.get('qtd_folhas', '—'))
        linha += 4
    elif "ESQUADRIA" in tipo_escopo.upper() or "VIDRO" in tipo_escopo.upper():
        info_row(ws, linha,   "PESO TOTAL:",      campos_extras.get('peso_total', '—'))
        info_row(ws, linha+1, "QTD FOLHAS PROJ:", campos_extras.get('qtd_folhas', '—'))
        linha += 2
    else:  # Terceirizada
        info_row(ws, linha,   "EMPRESA RESP.:",   campos_extras.get('empresa', '—'))
        linha += 1

    linha += 1

    # ── OBSERVAÇÕES ────────────────────────────────────────
    ws.merge_cells(f"A{linha}:F{linha}")
    ws[f"A{linha}"] = "LIBERAÇÃO PARA PRODUÇÃO"
    ws[f"A{linha}"].font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    ws[f"A{linha}"].fill = fill_az
    ws[f"A{linha}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[linha].height = 20
    linha += 1

    obs_txt = campos_extras.get('observacoes', '')
    ws.merge_cells(f"A{linha}:F{linha+2}")
    ws[f"A{linha}"] = obs_txt
    ws[f"A{linha}"].font = Font(name="Arial", size=11)
    ws[f"A{linha}"].alignment = Alignment(wrap_text=True, vertical="top")
    for r in range(linha, linha+3):
        for c in range(1, 7):
            ws.cell(r, c).border = borda
    ws.row_dimensions[linha].height = 45
    linha += 3

    linha += 1

    # ── TABELA DE PEÇAS ────────────────────────────────────
    if not pecas_df.empty:
        ws.merge_cells(f"A{linha}:F{linha}")
        ws[f"A{linha}"] = "RELAÇÃO DE PEÇAS"
        ws[f"A{linha}"].font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        ws[f"A{linha}"].fill = fill_az
        ws[f"A{linha}"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[linha].height = 20
        linha += 1

        titulos_p = ["#", "CÓDIGO", "DESCRIÇÃO / LOCAL", "MEDIDA", "QTD", "SALDO"]
        for ci, t in enumerate(titulos_p, 1):
            cel = ws.cell(linha, ci, t)
            cel.font = Font(name="Arial", size=10, bold=True)
            cel.fill = fill_gz
            cel.alignment = Alignment(horizontal="center")
            cel.border = borda
        linha += 1

        for i, (_, peca) in enumerate(pecas_df.iterrows(), 1):
            bg = fill_lz if i % 2 == 0 else PatternFill()
            vals = [i, peca.get('codigo',''), peca.get('localizacao',''), peca.get('medida',''),
                    peca.get('qtd_total',0), peca.get('saldo',0)]
            for ci, v in enumerate(vals, 1):
                cel = ws.cell(linha, ci, v)
                cel.font = Font(name="Arial", size=10)
                cel.alignment = Alignment(horizontal="center", vertical="center")
                cel.border = borda
                if i % 2 == 0:
                    cel.fill = fill_lz
            ws.row_dimensions[linha].height = 16
            linha += 1

        # Total
        ws.merge_cells(f"A{linha}:D{linha}")
        ws.cell(linha, 1, "TOTAL DE PEÇAS:").font = Font(name="Arial", size=10, bold=True)
        ws.cell(linha, 5, int(pecas_df['qtd_total'].sum())).font = Font(name="Arial", size=10, bold=True)
        for c in range(1, 7):
            ws.cell(linha, c).border = borda
        linha += 2

    # ── ASSINATURAS ────────────────────────────────────────
    assinaturas = ["Responsável PCP", "Responsável Produção", "Conferência"]
    for ass in assinaturas:
        ws.merge_cells(f"A{linha}:D{linha}")
        ws.cell(linha, 1).border = bd_bot
        ws.cell(linha+1, 1, ass).font = Font(name="Arial", size=10, bold=True)
        ws.cell(linha, 5, "____/____/______").font = Font(name="Arial", size=10)
        linha += 3

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def gerar_romaneio_xlsx(lote_row, pecas_df, endereco_obra: str, digitado_por: str) -> bytes:
    """Gera o romaneio .xlsx com as peças do lote."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Romaneio"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "portrait"

    bd = Side(style='thin', color="000000")
    borda = Border(left=bd, right=bd, top=bd, bottom=bd)
    fill_cab = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    fill_sub = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14

    # Cabeçalho
    ws.merge_cells("A1:E1")
    ws["A1"] = "PASSOLD SISTEMAS DE FACHADAS"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = fill_cab
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:E2")
    ws["A2"] = "ROMANEIO DE DESPACHO"
    ws["A2"].font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    ws["A2"].fill = fill_cab
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    # Informações do lote
    infos = [
        ("OP:", str(lote_row.get('Num_OP', '—'))),
        ("Obra:", str(lote_row.get('Obra_Vinculada', '—'))),
        ("Lote:", str(lote_row.get('Cod_Lote', '—'))),
        ("Endereço:", str(endereco_obra)),
        ("Digitado por:", str(digitado_por)),
        ("Data:", datetime.now(FUSO_BR).strftime('%d/%m/%Y')),
    ]
    linha = 3
    for label, valor in infos:
        ws.cell(linha, 1, label).font = Font(name="Arial", size=11, bold=True)
        ws.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=5)
        ws.cell(linha, 2, valor).font = Font(name="Arial", size=11)
        for c in range(1, 6):
            ws.cell(linha, c).border = borda
        linha += 1

    linha += 1

    # Cabeçalho da tabela de peças
    titulos = ["CÓDIGO", "LOCALIZAÇÃO", "MEDIDA", "QTD TOTAL", "QTD ENVIADA"]
    for col, titulo in enumerate(titulos, 1):
        cel = ws.cell(linha, col, titulo)
        cel.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        cel.fill = fill_cab
        cel.alignment = Alignment(horizontal="center", vertical="center")
        cel.border = borda
    linha += 1

    # Peças
    for i, (_, peca) in enumerate(pecas_df.iterrows()):
        bg = fill_sub if i % 2 == 0 else PatternFill()
        dados = [
            peca.get('codigo', ''),
            peca.get('localizacao', ''),
            peca.get('medida', ''),
            peca.get('qtd_total', 0),
            peca.get('qtd_enviada', 0),
        ]
        for col, val in enumerate(dados, 1):
            cel = ws.cell(linha, col, val)
            cel.font = Font(name="Arial", size=11)
            cel.alignment = Alignment(horizontal="center", vertical="center")
            cel.border = borda
            if i % 2 == 0:
                cel.fill = fill_sub
        linha += 1

    # Totais
    linha += 1
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=3)
    ws.cell(linha, 1, "TOTAL DE PEÇAS:").font = Font(name="Arial", size=11, bold=True)
    ws.cell(linha, 4, int(pecas_df['qtd_total'].sum()) if not pecas_df.empty else 0).font = Font(name="Arial", size=11, bold=True)
    for c in range(1, 6):
        ws.cell(linha, c).border = borda

    # Assinaturas
    linha += 3
    assinaturas = [
        "Conferência Interna",
        "Nome Motorista / Data",
        "Recebedor na Obra / Data",
        "Engenheiro Responsável",
    ]
    for ass in assinaturas:
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=3)
        ws.cell(linha, 1).border = Border(bottom=bd)
        ws.cell(linha, 1, "").font = Font(name="Arial", size=11)
        ws.cell(linha+1, 1, ass).font = Font(name="Arial", size=10, bold=True)
        ws.cell(linha, 4, "____/____/______").font = Font(name="Arial", size=11)
        linha += 3

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ========================================================
# FUNÇÕES DO SISTEMA DE MEDIÇÃO — agora no banco
# ========================================================
@st.cache_data(ttl=30)
def carregar_medicao_obras():
    conn = conectar_banco()
    try:
        df = pd.read_sql_query("SELECT * FROM medicao_obras ORDER BY id", conn)
    finally:
        liberar_conexao(conn)
    return df

@st.cache_data(ttl=30)
def carregar_medicao_servicos(obra_id: int):
    conn = conectar_banco()
    try:
        df = pd.read_sql_query(
            "SELECT * FROM medicao_servicos WHERE obra_id=%s ORDER BY id", conn, params=(obra_id,)
        )
    finally:
        liberar_conexao(conn)
    return df

@st.cache_data(ttl=30)
def carregar_medicao_subdivisoes(servico_id: int):
    conn = conectar_banco()
    try:
        df = pd.read_sql_query(
            "SELECT * FROM medicao_subdivisoes WHERE servico_id=%s ORDER BY id", conn, params=(servico_id,)
        )
    finally:
        liberar_conexao(conn)
    return df

@st.cache_data(ttl=30)
def carregar_medicao_historico(obra_id: int):
    conn = conectar_banco()
    try:
        df = pd.read_sql_query(
            "SELECT * FROM medicao_historico WHERE obra_id=%s ORDER BY periodo", conn, params=(obra_id,)
        )
    finally:
        liberar_conexao(conn)
    return df

def salvar_medicao_obra(nome, valor_m2, metragem):
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO medicao_obras (nome, valor_m2_global, metragem_geral) VALUES (%s,%s,%s) RETURNING id",
            (nome, valor_m2, metragem)
        )
        new_id = cursor.fetchone()[0]
        conn.commit()
        carregar_medicao_obras.clear()
        return new_id
    except Exception as e:
        conn.rollback()
        st.error(f"Erro ao salvar obra: {e}")
        return None
    finally:
        liberar_conexao(conn)

def excluir_medicao_obra(obra_id):
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM medicao_obras WHERE id=%s", (obra_id,))
        conn.commit()
        carregar_medicao_obras.clear()
        carregar_medicao_servicos.clear()
        carregar_medicao_subdivisoes.clear()
        carregar_medicao_historico.clear()
    except Exception as e:
        conn.rollback()
        st.error(f"Erro ao excluir obra: {e}")
    finally:
        liberar_conexao(conn)

def salvar_servicos_medicao(obra_id: int, periodo: str, servicos: list):
    """Salva serviços, subdivisões e snapshot do histórico para o período."""
    conn = conectar_banco()
    try:
        cursor = conn.cursor()
        # Apaga serviços antigos e recria (substituição total)
        cursor.execute("DELETE FROM medicao_servicos WHERE obra_id=%s", (obra_id,))
        total_medido = 0.0
        snapshot = []
        for srv in servicos:
            cursor.execute(
                "INSERT INTO medicao_servicos (obra_id, nome, valor_m2_servico) VALUES (%s,%s,%s) RETURNING id",
                (obra_id, srv['nome'], srv['valor_m2_servico'])
            )
            srv_id = cursor.fetchone()[0]
            subs_snap = []
            for sub in srv['subdivisoes']:
                cursor.execute(
                    "INSERT INTO medicao_subdivisoes (servico_id, nome, m2, percentual) VALUES (%s,%s,%s,%s)",
                    (srv_id, sub['nome'], sub['m2'], sub['percentual'])
                )
                val = sub['m2'] * (srv['valor_m2_servico'] * (sub['percentual'] / 100))
                total_medido += val
                subs_snap.append({**sub, 'subtotal': val})
            snapshot.append({**srv, 'subdivisoes': subs_snap})
        import json
        cursor.execute("""
            INSERT INTO medicao_historico (obra_id, periodo, total_medido, snapshot)
            VALUES (%s,%s,%s,%s)
            ON CONFLICT (obra_id, periodo)
            DO UPDATE SET total_medido=EXCLUDED.total_medido, snapshot=EXCLUDED.snapshot, criado_em=NOW()
        """, (obra_id, periodo, total_medido, json.dumps(snapshot)))
        conn.commit()
        carregar_medicao_servicos.clear()
        carregar_medicao_subdivisoes.clear()
        carregar_medicao_historico.clear()
        return total_medido
    except Exception as e:
        conn.rollback()
        st.error(f"Erro ao salvar serviços: {e}")
        return None
    finally:
        liberar_conexao(conn)

# ========================================================
# HELPER — BLOCOS SEMANAIS
# ========================================================
def blocos_semanais(df_input):
    if df_input.empty:
        st.info("Nenhum lote encontrado.")
        return
    df_input = df_input.copy()
    df_input['Ano_Semana'] = df_input['Data_Producao_Programada'].dt.isocalendar().year
    df_input['Num_Semana'] = df_input['Data_Producao_Programada'].dt.isocalendar().week

    def fmt_sem(r):
        try:
            s = pd.to_datetime(f"{int(r['Ano_Semana'])}-W{int(r['Num_Semana'])}-1", format="%G-W%V-%u")
            return f"Semana {int(r['Num_Semana']):02d} ({s.strftime('%d/%m')} – {(s + timedelta(days=6)).strftime('%d/%m/%Y')})"
        except Exception:
            return f"Semana {r['Num_Semana']}"

    df_input['Periodo'] = df_input.apply(fmt_sem, axis=1)
    semanas_unicas = df_input[['Ano_Semana', 'Num_Semana', 'Periodo']].drop_duplicates().sort_values(['Ano_Semana', 'Num_Semana'])

    for _, sem_row in semanas_unicas.iterrows():
        lotes_sem = df_input[df_input['Periodo'] == sem_row['Periodo']]
        total_m2  = lotes_sem['M2_Item'].sum()
        n_lotes   = len(lotes_sem)
        with st.expander(f"📅 {sem_row['Periodo']}  —  {n_lotes} lote(s)  |  {total_m2:.2f} m²", expanded=False):
            for _, lrow in lotes_sem.iterrows():
                _dt_ini = pd.to_datetime(lrow['Data_Producao_Programada'])
                _dt_fim = pd.to_datetime(lrow['Data_Limite_Obra'])
                dt_ini = _dt_ini.strftime('%d/%m/%Y') if pd.notna(_dt_ini) else '—'
                dt_fim = _dt_fim.strftime('%d/%m/%Y') if pd.notna(_dt_fim) else '—'
                st.markdown(
                    f'<span class="badge-obra">{lrow["Obra_Vinculada"]}</span>&nbsp;'
                    f'<span class="badge-edt">{lrow["EDT_Vinculado"]}</span>&nbsp;'
                    f'<span class="badge-lote">{lrow["Cod_Lote"]}</span>',
                    unsafe_allow_html=True
                )
                st.markdown(f"**{lrow['Tipo_Material']}** &nbsp;|&nbsp; `{int(lrow['Qtd_Caixas'])} caixas` — {lrow['M2_Item']:.2f} m²")
                st.caption(f"Período: {dt_ini} a {dt_fim} &nbsp;|&nbsp; {lrow['Romaneio_Chapas']}")
                op_txt = lrow['Num_OP'] if lrow.get('Num_OP') else "Aguardando OP"
                st.caption(f"OP: {op_txt} &nbsp;|&nbsp; {lrow.get('Fase_Produtiva', '—')}")
                st.markdown("---")

# ========================================================
# LOGIN
# ========================================================
# ========================================================
# LOGIN
# ========================================================
if 'autenticado' not in st.session_state:
    st.session_state.autenticado      = False
    st.session_state.usuario_nome     = ""
    st.session_state.usuario_setor    = ""
    st.session_state.usuario_login    = ""
    st.session_state.ultima_atividade = None

# ── Verificar timeout de sessão ──────────────────────────
if st.session_state.autenticado and verificar_timeout_sessao():
    st.session_state.autenticado      = False
    st.session_state.usuario_nome     = ""
    st.session_state.usuario_setor    = ""
    st.session_state.usuario_login    = ""
    st.session_state.ultima_atividade = None
    st.warning(f"⏱️ Sessão encerrada por inatividade ({TIMEOUT_SESSAO_HORAS}h). Faça login novamente.")

if not st.session_state.autenticado:
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        ci1, ci2, ci3 = st.columns([1, 2, 1])
        with ci2:
            st.image("assets/LOGO_BAUDENPASSOLD.png", use_container_width=True)
        st.markdown("""
        <p style='text-align:center; color:#64748B; font-size:13px; margin-top:8px; letter-spacing:0.05em;'>
        PCP & Controle Operacional
        </p>
        """, unsafe_allow_html=True)
        with st.form("form_login", border=True):
            st.markdown("<p style='text-align:center;font-size:16px;font-weight:600;color:#0F172A;margin-bottom:8px;'>Acesso ao Sistema</p>", unsafe_allow_html=True)
            user_input = st.text_input("Usuário:")
            pass_input = st.text_input("Senha:", type="password")

            if user_input.strip():
                bloqueado, seg_rest = verificar_bloqueio(user_input.strip())
                restantes = tentativas_restantes(user_input.strip())
                if bloqueado:
                    minutos  = seg_rest // 60
                    segundos = seg_rest % 60
                    st.error(f"🔒 Conta bloqueada. Aguarde **{minutos}m {segundos}s** para tentar novamente.")
                elif restantes < MAX_TENTATIVAS_LOGIN:
                    st.warning(f"⚠️ {restantes} tentativa(s) restante(s) antes do bloqueio.")

            if st.form_submit_button("Entrar", use_container_width=True):
                usuario_limpo = user_input.strip()
                bloqueado, seg_rest = verificar_bloqueio(usuario_limpo)
                if bloqueado:
                    st.error(f"🔒 Conta bloqueada. Aguarde {seg_rest // 60} minuto(s).")
                else:
                    with st.spinner("Verificando..."):
                        dados = verificar_login(usuario_limpo, pass_input)
                    if dados:
                        resetar_tentativas(usuario_limpo)
                        st.session_state.autenticado      = True
                        st.session_state.usuario_nome     = dados[0]
                        st.session_state.usuario_setor    = dados[1]
                        st.session_state.usuario_login    = usuario_limpo
                        st.session_state.ultima_atividade = datetime.now()
                        registrar_auditoria(usuario_limpo, "LOGIN", f"Login bem-sucedido — {dados[1]}")
                        st.rerun()
                    else:
                        registrar_tentativa_falha(usuario_limpo)
                        restantes = tentativas_restantes(usuario_limpo)
                        if restantes > 0:
                            st.error(f"❌ Usuário ou senha inválidos. {restantes} tentativa(s) restante(s).")
                        else:
                            st.error(f"🔒 Muitas tentativas. Conta bloqueada por {BLOQUEIO_MINUTOS} minutos.")
    st.stop()
    # Registrar atividade a cada interação
registrar_atividade()

df_banco_macro = carregar_macro()
df_banco_micro = carregar_micro()
df_projetos = carregar_projetos()

setor = st.session_state.usuario_setor

# ── Mapeamento de páginas por setor ──────────────────────────────────────────
GRUPOS_NAV = {
    "🏭  Produção": {
        "Painel da Producao - ACM":       ("📋  Painel ACM",         ["Master","Producao","Diretoria","Engenharia","PCP"]),
        "Painel TV — ACM":                ("📺  TV — ACM",           ["Master","Producao","Diretoria","Medicao","PCP"]),
        "Painel da Producao - Esquadrias":("📋  Painel Esquadrias",  ["Master","Esquadria","Producao","Diretoria","Engenharia","PCP"]),
        "Painel TV — Esquadrias":         ("📺  TV — Esquadrias",    ["Master","Esquadria","Producao","Diretoria","Medicao","PCP"]),
        "Liberar OPs da Semana":          ("🔓  Liberar OPs",        ["Master","PCP"]),
    },
    "📐  Engenharia": {
        "Painel de Engenharia": ("📐  Painel de Engenharia", ["Master","Engenharia"]),
        "Visao Macro":          ("🗺️  Visão Macro",          ["Master","Diretoria","Medicao"]),
        "Cadastrar Obra":       ("🏗️  Cadastrar Obra",       ["Master"]),
        "Vincular Datas":       ("📅  Vincular Datas",       ["Master"]),
    },
    "🚚  Operações": {
        "Logistica":      ("🚚  Logística",       ["Master","Logistica"]),
        "Almoxarifado":   ("📦  Almoxarifado",    ["Master","Almoxarifado"]),
        "Romaneio Manual": ("📋  Romaneio Manual", ["Master","Almoxarifado","PCP"]),
    },
    "📊  Relatórios": {
        "Relatorio Geral":    ("📊  Relatório Geral",    ["Master","Diretoria","PCP","Medicao"]),
        "Sistema de Medicao": ("📏  Sistema de Medição", ["Master","Medicao"]),
    },
    "⚙️  Sistema": {
        "Configuracoes": ("⚙️  Configurações", ["Master"]),
    },
}

# ── Constrói lista de páginas disponíveis para este setor ────────────────────
paginas_disponiveis = ["Dashboard"]

with st.sidebar:
    # Logo
    try:
        st.image("assets/LOGO_BAUDENPASSOLD.png", use_container_width=True)
    except Exception:
        st.markdown('<div class="sidebar-logo"><div class="sidebar-logo-name">PSF</div><div class="sidebar-logo-sub">Gestão Industrial</div></div>', unsafe_allow_html=True)

    st.markdown("<hr>", unsafe_allow_html=True)

    # Seletor de obra
    st.markdown('<p class="nav-section-label">Obra de trabalho</p>', unsafe_allow_html=True)
    if not df_banco_macro.empty:
        obras_lista      = ["Todas as Obras"] + sorted(df_banco_macro['Obra'].unique().tolist())
        obra_selecionada = st.selectbox("Obra:", obras_lista, label_visibility="collapsed", key="obra_global")
        if obra_selecionada == "Todas as Obras":
            obra_selecionada = None
        df_macro_filtrado = carregar_macro_por_obra(obra_selecionada) if obra_selecionada else pd.DataFrame()
    else:
        obra_selecionada  = None
        df_macro_filtrado = pd.DataFrame()
        st.caption("Nenhuma obra cadastrada.")

    st.markdown("<hr>", unsafe_allow_html=True)

    # Garante pagina_atual válida
    if "pagina_atual" not in st.session_state:
        st.session_state.pagina_atual = "Dashboard"

    # Menu por grupos com botões reais
    st.markdown("""
    <style>
    section[data-testid="stSidebar"] .stButton > button {
        background: rgba(255,255,255,0.06) !important;
        color: #E2E8F0 !important;
        border: none !important;
        border-radius: 6px !important;
        padding: 9px 14px !important;
        font-size: 0.88rem !important;
        font-weight: 500 !important;
        width: 100% !important;
        box-shadow: none !important;
        margin: 2px 0 !important;
        text-align: left !important;
        justify-content: flex-start !important;
    }
    section[data-testid="stSidebar"] .stButton > button:hover {
        background: rgba(255,255,255,0.14) !important;
        color: #fff !important;
        transform: none !important;
        box-shadow: none !important;
    }
    section[data-testid="stSidebar"] .stButton > button p {
        color: #E2E8F0 !important;
        font-size: 0.88rem !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # Dashboard sempre primeiro
    _ativo = st.session_state.pagina_atual == "Dashboard"
    _estilo_ativo = "background:rgba(59,130,246,0.2)!important;color:#fff!important;font-weight:700!important;border-left:3px solid #3B82F6!important;"
    if _ativo:
        st.markdown(f'<div style="{_estilo_ativo}border-radius:6px;padding:8px 12px;font-size:0.87rem;margin:1px 0;">🏠  Dashboard</div>', unsafe_allow_html=True)
    else:
        if st.button("🏠  Dashboard", key="nav_Dashboard"):
            st.session_state.pagina_atual = "Dashboard"
            st.rerun()

    for grupo, itens in GRUPOS_NAV.items():
        itens_grupo = [(k, v[0]) for k, v in itens.items() if setor in v[1]]
        if not itens_grupo:
            continue
        grupo_label = grupo.split("  ")[1] if "  " in grupo else grupo
        st.markdown(f'<p class="nav-section-label">{grupo_label}</p>', unsafe_allow_html=True)
        for chave, label in itens_grupo:
            paginas_disponiveis.append(chave)
            _ativo = st.session_state.pagina_atual == chave
            if _ativo:
                st.markdown(f'<div style="{_estilo_ativo}border-radius:6px;padding:8px 12px;font-size:0.87rem;margin:1px 0;">{label}</div>', unsafe_allow_html=True)
            else:
                if st.button(label, key=f"nav_{chave}"):
                    st.session_state.pagina_atual = chave
                    st.rerun()

    st.markdown("<hr>", unsafe_allow_html=True)
    ultima_at  = st.session_state.get("ultima_atividade")
    timeout_em = ultima_at + timedelta(hours=TIMEOUT_SESSAO_HORAS) if ultima_at else None
    st.caption(f"👤 {st.session_state.usuario_nome}  ·  {setor}")
    if timeout_em:
        st.caption(f"⏱️ Sessão até {timeout_em.strftime('%H:%M')}")

    with st.expander("🔑 Trocar minha senha"):
        with st.form("form_trocar_senha", clear_on_submit=True):
            senha_atual_inp = st.text_input("Senha atual:", type="password", key="senha_atual_troca")
            senha_nova_inp  = st.text_input("Nova senha:", type="password", key="senha_nova_troca")
            senha_conf_inp  = st.text_input("Confirmar nova senha:", type="password", key="senha_conf_troca")
            if st.form_submit_button("Salvar nova senha"):
                if not senha_atual_inp or not senha_nova_inp:
                    st.error("Preencha todos os campos.")
                elif len(senha_nova_inp) < 8:
                    st.error("A nova senha precisa ter pelo menos 8 caracteres.")
                elif senha_nova_inp != senha_conf_inp:
                    st.error("A confirmação não confere com a nova senha.")
                else:
                    ok, msg = trocar_propria_senha(
                        st.session_state.usuario_login, senha_atual_inp, senha_nova_inp
                    )
                    if ok:
                        registrar_auditoria(st.session_state.usuario_nome, "TROCAR_SENHA", "Senha alterada pelo próprio usuário.")
                        st.success(msg)
                    else:
                        st.error(msg)

    if st.button("↩ Sair", key="btn_sair_sidebar"):
        registrar_auditoria(st.session_state.usuario_nome, "LOGOUT", "Logout manual.")
        st.session_state.autenticado      = False
        st.session_state.ultima_atividade = None
        st.rerun()

# ── Topbar ────────────────────────────────────────────────────────────────────
agora_str = datetime.now(FUSO_BR).strftime("%d/%m/%Y  %H:%M")

# Busca notificações recentes do log de auditoria
_ICONES_ACAO = {
    "LOGIN": "🔐", "LOGOUT": "🚪",
    "LIBERAR_OPS": "🔓", "LANCAMENTO_PECAS": "📦",
    "EXCLUIR_FRENTE": "🗑️", "EXCLUIR_OBRA": "🗑️",
    "CRIAR_USUARIO": "👤", "OP_AVULSA": "📋",
    "SOLICITAR_OP": "📝", "VINCULAR_SOLICITACAO_OP": "🔗",
    "GERAR_LOTE": "⚙️", "EXCLUIR_LOTE": "🗑️",
    "CADASTRAR_OBRA": "🏗️", "CADASTRAR_FRENTE": "📐",
    "ENVIO_TOTAL": "🚚", "DESPACHO_LOGISTICA": "🚚",
}

# Filtros de notificação por setor — cada setor vê só o que lhe interessa
_FILTROS_NOTIF = {
    "Master":        None,   # vê tudo (exceto login/logout)
    "PCP":           ["SOLICITAR_OP", "LIBERAR_OPS", "GERAR_LOTE", "CADASTRAR_OBRA", "CADASTRAR_FRENTE", "VINCULAR_SOLICITACAO_OP"],
    "Producao":      ["LIBERAR_OPS", "LANCAMENTO_PECAS", "GERAR_LOTE"],
    "Engenharia":    ["CADASTRAR_FRENTE", "VINCULAR_SOLICITACAO_OP", "SOLICITAR_OP"],
    "Diretoria":     ["CADASTRAR_OBRA", "GERAR_LOTE", "ENVIO_TOTAL", "LIBERAR_OPS"],
    "Logistica":     ["ENVIO_TOTAL", "DESPACHO_LOGISTICA", "GERAR_LOTE"],
    "Almoxarifado":  ["SOLICITAR_OP", "OP_AVULSA"],
    "Medicao":       ["CADASTRAR_OBRA", "VINCULAR_SOLICITACAO_OP"],
    "Esquadria":     ["LIBERAR_OPS", "GERAR_LOTE", "LANCAMENTO_PECAS"],
}

_TITULOS_NOTIF = {
    "SOLICITAR_OP":           "Nova solicitação de OP recebida",
    "LIBERAR_OPS":            "OPs liberadas para fábrica",
    "LANCAMENTO_PECAS":       "Lançamento de peças registrado",
    "GERAR_LOTE":             "Novo lote gerado",
    "CADASTRAR_OBRA":         "Nova obra cadastrada",
    "CADASTRAR_FRENTE":       "Nova frente cadastrada",
    "VINCULAR_SOLICITACAO_OP":"Solicitação vinculada a OP",
    "OP_AVULSA":              "OP avulsa lançada",
    "ENVIO_TOTAL":            "Envio total registrado",
    "DESPACHO_LOGISTICA":     "Despacho logístico registrado",
    "EXCLUIR_OBRA":           "Obra excluída",
    "EXCLUIR_FRENTE":         "Frente excluída",
    "EXCLUIR_LOTE":           "Lote excluído",
    "CRIAR_USUARIO":          "Novo usuário criado",
}

_ACOES_AUDITORIA = ["LOGIN", "LOGOUT"]   # só Master vê no painel de auditoria

@st.cache_data(ttl=15)
def _carregar_notificacoes(setor_usr: str):
    conn = conectar_banco()
    try:
        filtro = _FILTROS_NOTIF.get(setor_usr)
        if filtro is None:
            # Master: tudo menos login/logout
            df = pd.read_sql_query(
                "SELECT usuario, acao, detalhes, criado_em FROM auditoria_log "
                "WHERE acao NOT IN ('LOGIN','LOGOUT') ORDER BY id DESC LIMIT 30",
                conn
            )
        else:
            placeholders = ",".join(["%s"] * len(filtro))
            df = pd.read_sql_query(
                f"SELECT usuario, acao, detalhes, criado_em FROM auditoria_log "
                f"WHERE acao IN ({placeholders}) ORDER BY id DESC LIMIT 30",
                conn, params=filtro
            )
    except Exception:
        df = pd.DataFrame()
    finally:
        liberar_conexao(conn)
    return df

_df_notif = _carregar_notificacoes(setor)
_n_notif  = len(_df_notif)

if "notif_aberto" not in st.session_state:
    st.session_state.notif_aberto = False

tb_esq, tb_meio, tb_dir = st.columns([4, 1, 3])
with tb_esq:
    st.markdown('<span style="font-size:1.1rem;font-weight:700;color:#1E3A5F;">Passold <span style="color:#1A56DB;">Sistemas de Fachadas</span></span>', unsafe_allow_html=True)

with tb_meio:
    _label_sino = f"🔔 {_n_notif}" if _n_notif > 0 else "🔔"
    if st.button(_label_sino, key="btn_notif", help="Notificações recentes", use_container_width=True):
        st.session_state.notif_aberto = not st.session_state.notif_aberto

with tb_dir:
    st.markdown(f"""
    <div style="display:flex;align-items:center;justify-content:flex-end;gap:16px;padding:10px 0;">
        <span style="background:#EFF6FF;color:#1A56DB;padding:3px 12px;border-radius:20px;
                     font-size:0.75rem;font-weight:700;border:1px solid rgba(26,86,219,0.2);">{setor}</span>
        <span style="font-size:0.82rem;color:#64748B;">👤 <strong style="color:#1E293B;">{st.session_state.usuario_nome}</strong></span>
        <span style="font-size:0.78rem;color:#64748B;font-family:'JetBrains Mono',monospace;">🕐 {agora_str}</span>
    </div>
    """, unsafe_allow_html=True)

st.markdown('<hr style="margin:0 0 16px 0;border-color:#E2E8F0;">', unsafe_allow_html=True)

# ── Painel de notificações ────────────────────────────────
if st.session_state.notif_aberto:
    with st.container():
        st.markdown("""
        <div style="background:#fff;border:1px solid #E2E8F0;border-radius:12px;
                    box-shadow:0 8px 24px rgba(0,0,0,0.12);padding:0;overflow:hidden;
                    margin-bottom:16px;">
            <div style="background:#1E3A5F;padding:14px 20px;display:flex;
                        align-items:center;justify-content:space-between;">
                <span style="color:#fff;font-weight:700;font-size:0.95rem;">🔔 Notificações Recentes</span>
                <span style="color:#94A3B8;font-size:0.75rem;">Últimas 30 ações do sistema</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

        if _df_notif.empty:
            st.markdown("""
            <div style="padding:24px;text-align:center;color:#94A3B8;">
                <div style="font-size:2rem;margin-bottom:8px;">📭</div>
                <div>Nenhuma notificação para o seu setor.</div>
            </div>""", unsafe_allow_html=True)
        else:
            for _, row in _df_notif.iterrows():
                _acao_raw = str(row.get('acao', ''))
                icone  = _ICONES_ACAO.get(_acao_raw, "📌")
                titulo = _TITULOS_NOTIF.get(_acao_raw, _acao_raw.replace('_', ' ').title())
                det    = str(row.get('detalhes', ''))
                det    = det[:90] + ("…" if len(det) > 90 else "")
                usr    = str(row.get('usuario', ''))
                hora   = str(row.get('criado_em', ''))[:16]
                st.markdown(f"""
                <div style="display:flex;align-items:flex-start;gap:14px;padding:12px 20px;
                            border-bottom:1px solid #F1F5F9;background:#fff;">
                    <div style="width:38px;height:38px;border-radius:10px;background:#EFF6FF;
                                display:flex;align-items:center;justify-content:center;
                                font-size:1.2rem;flex-shrink:0;">{icone}</div>
                    <div style="flex:1;min-width:0;">
                        <div style="font-weight:700;font-size:0.85rem;color:#1E293B;">{titulo}</div>
                        <div style="font-size:0.78rem;color:#64748B;margin-top:3px;
                                    white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">{det}</div>
                        <div style="font-size:0.7rem;color:#94A3B8;margin-top:4px;">
                            👤 {usr} &nbsp;·&nbsp; 🕐 {hora}
                        </div>
                    </div>
                </div>
                """, unsafe_allow_html=True)

        if st.button("✕ Fechar notificações", key="fechar_notif"):
            st.session_state.notif_aberto = False
            st.rerun()

# ── Remove o header antigo (evita duplicar) ───────────────────────────────────
abas_disponiveis = paginas_disponiveis   # compatibilidade com código abaixo

# ── Compatibilidade: aba_objeto agora é um container único ────────────────────
class _FakePage:
    """Contexto vazio — substitui o with aba_objeto: das tabs antigas."""
    def __enter__(self): return self
    def __exit__(self, *_): pass

aba_objeto = _FakePage()

# ── Renderiza só a página selecionada ────────────────────────────────────────
if True:  # mantém indentação do bloco original
    nome_aba = st.session_state.pagina_atual

# ── DASHBOARD ─────────────────────────────────────────────────────────────────
if nome_aba == "Dashboard":
    with aba_objeto:
        st.markdown("""
        <div class="page-header">
            <div class="page-header-left">
                <h2>Dashboard</h2>
                <p>Resumo geral da operação em tempo real</p>
            </div>
            <span class="page-icon">🏠</span>
        </div>
        """, unsafe_allow_html=True)

        df_micro_dash = carregar_micro()
        df_macro_dash = carregar_macro()

        ops_pendentes    = len(df_micro_dash[df_micro_dash['Status_Item'] == 'Pendente'])           if not df_micro_dash.empty else 0
        ops_liberadas    = len(df_micro_dash[df_micro_dash['Status_Item'] == 'Liberado para Fabrica']) if not df_micro_dash.empty else 0
        ops_concluidas   = len(df_micro_dash[df_micro_dash['Status_Item'] == 'Concluido'])          if not df_micro_dash.empty else 0
        total_obras      = df_micro_dash['Obra_Vinculada'].nunique()                                 if not df_micro_dash.empty else 0

        hoje = datetime.now().date()
        if not df_micro_dash.empty and 'Data_Limite_Obra' in df_micro_dash.columns:
            df_ativos = df_micro_dash[~df_micro_dash['Status_Item'].isin(['Concluido','Cancelado'])].copy()
            df_ativos['Data_Limite_Obra'] = pd.to_datetime(df_ativos['Data_Limite_Obra'], errors='coerce')
            lotes_atrasados = len(df_ativos[df_ativos['Data_Limite_Obra'].dt.date < hoje])
        else:
            lotes_atrasados = 0

        c1, c2, c3, c4, c5 = st.columns(5)
        with c1:
            st.markdown(f"""
            <div class="dash-card">
                <div class="dash-card-title">OPs Pendentes</div>
                <div class="dash-card-value orange">{ops_pendentes}</div>
                <div class="dash-card-sub">aguardando liberação</div>
            </div>""", unsafe_allow_html=True)
        with c2:
            st.markdown(f"""
            <div class="dash-card">
                <div class="dash-card-title">Liberadas p/ Fábrica</div>
                <div class="dash-card-value">{ops_liberadas}</div>
                <div class="dash-card-sub">em produção</div>
            </div>""", unsafe_allow_html=True)
        with c3:
            st.markdown(f"""
            <div class="dash-card">
                <div class="dash-card-title">Concluídas</div>
                <div class="dash-card-value green">{ops_concluidas}</div>
                <div class="dash-card-sub">finalizadas</div>
            </div>""", unsafe_allow_html=True)
        with c4:
            cor_atr = "red" if lotes_atrasados > 0 else "green"
            st.markdown(f"""
            <div class="dash-card">
                <div class="dash-card-title">Lotes Atrasados</div>
                <div class="dash-card-value {cor_atr}">{lotes_atrasados}</div>
                <div class="dash-card-sub">prazo vencido</div>
            </div>""", unsafe_allow_html=True)
        with c5:
            st.markdown(f"""
            <div class="dash-card">
                <div class="dash-card-title">Obras Ativas</div>
                <div class="dash-card-value">{total_obras}</div>
                <div class="dash-card-sub">em andamento</div>
            </div>""", unsafe_allow_html=True)

        # ── Faixa de alertas visível ──────────────────────────────
        alertas = []
        if lotes_atrasados > 0:
            alertas.append(("danger", "🔴", f"{lotes_atrasados} lote(s) atrasado(s)", "Prazo vencido — acesse Relatório Geral"))
        if ops_pendentes > 0:
            alertas.append(("warn", "⚠️", f"{ops_pendentes} OP(s) aguardando liberação", "Acesse Liberar OPs da Semana"))
        if not df_macro_dash.empty and 'Status_Engenharia' in df_macro_dash.columns:
            df_macro_eng = df_macro_dash[df_macro_dash['Detalhado'] != False] if 'Detalhado' in df_macro_dash.columns else df_macro_dash
            criticos_eng = len(df_macro_eng[df_macro_eng['Status_Engenharia'].str.contains('Revisao|Aguardando', na=False)])
            if criticos_eng > 0:
                alertas.append(("warn", "📐", f"{criticos_eng} frente(s) em revisão ou aguardando", "Acesse Painel de Engenharia"))

        if alertas:
            st.markdown("---")
            st.markdown("#### 🚨 Alertas do Sistema")
            cols_alertas = st.columns(len(alertas))
            estilos = {
                "danger": ("background:#FEF2F2;border:1px solid #FECACA;border-left:5px solid #DC2626;", "#DC2626"),
                "warn":   ("background:#FFFBEB;border:1px solid #FDE68A;border-left:5px solid #D97706;", "#D97706"),
            }
            for col, (tipo, icone, titulo, subtitulo) in zip(cols_alertas, alertas):
                estilo, cor = estilos[tipo]
                with col:
                    st.markdown(f"""
                    <div style="{estilo} border-radius:10px; padding:16px 20px;">
                        <div style="font-size:1.6rem; margin-bottom:6px;">{icone}</div>
                        <div style="font-size:1rem; font-weight:800; color:{cor}; line-height:1.2;">{titulo}</div>
                        <div style="font-size:0.75rem; color:#6B7280; margin-top:4px;">{subtitulo}</div>
                    </div>
                    """, unsafe_allow_html=True)
        else:
            st.markdown("---")
            st.markdown("""
            <div style="background:#ECFDF5;border:1px solid #A7F3D0;border-left:5px solid #059669;
                        border-radius:10px;padding:14px 20px;display:flex;align-items:center;gap:12px;">
                <span style="font-size:1.4rem;">✅</span>
                <div>
                    <strong style="color:#065F46;">Tudo em ordem!</strong>
                    <span style="color:#6B7280;font-size:0.85rem;margin-left:8px;">Nenhum alerta crítico no momento.</span>
                </div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        col_esq, col_dir = st.columns([3, 2])

        with col_esq:
            st.markdown("#### 📋 Status dos Lotes")
            if not df_micro_dash.empty:
                status_counts = df_micro_dash['Status_Item'].value_counts().reset_index()
                status_counts.columns = ['Status', 'Quantidade']
                cores_status = {
                    'Pendente':                '#D97706',
                    'Liberado para Fabrica':   '#0EA5E9',
                    'Parcialmente Concluido':  '#8B5CF6',
                    'Concluido':               '#059669',
                    'Cancelado':               '#DC2626',
                }
                fig = px.bar(status_counts, x='Status', y='Quantidade',
                             color='Status',
                             color_discrete_map=cores_status,
                             template='plotly_white')
                fig.update_layout(
                    height=280, showlegend=False,
                    margin=dict(l=0, r=0, t=10, b=0),
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)',
                    font=dict(family='Inter', size=12),
                    xaxis=dict(title='', tickfont=dict(size=11)),
                    yaxis=dict(title='Qtd', gridcolor='#F1F5F9'),
                )
                fig.update_traces(marker_line_width=0)
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.markdown('<div class="empty-state"><div class="empty-icon">📊</div><h4>Sem dados</h4><p>Nenhum lote cadastrado ainda.</p></div>', unsafe_allow_html=True)

        with col_dir:
            st.markdown("#### 📦 Lotes por Obra")
            if not df_micro_dash.empty:
                por_obra = df_micro_dash.groupby('Obra_Vinculada').size().reset_index(name='Lotes')
                por_obra = por_obra.sort_values('Lotes', ascending=False).head(8)
                st.dataframe(por_obra, hide_index=True, use_container_width=True,
                             column_config={"Obra_Vinculada": "Obra", "Lotes": "Qtd Lotes"})

elif nome_aba not in abas_disponiveis:
    st.warning("Página não encontrada.")

# ── Bloco de compatibilidade — substitui o loop de tabs ──────────────────────
# O código abaixo usa `if nome_aba == "X":` diretamente (sem for loop)
if False:  # bloco morto — mantém o parser feliz com o código legado abaixo
    for nome_aba, aba_objeto in zip([], []):
        pass

for nome_aba, aba_objeto in [(st.session_state.pagina_atual, _FakePage())]:

    # ==================================================
    # PAINEL DA PRODUCAO ACM
    # ==================================================
    if nome_aba == "Painel da Producao - ACM":
        with aba_objeto:
            st.markdown('<div class="page-header"><div class="page-header-left"><h2>Painel da Produção — ACM</h2><p>Acompanhamento em tempo real dos lotes e OPs de ACM</p></div><span class="page-icon">📋</span></div>', unsafe_allow_html=True)
            obras_tv = ["Todas as obras"] + (
                list(df_banco_micro['Obra_Vinculada'].dropna().unique()) if not df_banco_micro.empty else []
            )
            obra_tv = st.selectbox("Filtrar por obra:", obras_tv, key="sb_obra_tv")

            st.markdown("### Previsão de Entrada em Produção")
            st.caption("Lotes planejados no Vincular Datas — ainda não liberados oficialmente.")
            df_prev = df_banco_micro[df_banco_micro['Escopo'] == 'ACM'].copy() if not df_banco_micro.empty else pd.DataFrame()
            if obra_tv != "Todas as obras" and not df_prev.empty:
                df_prev = df_prev[df_prev['Obra_Vinculada'] == obra_tv]
            df_prev_pend = df_prev[df_prev['Status_Item'] == 'Pendente'].copy() if not df_prev.empty else pd.DataFrame()
            if df_prev_pend.empty:
                st.info("Nenhuma previsão pendente.")
            else:
                blocos_semanais(df_prev_pend)

            st.markdown("---")
            st.markdown("### Calendário de Produção — OPs Liberadas")
            st.caption("Apenas lotes liberados oficialmente na aba 'Liberar OPs da Semana'.")
            if not df_banco_micro.empty:
                df_base = df_banco_micro[
                    df_banco_micro['Status_Item'].isin(["Liberado para Fabrica", "Parcialmente Concluido"]) &
                    (df_banco_micro['Escopo'] == 'ACM')
                ].copy()
                if obra_tv != "Todas as obras":
                    df_base = df_base[df_base['Obra_Vinculada'] == obra_tv]
                if not df_base.empty:
                    registros_exp = []
                    for _, row in df_base.iterrows():
                        dt_ini = pd.to_datetime(row['Data_Producao_Programada']).date()
                        dt_fim = pd.to_datetime(row['Data_Limite_Obra']).date()
                        ini_ultima_semana = (pd.to_datetime(dt_fim) - timedelta(days=6)).date()
                        ini_ultima_semana = max(ini_ultima_semana, dt_ini)
                        dia = dt_ini
                        while dia <= dt_fim:
                            if dia.weekday() < 5:  # sem producao aos finais de semana
                                r = row.to_dict()
                                r['_dia'] = dia
                                r['_pode_concluir'] = (dia >= ini_ultima_semana)
                                registros_exp.append(r)
                            dia += timedelta(days=1)
                    df_exp = pd.DataFrame(registros_exp) if registros_exp else pd.DataFrame()

                    if "prog_mes" not in st.session_state:
                        st.session_state.prog_mes = hoje_projeto().month
                    if "prog_ano" not in st.session_state:
                        st.session_state.prog_ano = hoje_projeto().year

                    c1, c2, c3 = st.columns([1, 2, 1])
                    with c1:
                        if st.button("Mes Anterior", use_container_width=True, key="btn_ant"):
                            st.session_state.prog_mes -= 1
                            if st.session_state.prog_mes == 0:
                                st.session_state.prog_mes = 12
                                st.session_state.prog_ano -= 1
                            st.rerun()
                    with c2:
                        nomes_meses = ["", "Janeiro", "Fevereiro", "Marco", "Abril", "Maio", "Junho",
                                       "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
                        st.markdown(
                            f"<h3 style='text-align:center;color:#1E3A8A;margin:0;'>"
                            f"{nomes_meses[st.session_state.prog_mes]} / {st.session_state.prog_ano}</h3>",
                            unsafe_allow_html=True
                        )
                    with c3:
                        if st.button("Proximo Mes", use_container_width=True, key="btn_prox"):
                            st.session_state.prog_mes += 1
                            if st.session_state.prog_mes == 13:
                                st.session_state.prog_mes = 1
                                st.session_state.prog_ano += 1
                            st.rerun()

                    st.markdown("---")
                    cal     = py_calendar.Calendar(firstweekday=6)
                    semanas = cal.monthdatescalendar(st.session_state.prog_ano, st.session_state.prog_mes)
                    cols_h  = st.columns(7)
                    for i, nome in enumerate(["Dom", "Seg", "Ter", "Qua", "Qui", "Sex", "Sab"]):
                        cols_h[i].markdown(
                            f"<div style='text-align:center;font-weight:600;color:#475569;padding:4px 0;font-size:13px;'>{nome}</div>",
                            unsafe_allow_html=True
                        )
                    for semana in semanas:
                        cols = st.columns(7)
                        for i, data_dia in enumerate(semana):
                            with cols[i]:
                                if data_dia.month == st.session_state.prog_mes:
                                    lotes_dia = df_exp[df_exp['_dia'] == data_dia] if not df_exp.empty else pd.DataFrame()
                                    n_lotes   = lotes_dia['Cod_Lote'].nunique() if not lotes_dia.empty else 0
                                    eh_hoje   = (data_dia == hoje_projeto().date())
                                    if n_lotes > 0:
                                        obras_d   = lotes_dia['Obra_Vinculada'].unique()
                                        label_o   = obras_d[0] if len(obras_d) == 1 else f"{len(obras_d)} obras"
                                        btn_label = f"{data_dia.day}  |  {n_lotes} lote(s)\n{label_o}"
                                        if st.button(btn_label, key=f"btn_{data_dia}", use_container_width=True):
                                            st.session_state.dia_clicado_tv = data_dia
                                    else:
                                        css_class = "cal-day-today" if eh_hoje else "cal-day-empty"
                                        st.markdown(
                                            f"<div class='{css_class}'><span style='color:#94A3B8;font-size:15px;'>{data_dia.day}</span>"
                                            f"<br><span style='color:#CBD5E1;font-size:10px;'>—</span></div>",
                                            unsafe_allow_html=True
                                        )
                                else:
                                    st.markdown('<div style="height:70px;"></div>', unsafe_allow_html=True)

                    st.markdown("---")
                    if "dia_clicado_tv" not in st.session_state:
                        st.session_state.dia_clicado_tv = hoje_projeto().date()
                    dia_sel   = st.session_state.dia_clicado_tv
                    st.subheader(f"Lotes em producao — {dia_sel.strftime('%d/%m/%Y')}")
                    lotes_sel = (
                        df_exp[df_exp['_dia'] == dia_sel].drop_duplicates(subset=['id'])
                        if not df_exp.empty else pd.DataFrame()
                    )
                    if lotes_sel.empty:
                        st.info("Clique em um dia com lotes no calendario acima.")
                    else:
                        kc1, kc2, kc3 = st.columns(3)
                        kc1.metric("Lotes em producao", lotes_sel['Cod_Lote'].nunique())
                        kc2.metric("Total caixas", int(lotes_sel['Qtd_Caixas'].sum()))
                        kc3.metric("Total m²", f"{lotes_sel['M2_Item'].sum():.2f}")
                        st.markdown("---")
                        for _, row in lotes_sel.iterrows():
                            # Parcialmente Concluido sempre pode concluir o restante
                            eh_parcial    = row.get('Status_Item', '') == 'Parcialmente Concluido'
                            pode_concluir = bool(row.get('_pode_concluir', False)) or eh_parcial
                            dt_i = pd.to_datetime(row['Data_Producao_Programada']).strftime('%d/%m/%Y')
                            dt_f = pd.to_datetime(row['Data_Limite_Obra']).strftime('%d/%m/%Y')
                            border_color = "#D97706" if eh_parcial else ("#1A56DB" if pode_concluir else "#3B82F6")
                            bg_color     = "#FFFBEB" if eh_parcial else ("#FFF7ED" if pode_concluir else "#F8FAFC")
                            st.markdown(
                                f"<div style='border-left:4px solid {border_color};background:{bg_color};"
                                f"padding:12px 16px;border-radius:6px;margin-bottom:4px;'></div>",
                                unsafe_allow_html=True
                            )
                            with st.container(border=True):
                                cd, ca = st.columns([4, 1])
                                with cd:
                                    st.markdown(
                                        f'<span class="badge-obra">{row["Obra_Vinculada"]}</span>&nbsp;'
                                        f'<span class="badge-edt">{row["EDT_Vinculado"]}</span>&nbsp;'
                                        f'<span class="badge-lote">{row["Cod_Lote"]}</span>',
                                        unsafe_allow_html=True
                                    )
                                    st.markdown(f"**{row['Tipo_Material']}** &nbsp;|&nbsp; `{int(row['Qtd_Caixas'])} caixas` — {row['M2_Item']:.2f} m²")
                                    st.caption(f"Periodo: {dt_i} a {dt_f} &nbsp;|&nbsp; {row['Romaneio_Chapas']}")
                                    op_txt = row['Num_OP'] if row.get('Num_OP') else "Aguardando OP"
                                    st.caption(f"OP: {op_txt} &nbsp;|&nbsp; {row.get('Fase_Produtiva', '—')}")
                                    em_parada_op = bool(row.get('Em_Parada', False))
                                    _motivo_op_raw = row.get('Motivo_Parada')
                                    motivo_op    = html_escape(str(_motivo_op_raw)) if pd.notna(_motivo_op_raw) else ''
                                    if em_parada_op:
                                        st.markdown(f"<span style='color:#DC2626;font-size:12px;font-weight:700;'>⛔ EM PARADA — {motivo_op}</span>", unsafe_allow_html=True)
                                    elif eh_parcial:
                                        st.markdown("<span style='color:#D97706;font-size:12px;font-weight:700;'>🟠 Envio parcial registrado — ainda há peças pendentes</span>", unsafe_allow_html=True)
                                    elif pode_concluir:
                                        st.markdown("<span style='color:#1A56DB;font-size:12px;font-weight:600;'>Ultima semana — liberado para concluir</span>", unsafe_allow_html=True)
                                    else:
                                        dias_restantes = (pd.to_datetime(row['Data_Limite_Obra']).date() - dia_sel).days
                                        st.markdown(f"<span style='color:#3B82F6;font-size:12px;'>Em producao — {dias_restantes} dias ate o prazo</span>", unsafe_allow_html=True)
                                with ca:
                                    if setor in ["Producao", "Master"]:
                                        if em_parada_op:
                                            if st.button("▶ Retomar", key=f"retomar_{row['id']}", use_container_width=True):
                                                salvar_parada_op(row['id'], False, '', st.session_state.usuario_nome)
                                                st.toast("OP retomada!")
                                                st.rerun()
                                        else:
                                            if st.button("⏸ Parada", key=f"parada_{row['id']}", use_container_width=True):
                                                st.session_state[f"modal_parada_{row['id']}"] = not st.session_state.get(f"modal_parada_{row['id']}", False)
                                                st.rerun()
                                    if pode_concluir:
                                        st.write("")
                                        if st.button("✅ Pronto", key=f"baixa_{row['id']}", type="primary", use_container_width=True):
                                            st.session_state[f"modal_pronto_{row['id']}"] = not st.session_state.get(f"modal_pronto_{row['id']}", False)
                                            st.rerun()
                                    elif setor not in ["Producao", "Master"]:
                                        st.markdown("<div style='text-align:center;color:#94A3B8;font-size:12px;padding:8px;'>Em producao</div>", unsafe_allow_html=True)

                            # ── ARQUIVOS DA OP ────────────────────────────────
                            arqs_op = carregar_arquivos_op(int(row['id']))
                            if arqs_op:
                                with st.expander(f"📎 {len(arqs_op)} arquivo(s) anexado(s)", expanded=False):
                                    for arq in arqs_op:
                                        arq_id, arq_nome, arq_tipo, arq_enviado_por, _ = arq
                                        ca1, ca2 = st.columns([5, 1])
                                        ca1.markdown(f"📄 **{html_escape(arq_nome)}**  \n<small style='color:#94A3B8'>{html_escape(arq_enviado_por)}</small>", unsafe_allow_html=True)
                                        conteudo_arq = carregar_conteudo_arquivo(arq_id)
                                        if conteudo_arq:
                                            _, _, bytes_arq = conteudo_arq
                                            ca2.download_button(
                                                "⬇️", data=bytes(bytes_arq),
                                                file_name=arq_nome, mime=arq_tipo or "application/octet-stream",
                                                key=f"acm_dl_{arq_id}"
                                            )

                            # ── MODAL EM LARGURA TOTAL ─────────────────────────
                            if st.session_state.get(f"modal_pronto_{row['id']}", False):
                                with st.container(border=True):
                                    st.markdown(f"#### 📦 `{row['Cod_Lote']}` — {row['Obra_Vinculada']} — Tipo de Envio")
                                    limite_desp = None
                                    if not df_banco_macro.empty:
                                        fr = df_banco_macro[df_banco_macro['EDT'] == row['EDT_Vinculado']]
                                        if not fr.empty:
                                            limite_desp = fr.iloc[0].get('Data_Limite_Despacho')

                                    me1, me2 = st.columns([2, 3])
                                    with me1:
                                        tipo_envio = st.radio(
                                            "Este envio é:",
                                            ["Envio Total", "Envio Parcial"],
                                            key=f"tipo_envio_{row['id']}",
                                            horizontal=True
                                        )

                                    if tipo_envio == "Envio Total":
                                        _pecas_lote_check = carregar_pecas_lote(int(row['id']))
                                        with me2:
                                            if _pecas_lote_check.empty:
                                                st.warning("⚠️ Nenhuma peça lançada neste lote. Lance as peças na Seção 2 antes de concluir.")
                                            else:
                                                st.info("Todas as peças serão marcadas como concluídas e enviadas para logística.")
                                        bt1, bt2, _ = st.columns([2, 2, 4])
                                        with bt1:
                                            if st.button("✅ Confirmar Total", key=f"conf_total_{row['id']}", type="primary", use_container_width=True, disabled=_pecas_lote_check.empty):
                                                conn = conectar_banco()
                                                try:
                                                    cursor = conn.cursor()
                                                    cursor.execute(
                                                        "UPDATE itens_detalhado SET Status_Item='Concluido' WHERE id=%s",
                                                        (row['id'],)
                                                    )
                                                    cursor.execute(
                                                        "UPDATE op_pecas SET qtd_enviada=qtd_total, saldo=0 WHERE lote_id=%s",
                                                        (row['id'],)
                                                    )
                                                    conn.commit()
                                                except Exception as e:
                                                    conn.rollback()
                                                    st.error(f"Erro: {e}")
                                                finally:
                                                    liberar_conexao(conn)
                                                carregar_micro.clear()
                                                carregar_macro.clear()
                                                carregar_fila_logistica.clear()
                                                carregar_pecas_lote.clear()
                                                carregar_todas_pecas_obra.clear()
                                                enviar_para_logistica(row, limite_desp if prazo_valido(limite_desp) else pd.NaT)
                                                st.session_state[f"modal_pronto_{row['id']}"] = False
                                                st.toast(f"✅ {row['Cod_Lote']} concluido!")
                                                time.sleep(0.5)
                                                st.rerun()
                                        with bt2:
                                            if st.button("Cancelar", key=f"cancel_total_{row['id']}", use_container_width=True):
                                                st.session_state[f"modal_pronto_{row['id']}"] = False
                                                st.rerun()

                                    else:
                                        # Envio Parcial
                                        df_pecas_parc = carregar_pecas_lote(int(row['id']))
                                        if df_pecas_parc.empty:
                                            st.warning("⚠️ Nenhuma peça lançada para este lote. Lance as peças na aba 'Liberar OPs da Semana' primeiro.")
                                            if st.button("Fechar", key=f"cancel_parc_{row['id']}"):
                                                st.session_state[f"modal_pronto_{row['id']}"] = False
                                                st.rerun()
                                        else:
                                            st.caption("Informe quantas unidades de cada peça estão prontas para este envio:")
                                            st.markdown("---")

                                            pecas_envio = []
                                            # Cabeçalho da tabela
                                            th1, th2, th3, th4, th5 = st.columns([3, 2, 2, 1, 2])
                                            th1.markdown("**Código**")
                                            th2.markdown("**Localização**")
                                            th3.markdown("**Medida**")
                                            th4.markdown("**Saldo**")
                                            th5.markdown("**Enviar agora**")
                                            st.markdown("<hr style='margin:4px 0;'>", unsafe_allow_html=True)

                                            for _, peca in df_pecas_parc.iterrows():
                                                saldo_peca = int(peca.get('saldo', 0))
                                                if saldo_peca <= 0:
                                                    pc1, pc2, pc3, pc4, pc5 = st.columns([3, 2, 2, 1, 2])
                                                    pc1.markdown(f"~~{peca['codigo']}~~")
                                                    pc5.markdown("✅ enviado")
                                                    continue
                                                pc1, pc2, pc3, pc4, pc5 = st.columns([3, 2, 2, 1, 2])
                                                pc1.markdown(f"**{peca['codigo']}**")
                                                pc2.markdown(peca.get('localizacao', '—'))
                                                pc3.markdown(peca.get('medida', '—'))
                                                pc4.markdown(f"`{saldo_peca}`")
                                                with pc5:
                                                    qtd_enviar = st.number_input(
                                                        "",
                                                        min_value=0,
                                                        max_value=saldo_peca,
                                                        value=saldo_peca,
                                                        key=f"parc_{row['id']}_{peca['id']}",
                                                        label_visibility="collapsed"
                                                    )
                                                if qtd_enviar > 0:
                                                    pecas_envio.append({
                                                        "peca_id":    int(peca['id']),
                                                        "codigo":     peca['codigo'],
                                                        "qtd_enviar": qtd_enviar,
                                                        "saldo_atual": saldo_peca
                                                    })

                                            st.markdown("---")
                                            bp1, bp2, _ = st.columns([2, 2, 4])
                                            with bp1:
                                                if st.button("🟠 Confirmar Envio Parcial", key=f"conf_parc_{row['id']}", type="primary", use_container_width=True):
                                                    if not pecas_envio:
                                                        st.error("Selecione pelo menos uma peça.")
                                                    else:
                                                        conn = conectar_banco()
                                                        try:
                                                            cursor = conn.cursor()
                                                            todas_zeradas = True
                                                            for pe in pecas_envio:
                                                                novo_saldo = pe['saldo_atual'] - pe['qtd_enviar']
                                                                if novo_saldo > 0:
                                                                    todas_zeradas = False
                                                                cursor.execute("""
                                                                    UPDATE op_pecas
                                                                    SET qtd_enviada = qtd_enviada + %s,
                                                                        saldo = saldo - %s
                                                                    WHERE id=%s
                                                                """, (pe['qtd_enviar'], pe['qtd_enviar'], pe['peca_id']))
                                                            novo_status = 'Concluido' if todas_zeradas else 'Parcialmente Concluido'
                                                            cursor.execute(
                                                                "UPDATE itens_detalhado SET Status_Item=%s WHERE id=%s",
                                                                (novo_status, int(row['id']))
                                                            )
                                                            conn.commit()
                                                        except Exception as e:
                                                            conn.rollback()
                                                            st.error(f"Erro: {e}")
                                                        finally:
                                                            liberar_conexao(conn)
                                                        # Limpa TODOS os caches explicitamente
                                                        carregar_micro.clear()
                                                        carregar_macro.clear()
                                                        carregar_pecas_lote.clear()
                                                        carregar_todas_pecas_obra.clear()
                                                        carregar_fila_logistica.clear()
                                                        enviar_para_logistica(row, limite_desp if prazo_valido(limite_desp) else pd.NaT)
                                                        st.session_state[f"modal_pronto_{row['id']}"] = False
                                                        emoji_t = "✅" if todas_zeradas else "🟠"
                                                        st.toast(f"{emoji_t} {row['Cod_Lote']} — {'Concluido!' if todas_zeradas else 'Envio parcial registrado!'}")
                                                        time.sleep(0.5)
                                                        st.rerun()
                                            with bp2:
                                                if st.button("Cancelar", key=f"cancel_parc2_{row['id']}", use_container_width=True):
                                                    st.session_state[f"modal_pronto_{row['id']}"] = False
                                                    st.rerun()

                            # ── MODAL PARADA ───────────────────────────────────
                            if st.session_state.get(f"modal_parada_{row['id']}", False):
                                with st.container(border=True):
                                    st.markdown(f"#### ⛔ Registrar Parada — `{row['Cod_Lote']}` | {row['Obra_Vinculada']}")
                                    motivo_inp = st.text_input(
                                        "Motivo da parada:",
                                        placeholder="Ex: Falta de material, aguardando chapa...",
                                        key=f"motivo_inp_{row['id']}"
                                    )
                                    pb1, pb2, _ = st.columns([2, 2, 4])
                                    with pb1:
                                        if st.button("⛔ Confirmar Parada", key=f"conf_parada_{row['id']}", type="primary", use_container_width=True):
                                            if not motivo_inp.strip():
                                                st.error("Digite o motivo da parada.")
                                            else:
                                                salvar_parada_op(row['id'], True, motivo_inp.strip(), st.session_state.usuario_nome)
                                                st.session_state[f"modal_parada_{row['id']}"] = False
                                                st.toast(f"⛔ Parada registrada em {row['Cod_Lote']}!")
                                                time.sleep(0.5)
                                                st.rerun()
                                    with pb2:
                                        if st.button("Cancelar", key=f"cancel_parada_{row['id']}", use_container_width=True):
                                            st.session_state[f"modal_parada_{row['id']}"] = False
                                            st.rerun()

                else:
                    st.info("Nenhuma OP liberada para este filtro ainda.")
            else:
                st.info("Nenhum lote liberado no sistema ainda.")

    # ==================================================
    # PAINEL TV — ACM
    # ==================================================
    elif nome_aba == "Painel TV — ACM":
        with aba_objeto:
            st.markdown('<div class="page-header"><div class="page-header-left"><h2>Painel TV — ACM</h2><p>Visão de chão de fábrica para exibição em televisão</p></div><span class="page-icon">📺</span></div>', unsafe_allow_html=True)
            if 'tv_last_refresh' not in st.session_state:
                st.session_state.tv_last_refresh = time.time()
            agora   = time.time()
            elapsed = agora - st.session_state.tv_last_refresh
            segundos_restantes = max(0, int(30 - elapsed))

            st.markdown(f"""
            <div style='background: linear-gradient(135deg, #0F172A 0%, #1E3A8A 100%);
                        padding: 18px 28px; border-radius: 12px; display: flex;
                        align-items: center; justify-content: space-between; margin-bottom: 24px;'>
                <div>
                    <span style='color:#FFFFFF;font-size:26px;font-weight:800;letter-spacing:-0.03em;'>
                        Passold — Painel de Produção ACM
                    </span><br>
                    <span style='color:#93C5FD;font-size:13px;'>
                        {datetime.now(FUSO_BR).strftime('%d/%m/%Y  %H:%M')}
                    </span>
                </div>
                <div style='text-align:right;'>
                    <span style='color:#94A3B8;font-size:12px;'>Atualiza em</span><br>
                    <span style='color:#FCD34D;font-size:22px;font-weight:700;'>{segundos_restantes}s</span>
                </div>
            </div>
            """, unsafe_allow_html=True)

            df_tv = carregar_micro()
            if not df_tv.empty:
                df_tv = df_tv[df_tv['Escopo'] == 'ACM']
            if df_tv.empty:
                st.markdown("<div style='text-align:center;padding:60px;color:#94A3B8;font-size:20px;'>Nenhum lote cadastrado.</div>", unsafe_allow_html=True)
            else:
                def urgencia(row):
                    if row.get('Status_Item') == 'Concluido': return 'concluido'
                    prazo = row.get('Data_Limite_Obra')
                    if not prazo_valido(prazo): return 'sem_prazo'
                    dias = (pd.to_datetime(prazo).normalize() - hoje_projeto()).days
                    if dias < 0:   return 'vencido'
                    if dias <= 3:  return 'critico'
                    if dias <= 7:  return 'atencao'
                    return 'ok'

                df_tv['_urgencia'] = df_tv.apply(urgencia, axis=1)
                df_tv['_dias_restantes'] = df_tv['Data_Limite_Obra'].apply(
                    lambda x: (pd.to_datetime(x).normalize() - hoje_projeto()).days if prazo_valido(x) else 9999
                )
                df_tv = df_tv.sort_values('_dias_restantes')

                STATUS_EMOJI = {
                    'Pendente':                ('⏳', '#64748B', '#F1F5F9'),
                    'Liberado para Fabrica':   ('🔧', '#1D4ED8', '#EFF6FF'),
                    'Parcialmente Concluido':  ('🟠', '#D97706', '#FFFBEB'),
                    'Concluido':               ('✅', '#15803D', '#F0FDF4'),
                }
                URG_CONFIG = {
                    'vencido':   {'border': '#DC2626', 'bg': '#FEF2F2', 'tag': '🔴 VENCIDO',  'tag_color': '#DC2626'},
                    'critico':   {'border': '#1A56DB', 'bg': '#EFF6FF', 'tag': '🔵 URGENTE',  'tag_color': '#1A56DB'},
                    'atencao':   {'border': '#D97706', 'bg': '#FFFBEB', 'tag': '🟡 ATENÇÃO',  'tag_color': '#D97706'},
                    'ok':        {'border': '#059669', 'bg': '#F0FDF4', 'tag': '🟢 NO PRAZO', 'tag_color': '#059669'},
                    'sem_prazo': {'border': '#94A3B8', 'bg': '#F8FAFC', 'tag': '⚪ SEM PRAZO','tag_color': '#94A3B8'},
                    'concluido': {'border': '#15803D', 'bg': '#F0FDF4', 'tag': '✅ CONCLUÍDO','tag_color': '#15803D'},
                }

                criticos = df_tv[df_tv['_urgencia'].isin(['vencido', 'critico'])]
                if not criticos.empty:
                    st.markdown(f"""
                    <div style='background:#FEF2F2;border:1px solid #FECACA;border-radius:8px;padding:10px 18px;margin-bottom:16px;'>
                        <span style='color:#DC2626;font-weight:700;font-size:15px;'>⚠️ {len(criticos)} lote(s) crítico(s) ou vencido(s)</span>
                    </div>
                    """, unsafe_allow_html=True)

                producao  = df_tv[df_tv['Status_Item'].isin(['Liberado para Fabrica', 'Parcialmente Concluido'])].sort_values('_dias_restantes')
                pendentes = df_tv[df_tv['Status_Item'] == 'Pendente'].sort_values('_dias_restantes')
                concluidos = df_tv[df_tv['Status_Item'] == 'Concluido'].sort_values('Data_Limite_Obra', ascending=False)

                def _card_lote(row, key_prefix):
                    urg  = row['_urgencia']
                    cfg  = URG_CONFIG[urg]
                    dias = row['_dias_restantes']
                    dias_txt  = f"Vencido há {abs(dias)}d" if dias < 0 else (f"Faltam {dias} dia(s)" if dias < 9999 else "Sem prazo")
                    em, ec, _ = STATUS_EMOJI.get(row['Status_Item'], ('❓', '#64748B', '#F8FAFC'))
                    prazo_fmt = pd.to_datetime(row['Data_Limite_Obra']).strftime('%d/%m/%Y') if prazo_valido(row['Data_Limite_Obra']) else '—'
                    op_txt    = row['Num_OP'] if row.get('Num_OP') else 'S/ OP'
                    arqs_tv   = carregar_arquivos_op(int(row['id']))
                    clipe_badge = f"<div style='margin-top:6px;font-size:10px;color:#475569;'>📎 {len(arqs_tv)} arquivo(s)</div>" if arqs_tv else ""
                    st.markdown(f"""
                    <div style='border:1.5px solid {cfg["border"]};background:{cfg["bg"]};border-radius:8px;padding:10px 12px;margin-bottom:8px;box-shadow:0 1px 4px rgba(0,0,0,0.06);'>
                        <div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:6px;'>
                            <span style='font-size:10px;font-weight:700;color:{cfg["tag_color"]};border:1px solid {cfg["border"]};padding:1px 6px;border-radius:4px;'>{cfg["tag"]}</span>
                            <span style='font-size:10px;color:#64748B;font-weight:600;'>{dias_txt}</span>
                        </div>
                        <div style='font-size:14px;font-weight:800;color:#0F172A;margin-bottom:2px;'>{row["Obra_Vinculada"]}</div>
                        <div style='font-size:11px;color:#475569;margin-bottom:8px;'>{row["Tipo_Material"]} · {row["Romaneio_Chapas"] or "—"}</div>
                        <div style='display:grid;grid-template-columns:1fr 1fr;gap:5px;'>
                            <div style='background:white;border-radius:5px;padding:5px 7px;'><div style='font-size:8px;color:#94A3B8;text-transform:uppercase;'>OP</div><div style='font-size:12px;font-weight:700;color:#1E293B;'>{op_txt}</div></div>
                            <div style='background:white;border-radius:5px;padding:5px 7px;'><div style='font-size:8px;color:#94A3B8;text-transform:uppercase;'>M²</div><div style='font-size:12px;font-weight:700;color:#1E293B;'>{row["M2_Item"]:.2f}</div></div>
                        </div>
                        <div style='margin-top:6px;font-size:10px;font-weight:700;color:{cfg["tag_color"]};'>Prazo: {prazo_fmt}</div>
                        {clipe_badge}
                    </div>
                    """, unsafe_allow_html=True)
                    if arqs_tv:
                        with st.expander("📂 Ver arquivos", expanded=False):
                            for arq in arqs_tv:
                                arq_id, arq_nome, arq_tipo, _, _ = arq
                                conteudo_arq = carregar_conteudo_arquivo(arq_id)
                                if conteudo_arq:
                                    _, _, bytes_arq = conteudo_arq
                                    st.download_button(
                                        f"⬇️ {arq_nome}", data=bytes(bytes_arq),
                                        file_name=arq_nome, mime=arq_tipo or "application/octet-stream",
                                        key=f"tv_{key_prefix}_dl_{arq_id}"
                                    )

                col_prod, col_pend, col_conc = st.columns(3)
                colunas_tv = [
                    (col_prod, "prod", "#EFF6FF", "#1D4ED8", "🔧 EM PRODUÇÃO", producao,  "Nenhum lote em produção."),
                    (col_pend, "pend", "#F1F5F9", "#475569", "⏳ PENDENTE",     pendentes, "Nenhum lote pendente."),
                    (col_conc, "conc", "#F0FDF4", "#15803D", "✅ CONCLUÍDO",    concluidos, "Nenhum lote concluído."),
                ]
                for col, prefix, bg, fg, titulo, df_col, msg_vazio in colunas_tv:
                    with col:
                        with st.container(border=True):
                            st.markdown(f"""<div style='background:{bg};border-top:3px solid {fg};border-radius:6px;padding:6px 12px;margin-bottom:10px;text-align:center;'>
                                <span style='color:{fg};font-weight:800;font-size:13px;'>{titulo}</span><br>
                                <span style='color:{fg};font-size:19px;font-weight:800;'>{len(df_col)}</span>
                            </div>""", unsafe_allow_html=True)
                            if df_col.empty:
                                st.markdown(f"<div style='text-align:center;color:#94A3B8;padding:16px;font-size:12px;'>{msg_vazio}</div>", unsafe_allow_html=True)
                            else:
                                for _, row in df_col.iterrows():
                                    _card_lote(row, f"{prefix}_{row['id']}")

            progress_val = (30 - segundos_restantes) / 30
            st.markdown("<br>", unsafe_allow_html=True)
            st.progress(progress_val, text=f"Próxima atualização em {segundos_restantes}s")
            if segundos_restantes == 0:
                st.session_state.tv_last_refresh = time.time()
                st.rerun()

    # ==================================================
    # PAINEL DA PRODUCAO — ESQUADRIAS
    # ==================================================
    elif nome_aba == "Painel da Producao - Esquadrias":
        with aba_objeto:
            st.markdown('<div class="page-header"><div class="page-header-left"><h2>Painel da Produção — Esquadrias</h2><p>Acompanhamento em tempo real dos lotes e OPs de Esquadrias</p></div><span class="page-icon">📋</span></div>', unsafe_allow_html=True)
            obras_esq = ["Todas as obras"] + (
                list(df_banco_micro['Obra_Vinculada'].dropna().unique()) if not df_banco_micro.empty else []
            )
            obra_esq = st.selectbox("Filtrar por obra:", obras_esq, key="sb_obra_esq")

            st.markdown("### Previsão de Entrada em Produção")
            st.caption("Lotes planejados ainda não liberados oficialmente.")
            df_prev_esq = df_banco_micro[df_banco_micro['Escopo'] == 'Esquadria-Vidro'].copy() if not df_banco_micro.empty else pd.DataFrame()
            if not df_prev_esq.empty:
                if obra_esq != "Todas as obras":
                    df_prev_esq = df_prev_esq[df_prev_esq['Obra_Vinculada'] == obra_esq]
            df_prev_pend_esq = df_prev_esq[df_prev_esq['Status_Item'] == 'Pendente'].copy() if not df_prev_esq.empty else pd.DataFrame()
            if df_prev_pend_esq.empty:
                st.info("Nenhuma previsão pendente de Esquadrias.")
            else:
                blocos_semanais(df_prev_pend_esq)

            st.markdown("---")
            st.markdown("### Calendário de Produção — OPs Liberadas")
            st.caption("Apenas lotes liberados oficialmente na aba 'Liberar OPs da Semana'.")
            if not df_banco_micro.empty:
                df_base_esq = df_banco_micro[
                    df_banco_micro['Status_Item'].isin(["Liberado para Fabrica", "Parcialmente Concluido"]) &
                    (df_banco_micro['Escopo'] == 'Esquadria-Vidro')
                ].copy()
                if obra_esq != "Todas as obras":
                    df_base_esq = df_base_esq[df_base_esq['Obra_Vinculada'] == obra_esq]
                if not df_base_esq.empty:
                    registros_exp_esq = []
                    for _, row in df_base_esq.iterrows():
                        dt_ini = pd.to_datetime(row['Data_Producao_Programada']).date()
                        dt_fim = pd.to_datetime(row['Data_Limite_Obra']).date()
                        ini_ult = max((pd.to_datetime(dt_fim) - timedelta(days=6)).date(), dt_ini)
                        dia = dt_ini
                        while dia <= dt_fim:
                            if dia.weekday() < 5:  # sem producao aos finais de semana
                                r = row.to_dict(); r['_dia'] = dia; r['_pode_concluir'] = (dia >= ini_ult)
                                registros_exp_esq.append(r)
                            dia += timedelta(days=1)
                    df_exp_esq = pd.DataFrame(registros_exp_esq)

                    if "esq_mes" not in st.session_state:
                        st.session_state.esq_mes = hoje_projeto().month
                    if "esq_ano" not in st.session_state:
                        st.session_state.esq_ano = hoje_projeto().year

                    c1, c2, c3 = st.columns([1, 2, 1])
                    with c1:
                        if st.button("Mes Anterior", use_container_width=True, key="esq_ant"):
                            st.session_state.esq_mes -= 1
                            if st.session_state.esq_mes == 0:
                                st.session_state.esq_mes = 12; st.session_state.esq_ano -= 1
                            st.rerun()
                    with c2:
                        nomes_meses = ["","Janeiro","Fevereiro","Marco","Abril","Maio","Junho",
                                       "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
                        st.markdown(
                            f"<h3 style='text-align:center;color:#1E3A8A;margin:0;'>"
                            f"{nomes_meses[st.session_state.esq_mes]} / {st.session_state.esq_ano}</h3>",
                            unsafe_allow_html=True)
                    with c3:
                        if st.button("Proximo Mes", use_container_width=True, key="esq_prox"):
                            st.session_state.esq_mes += 1
                            if st.session_state.esq_mes == 13:
                                st.session_state.esq_mes = 1; st.session_state.esq_ano += 1
                            st.rerun()

                    st.markdown("---")
                    cal_esq = py_calendar.Calendar(firstweekday=6)
                    semanas_esq = cal_esq.monthdatescalendar(st.session_state.esq_ano, st.session_state.esq_mes)
                    cols_h_esq = st.columns(7)
                    for i, nome in enumerate(["Dom","Seg","Ter","Qua","Qui","Sex","Sab"]):
                        cols_h_esq[i].markdown(
                            f"<div style='text-align:center;font-weight:600;color:#475569;padding:4px 0;font-size:13px;'>{nome}</div>",
                            unsafe_allow_html=True)
                    for semana in semanas_esq:
                        cols = st.columns(7)
                        for i, data_dia in enumerate(semana):
                            with cols[i]:
                                if data_dia.month == st.session_state.esq_mes:
                                    lotes_dia = df_exp_esq[df_exp_esq['_dia'] == data_dia]
                                    n_lotes = lotes_dia['Cod_Lote'].nunique() if not lotes_dia.empty else 0
                                    eh_hoje = (data_dia == hoje_projeto().date())
                                    if n_lotes > 0:
                                        obras_d = lotes_dia['Obra_Vinculada'].unique()
                                        label_o = obras_d[0] if len(obras_d) == 1 else f"{len(obras_d)} obras"
                                        if st.button(f"{data_dia.day}  |  {n_lotes} lote(s)\n{label_o}",
                                                     key=f"esq_btn_{data_dia}", use_container_width=True):
                                            st.session_state.esq_dia_clicado = data_dia
                                    else:
                                        css = "cal-day-today" if eh_hoje else "cal-day-empty"
                                        st.markdown(
                                            f"<div class='{css}'><span style='color:#94A3B8;font-size:15px;'>{data_dia.day}</span>"
                                            f"<br><span style='color:#CBD5E1;font-size:10px;'>—</span></div>",
                                            unsafe_allow_html=True)
                                else:
                                    st.markdown('<div style="height:70px;"></div>', unsafe_allow_html=True)

                    st.markdown("---")
                    if "esq_dia_clicado" not in st.session_state:
                        st.session_state.esq_dia_clicado = hoje_projeto().date()
                    dia_sel_esq = st.session_state.esq_dia_clicado
                    st.subheader(f"Lotes em producao — {dia_sel_esq.strftime('%d/%m/%Y')}")
                    lotes_sel_esq = df_exp_esq[df_exp_esq['_dia'] == dia_sel_esq].drop_duplicates(subset=['id'])

                    if lotes_sel_esq.empty:
                        st.info("Clique em um dia com lotes no calendário acima.")
                    else:
                        ke1, ke2, ke3 = st.columns(3)
                        ke1.metric("Lotes em producao", lotes_sel_esq['Cod_Lote'].nunique())
                        ke2.metric("Total unidades", int(lotes_sel_esq['Qtd_Caixas'].sum()))
                        ke3.metric("Total kg", f"{lotes_sel_esq['Peso_Kg'].sum():.2f}" if 'Peso_Kg' in lotes_sel_esq.columns else "—")
                        st.markdown("---")

                        for _, row in lotes_sel_esq.iterrows():
                            eh_parcial = row.get('Status_Item', '') == 'Parcialmente Concluido'
                            pode_concluir = bool(row.get('_pode_concluir', False)) or eh_parcial
                            dt_i = pd.to_datetime(row['Data_Producao_Programada']).strftime('%d/%m/%Y')
                            dt_f = pd.to_datetime(row['Data_Limite_Obra']).strftime('%d/%m/%Y')
                            border_color = "#D97706" if eh_parcial else ("#1A56DB" if pode_concluir else "#3B82F6")
                            bg_color = "#FFFBEB" if eh_parcial else ("#FFF7ED" if pode_concluir else "#F8FAFC")
                            st.markdown(
                                f"<div style='border-left:4px solid {border_color};background:{bg_color};"
                                f"padding:12px 16px;border-radius:6px;margin-bottom:4px;'></div>",
                                unsafe_allow_html=True)
                            with st.container(border=True):
                                cd, ca = st.columns([4, 1])
                                with cd:
                                    st.markdown(
                                        f'<span class="badge-obra">{row["Obra_Vinculada"]}</span>&nbsp;'
                                        f'<span class="badge-edt">{row["EDT_Vinculado"]}</span>&nbsp;'
                                        f'<span class="badge-lote">{row["Cod_Lote"]}</span>',
                                        unsafe_allow_html=True)
                                    kg_val = row.get('Peso_Kg', 0.0) or 0.0
                                    st.markdown(f"**{row['Tipo_Material']}** &nbsp;|&nbsp; `{int(row['Qtd_Caixas'])} un` — **{kg_val:.2f} kg**")
                                    st.caption(f"Periodo: {dt_i} a {dt_f} &nbsp;|&nbsp; {row['Romaneio_Chapas']}")
                                    op_txt = row['Num_OP'] if row.get('Num_OP') else "Aguardando OP"
                                    st.caption(f"OP: {op_txt} &nbsp;|&nbsp; {row.get('Fase_Produtiva', '—')}")
                                    em_parada_esq = bool(row.get('Em_Parada', False))
                                    _motivo_esq_raw = row.get('Motivo_Parada')
                                    motivo_esq    = html_escape(str(_motivo_esq_raw)) if pd.notna(_motivo_esq_raw) else ''
                                    if em_parada_esq:
                                        st.markdown(f"<span style='color:#DC2626;font-size:12px;font-weight:700;'>⛔ EM PARADA — {motivo_esq}</span>", unsafe_allow_html=True)
                                    elif eh_parcial:
                                        st.markdown("<span style='color:#D97706;font-size:12px;font-weight:700;'>Envio parcial registrado — ainda há peças pendentes</span>", unsafe_allow_html=True)
                                    elif pode_concluir:
                                        st.markdown("<span style='color:#1A56DB;font-size:12px;font-weight:600;'>Ultima semana — liberado para concluir</span>", unsafe_allow_html=True)
                                    else:
                                        dias_rest = (pd.to_datetime(row['Data_Limite_Obra']).date() - dia_sel_esq).days
                                        st.markdown(f"<span style='color:#3B82F6;font-size:12px;'>Em producao — {dias_rest} dias ate o prazo</span>", unsafe_allow_html=True)
                                with ca:
                                    if setor in ["Producao", "Master"]:
                                        if em_parada_esq:
                                            if st.button("▶ Retomar", key=f"esq_retomar_{row['id']}", use_container_width=True):
                                                salvar_parada_op(row['id'], False, '', st.session_state.usuario_nome)
                                                st.toast("OP retomada!")
                                                st.rerun()
                                        else:
                                            if st.button("⏸ Parada", key=f"esq_parada_{row['id']}", use_container_width=True):
                                                st.session_state[f"esq_modal_parada_{row['id']}"] = not st.session_state.get(f"esq_modal_parada_{row['id']}", False)
                                                st.rerun()
                                    if pode_concluir:
                                        st.write("")
                                        if st.button("✅ Pronto", key=f"esq_baixa_{row['id']}", type="primary", use_container_width=True):
                                            st.session_state[f"esq_modal_{row['id']}"] = not st.session_state.get(f"esq_modal_{row['id']}", False)
                                            st.rerun()
                                    elif setor not in ["Producao", "Master"]:
                                        st.markdown("<div style='text-align:center;color:#94A3B8;font-size:12px;padding:8px;'>Em producao</div>", unsafe_allow_html=True)

                            if st.session_state.get(f"esq_modal_{row['id']}", False):
                                with st.container(border=True):
                                    st.markdown(f"#### `{row['Cod_Lote']}` — {row['Obra_Vinculada']} — Tipo de Envio")
                                    limite_desp_esq = None
                                    if not df_banco_macro.empty:
                                        fr = df_banco_macro[df_banco_macro['EDT'] == row['EDT_Vinculado']]
                                        if not fr.empty:
                                            limite_desp_esq = fr.iloc[0].get('Data_Limite_Despacho')
                                    me1, me2 = st.columns([2, 3])
                                    with me1:
                                        tipo_envio_esq = st.radio(
                                            "Este envio é:",
                                            ["Envio Total", "Envio Parcial"],
                                            key=f"esq_tipo_envio_{row['id']}",
                                            horizontal=True)
                                    if tipo_envio_esq == "Envio Total":
                                        with me2:
                                            st.info("Todas as peças serão marcadas como concluídas.")
                                        bt1, bt2, _ = st.columns([2, 2, 4])
                                        with bt1:
                                            if st.button("✅ Confirmar Total", key=f"esq_conf_total_{row['id']}", type="primary", use_container_width=True):
                                                conn = conectar_banco()
                                                try:
                                                    cursor = conn.cursor()
                                                    cursor.execute("UPDATE itens_detalhado SET Status_Item='Concluido' WHERE id=%s", (row['id'],))
                                                    cursor.execute("UPDATE op_pecas SET qtd_enviada=qtd_total, saldo=0 WHERE lote_id=%s", (row['id'],))
                                                    conn.commit()
                                                except Exception as e:
                                                    conn.rollback(); st.error(f"Erro: {e}")
                                                finally:
                                                    liberar_conexao(conn)
                                                carregar_micro.clear(); carregar_macro.clear(); carregar_fila_logistica.clear()
                                                carregar_pecas_lote.clear(); carregar_todas_pecas_obra.clear()
                                                enviar_para_logistica(row, limite_desp_esq if prazo_valido(limite_desp_esq) else pd.NaT)
                                                st.session_state[f"esq_modal_{row['id']}"] = False
                                                st.toast(f"✅ {row['Cod_Lote']} concluido!")
                                                time.sleep(0.5); st.rerun()
                                        with bt2:
                                            if st.button("Cancelar", key=f"esq_cancel_total_{row['id']}", use_container_width=True):
                                                st.session_state[f"esq_modal_{row['id']}"] = False; st.rerun()
                                    else:
                                        df_pecas_esq = carregar_pecas_lote(int(row['id']))
                                        if df_pecas_esq.empty:
                                            st.warning("Nenhuma peça lançada para este lote.")
                                            if st.button("Fechar", key=f"esq_cancel_parc_{row['id']}"):
                                                st.session_state[f"esq_modal_{row['id']}"] = False; st.rerun()
                                        else:
                                            st.caption("Informe quantas unidades de cada peça estão prontas:")
                                            st.markdown("---")
                                            pecas_envio_esq = []
                                            th1,th2,th3,th4,th5 = st.columns([3,2,2,1,2])
                                            th1.markdown("**Código**"); th2.markdown("**Localização**")
                                            th3.markdown("**Medida**"); th4.markdown("**Saldo**"); th5.markdown("**Enviar agora**")
                                            st.markdown("<hr style='margin:4px 0;'>", unsafe_allow_html=True)
                                            for _, peca in df_pecas_esq.iterrows():
                                                saldo_peca_esq = int(peca.get('saldo', 0))
                                                if saldo_peca_esq <= 0:
                                                    pc1, pc2, pc3, pc4, pc5 = st.columns([3, 2, 2, 1, 2])
                                                    pc1.markdown(f"~~{peca['codigo']}~~")
                                                    pc5.markdown("✅ enviado")
                                                    continue
                                                p1,p2,p3,p4,p5 = st.columns([3,2,2,1,2])
                                                p1.write(peca.get('codigo','—')); p2.write(peca.get('localizacao','—'))
                                                p3.write(peca.get('medida','—')); p4.write(saldo_peca_esq)
                                                qtd_env = p5.number_input("", min_value=0,
                                                    max_value=saldo_peca_esq,
                                                    value=saldo_peca_esq,
                                                    key=f"esq_env_{row['id']}_{peca['id']}", label_visibility="collapsed")
                                                if qtd_env > 0:
                                                    pecas_envio_esq.append({"peca_id": int(peca['id']), "qtd": qtd_env, "saldo": saldo_peca_esq})
                                            st.markdown("---")
                                            b1,b2,_ = st.columns([2,2,4])
                                            with b1:
                                                if st.button("✅ Confirmar Parcial", key=f"esq_conf_parc_{row['id']}", type="primary", use_container_width=True, disabled=not pecas_envio_esq):
                                                    conn = conectar_banco()
                                                    try:
                                                        cursor = conn.cursor()
                                                        todas_concluidas = True
                                                        for p in pecas_envio_esq:
                                                            novo_saldo_esq = p['saldo'] - p['qtd']
                                                            if novo_saldo_esq > 0:
                                                                todas_concluidas = False
                                                            cursor.execute("""
                                                                UPDATE op_pecas
                                                                SET qtd_enviada = qtd_enviada + %s,
                                                                    saldo = saldo - %s
                                                                WHERE id=%s
                                                            """, (p['qtd'], p['qtd'], p['peca_id']))
                                                        novo_status = 'Concluido' if todas_concluidas else 'Parcialmente Concluido'
                                                        cursor.execute("UPDATE itens_detalhado SET Status_Item=%s WHERE id=%s", (novo_status, row['id']))
                                                        conn.commit()
                                                    except Exception as e:
                                                        conn.rollback(); st.error(f"Erro: {e}")
                                                    finally:
                                                        liberar_conexao(conn)
                                                    carregar_micro.clear(); carregar_macro.clear(); carregar_fila_logistica.clear()
                                                    carregar_pecas_lote.clear(); carregar_todas_pecas_obra.clear()
                                                    enviar_para_logistica(row, limite_desp_esq if prazo_valido(limite_desp_esq) else pd.NaT)
                                                    st.session_state[f"esq_modal_{row['id']}"] = False
                                                    st.toast(f"Envio parcial de {row['Cod_Lote']} registrado!"); time.sleep(0.5); st.rerun()
                                            with b2:
                                                if st.button("Cancelar", key=f"esq_cancel_parc2_{row['id']}", use_container_width=True):
                                                    st.session_state[f"esq_modal_{row['id']}"] = False; st.rerun()

                            # ── MODAL PARADA ESQUADRIAS ────────────────────────
                            if st.session_state.get(f"esq_modal_parada_{row['id']}", False):
                                with st.container(border=True):
                                    st.markdown(f"#### ⛔ Registrar Parada — `{row['Cod_Lote']}` | {row['Obra_Vinculada']}")
                                    motivo_esq_inp = st.text_input(
                                        "Motivo da parada:",
                                        placeholder="Ex: Falta de material, aguardando perfil...",
                                        key=f"esq_motivo_inp_{row['id']}"
                                    )
                                    epb1, epb2, _ = st.columns([2, 2, 4])
                                    with epb1:
                                        if st.button("⛔ Confirmar Parada", key=f"esq_conf_parada_{row['id']}", type="primary", use_container_width=True):
                                            if not motivo_esq_inp.strip():
                                                st.error("Digite o motivo da parada.")
                                            else:
                                                salvar_parada_op(row['id'], True, motivo_esq_inp.strip(), st.session_state.usuario_nome)
                                                st.session_state[f"esq_modal_parada_{row['id']}"] = False
                                                st.toast(f"⛔ Parada registrada em {row['Cod_Lote']}!")
                                                time.sleep(0.5)
                                                st.rerun()
                                    with epb2:
                                        if st.button("Cancelar", key=f"esq_cancel_parada_{row['id']}", use_container_width=True):
                                            st.session_state[f"esq_modal_parada_{row['id']}"] = False
                                            st.rerun()

                else:
                    st.info("Nenhum lote de Esquadrias/Vidro liberado para produção.")
            else:
                st.info("Nenhum dado disponível.")

    # ==================================================
    # PAINEL TV — ESQUADRIAS
    # ==================================================
    elif nome_aba == "Painel TV — Esquadrias":
        with aba_objeto:
            st.markdown('<div class="page-header"><div class="page-header-left"><h2>Painel TV — Esquadrias</h2><p>Visão de chão de fábrica para exibição em televisão</p></div><span class="page-icon">📺</span></div>', unsafe_allow_html=True)

            if not df_banco_micro.empty:
                df_tv_esq = df_banco_micro[
                    df_banco_micro['Status_Item'].isin(["Liberado para Fabrica","Parcialmente Concluido"]) &
                    (df_banco_micro['Escopo'] == 'Esquadria-Vidro')
                ].copy()
            else:
                df_tv_esq = pd.DataFrame()

            hoje_ts = pd.Timestamp.now().normalize()
            if not df_tv_esq.empty:
                df_tv_esq['_limite'] = pd.to_datetime(df_tv_esq['Data_Limite_Obra'], errors='coerce')
                df_tv_esq['_dias_rest'] = (df_tv_esq['_limite'] - hoje_ts).dt.days
                urgentes_esq = df_tv_esq[df_tv_esq['_dias_rest'] <= 5].copy()

            st.markdown(f"""
            <div style='background:linear-gradient(135deg,#0F172A 0%,#1E3A8A 100%);
                        padding:18px 28px;border-radius:12px;display:flex;
                        align-items:center;justify-content:space-between;margin-bottom:24px;'>
                <div>
                    <div style='color:#FFFFFF;font-size:2rem;font-weight:800;letter-spacing:-0.03em;'>
                        Producao — Esquadrias & Vidro
                    </div>
                    <div style='color:#94A3B8;font-size:1rem;margin-top:4px;'>
                        {hoje_ts.strftime('%d/%m/%Y')} &nbsp;|&nbsp; {len(df_tv_esq)} lote(s) ativos
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)

            if not df_tv_esq.empty:
                tv_k1, tv_k2, tv_k3, tv_k4 = st.columns(4)
                tv_k1.metric("Lotes ativos", len(df_tv_esq))
                tv_k2.metric("Total unidades", int(df_tv_esq['Qtd_Caixas'].sum()))
                kg_col = df_tv_esq['Peso_Kg'].sum() if 'Peso_Kg' in df_tv_esq.columns else 0
                tv_k3.metric("Total kg", f"{kg_col:,.2f}")
                tv_k4.metric("Criticos (≤5 dias)", len(urgentes_esq) if not df_tv_esq.empty else 0)
                st.markdown("---")

                if not urgentes_esq.empty:
                    st.markdown(f"""
                    <div style='background:#FEF2F2;border:1px solid #FECACA;border-radius:8px;padding:10px 18px;margin-bottom:16px;'>
                        <span style='color:#DC2626;font-weight:700;font-size:15px;'>⚠️ {len(urgentes_esq)} lote(s) crítico(s) ou vencido(s)</span>
                    </div>""", unsafe_allow_html=True)

                cols_tv_esq = st.columns(3)
                for i, (_, row) in enumerate(df_tv_esq.sort_values('_dias_rest').iterrows()):
                    dias = row['_dias_rest']
                    if dias < 0:    cfg = {"border":"#DC2626","bg":"#FEF2F2","tag":"VENCIDO","tag_color":"#DC2626"}
                    elif dias <= 3: cfg = {"border":"#DC2626","bg":"#FEF2F2","tag":"CRITICO","tag_color":"#DC2626"}
                    elif dias <= 5: cfg = {"border":"#D97706","bg":"#FFFBEB","tag":"URGENTE","tag_color":"#D97706"}
                    else:           cfg = {"border":"#3B82F6","bg":"#EFF6FF","tag":"OK","tag_color":"#3B82F6"}
                    kg_v = row.get('Peso_Kg', 0.0) or 0.0
                    with cols_tv_esq[i % 3]:
                        st.markdown(f"""
                        <div style='border:2px solid {cfg["border"]};background:{cfg["bg"]};border-radius:10px;
                                    padding:18px 20px;margin-bottom:12px;box-shadow:0 4px 12px rgba(0,0,0,0.10);'>
                            <div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:10px;'>
                                <span style='font-size:11px;font-weight:700;color:{cfg["tag_color"]};border:1px solid {cfg["border"]};
                                             padding:2px 8px;border-radius:4px;'>{cfg["tag"]}</span>
                                <span style='font-size:11px;color:#64748B;'>{int(dias)} dias</span>
                            </div>
                            <div style='font-size:1.1rem;font-weight:700;color:#0F172A;margin-bottom:4px;'>{row['Cod_Lote']}</div>
                            <div style='font-size:0.85rem;color:#475569;margin-bottom:8px;'>{row['Obra_Vinculada']} — {row['Tipo_Material']}</div>
                            <div style='display:flex;gap:12px;'>
                                <span style='background:#F1F5F9;padding:3px 8px;border-radius:4px;font-size:12px;font-weight:600;'>
                                    {int(row['Qtd_Caixas'])} un
                                </span>
                                <span style='background:#F1F5F9;padding:3px 8px;border-radius:4px;font-size:12px;font-weight:600;'>
                                    {kg_v:.2f} kg
                                </span>
                            </div>
                            <div style='margin-top:10px;font-size:11px;color:#64748B;'>
                                Limite: {pd.to_datetime(row['Data_Limite_Obra']).strftime('%d/%m/%Y')}
                            </div>
                        </div>""", unsafe_allow_html=True)

                if "tv_esq_last_refresh" not in st.session_state:
                    st.session_state.tv_esq_last_refresh = time.time()
                segundos_restantes_esq = max(0, 300 - int(time.time() - st.session_state.tv_esq_last_refresh))
                st.caption(f"Auto-refresh em {segundos_restantes_esq}s")
                if segundos_restantes_esq == 0:
                    st.session_state.tv_esq_last_refresh = time.time(); st.rerun()
            else:
                st.info("Nenhum lote de Esquadrias/Vidro em produção no momento.")

    # ==================================================
    # LIBERAR OPS
    # ==================================================
    elif nome_aba == "Liberar OPs da Semana":
        with aba_objeto:
            if setor not in ["Master", "PCP"]:
                st.error("⛔ Acesso negado.")
                st.stop()
            st.markdown('<div class="page-header"><div class="page-header-left"><h2>Ordens de Produção</h2><p>Gerencie e libere Ordens de Produção para a fábrica</p></div><span class="page-icon">🔓</span></div>', unsafe_allow_html=True)

            # ── CARD 0: LOTE SEM FRENTE (uso frequente — fica no topo) ──
            with st.expander("➕ Liberação de OP  ·  Ancoragens, prisílias e materiais de apoio", expanded=True):
                st.caption("Para ancoragens, prisilias, corte de perfil e outros materiais de apoio — não precisa de frente detalhada, mas precisa de um Projeto já cadastrado em 'Cadastrar Obra'.")
                if df_projetos.empty:
                    st.info("Cadastre uma Obra e um Projeto em 'Cadastrar Obra' antes de lançar uma OP.")
                    av_obra, av_projeto, av_escopo = None, None, None
                else:
                    av1, av2, av3 = st.columns(3)
                    with av1:
                        obras_disp = sorted(df_projetos['Obra'].dropna().unique().tolist())
                        av_obra = st.selectbox("Obra:", obras_disp, key="av_obra")
                    with av2:
                        projetos_av = sorted(df_projetos[df_projetos['Obra'] == av_obra]['Numero_Projeto'].unique().tolist())
                        av_projeto = st.selectbox("Projeto:", projetos_av, key="av_projeto") if projetos_av else None
                    with av3:
                        # Escopo vem travado do que ja foi cadastrado pra esse Obra+Projeto em
                        # 'Cadastrar Obra', quando ha uma unica frente/escopo registrada — evita
                        # escolher um escopo diferente do que a obra ja define.
                        escopos_projeto = sorted(df_banco_macro[
                            (df_banco_macro['Obra'] == av_obra) & (df_banco_macro['Numero_Projeto'] == av_projeto)
                        ]['Tipo_Escopo'].dropna().unique().tolist()) if not df_banco_macro.empty and av_projeto else []
                        if len(escopos_projeto) == 1:
                            av_escopo = escopos_projeto[0]
                            st.text_input("Tipo de Escopo:", value=av_escopo, disabled=True, key="av_escopo_travado")
                        else:
                            av_escopo = st.selectbox("Tipo de Escopo:", ["ACM", "Esquadria-Vidro", "Terceirizada"], key="av_escopo")

                    st.caption("Este item entra como **Pendente** e recebe o número da OP na liberação, em 'Liberar OPs da Semana'.")

                # Lista de itens acumulados na sessão
                if "av_itens" not in st.session_state:
                    st.session_state.av_itens = []

                st.markdown("**Itens da OP:**")
                av_desc = st.text_input("Descrição do material:", key="av_desc",
                                         placeholder="Ex: Ancoragem estrutural, Prisilia, Corte de perfil...")
                av4, av5 = st.columns(2)
                if av_escopo == "ACM":
                    with av4:
                        av_qtd_cx = st.number_input("Qtd Caixas:", min_value=0, value=1, key="av_qtd_cx")
                    with av5:
                        av_m2 = st.number_input("m²:", min_value=0.0, value=0.0, step=0.1, key="av_m2")
                    av_peso = 0.0
                    av_unidade = "cx"
                elif av_escopo == "Esquadria-Vidro":
                    with av4:
                        av_qtd_cx = st.number_input("Quantidade (un):", min_value=0, value=1, key="av_qtd_cx")
                    with av5:
                        av_peso = st.number_input("Peso (kg):", min_value=0.0, value=0.0, step=0.1, key="av_peso")
                    av_m2 = 0.0
                    av_unidade = "un"
                else:
                    with av4:
                        av_unidade = st.selectbox("Unidade:", ["un", "kg", "m", "m²", "cx", "pç"], key="av_unidade")
                    with av5:
                        av_qtd_cx = st.number_input("Quantidade:", min_value=0, value=1, key="av_qtd_cx")
                    av_m2  = 0.0
                    av_peso = 0.0

                if st.button("➕ Adicionar Item", key="btn_add_item"):
                    if not av_desc.strip():
                        st.error("Informe a descrição do material antes de adicionar.")
                    else:
                        st.session_state.av_itens.append({
                            "desc": av_desc.strip().upper(),
                            "qtd": int(av_qtd_cx),
                            "m2": float(av_m2),
                            "peso": float(av_peso),
                            "unidade": av_unidade,
                        })
                        st.rerun()

                if st.session_state.av_itens:
                    st.markdown("**Itens adicionados:**")
                    for i, item in enumerate(st.session_state.av_itens):
                        col_desc, col_qtd, col_rem = st.columns([5, 2, 1])
                        with col_desc:
                            st.write(f"{i+1}. {item['desc']}")
                        with col_qtd:
                            if item['m2'] > 0:
                                st.write(f"{item['qtd']} cx | {item['m2']} m²")
                            elif item['peso'] > 0:
                                st.write(f"{item['qtd']} {item['unidade']} | {item['peso']} kg")
                            else:
                                st.write(f"{item['qtd']} {item['unidade']}")
                        with col_rem:
                            if st.button("🗑", key=f"rem_item_{i}"):
                                st.session_state.av_itens.pop(i)
                                st.rerun()

                av6, av7 = st.columns(2)
                with av6:
                    av_dt_ini = st.date_input("Entrada em produção:", value=datetime.now().date(),
                                               format="DD/MM/YYYY", key="av_dt_ini")
                with av7:
                    av_dt_fim = st.date_input("Data limite:", value=(datetime.now() + timedelta(days=7)).date(),
                                               format="DD/MM/YYYY", key="av_dt_fim")

                av_pav     = st.text_input("Pavimentos / Destino:", key="av_pav", placeholder="Ex: Pav 10 ao 15")
                av_destino = st.radio("Destino:", ["Envio para Obra", "Uso Interno"],
                                       horizontal=True, key="av_destino")

                if st.button("💾 Cadastrar Lote", key="btn_av", type="primary", disabled=df_projetos.empty):
                    if not st.session_state.av_itens:
                        st.error("Adicione ao menos um item antes de cadastrar.")
                    elif not av_projeto:
                        st.error("Selecione o Projeto (cadastre um em 'Cadastrar Obra' se ainda não existir).")
                    else:
                        conn_av2 = conectar_banco()
                        try:
                            cursor_av2 = conn_av2.cursor()
                            for item in st.session_state.av_itens:
                                cursor_av2.execute("""
                                    INSERT INTO itens_detalhado
                                    (Obra_Vinculada, EDT_Vinculado, Cod_Lote, Num_OP, Tipo_Material,
                                     Qtd_Caixas, M2_Item, Peso_Kg, Data_Producao_Programada, Data_Limite_Obra,
                                     Data_Despacho, Romaneio_Chapas, Status_Item, Dificuldade,
                                     Fase_Produtiva, Enviado_Logistica, Escopo, Numero_Projeto, Uso_Interno)
                                    VALUES (%s,'AVULSO',%s,NULL,%s,%s,%s,%s,%s,%s,%s,%s,'Pendente',1,%s,%s,%s,%s,%s)
                                """, (
                                    av_obra,
                                    f"AVULSO-{av_projeto}",
                                    item["desc"],
                                    item["qtd"],
                                    item["m2"],
                                    item["peso"],
                                    av_dt_ini.strftime('%Y-%m-%d'),
                                    av_dt_fim.strftime('%Y-%m-%d'),
                                    av_dt_fim.strftime('%Y-%m-%d'),
                                    f"PRJ-{av_projeto} | {av_pav}",
                                    f"LOTE SEM FRENTE — {av_escopo} | {'Envio para Obra' if av_destino == 'Envio para Obra' else 'Uso Interno'}",
                                    1 if av_destino == "Envio para Obra" else 0,
                                    av_escopo,
                                    av_projeto,
                                    av_destino == "Uso Interno"
                                ))
                            conn_av2.commit()
                            _limpar_cache_geral()
                        except Exception as e:
                            conn_av2.rollback()
                            st.error(f"Erro: {e}")
                        finally:
                            liberar_conexao(conn_av2)
                        registrar_auditoria(st.session_state.usuario_nome, "LOTE_SEM_FRENTE",
                            f"Projeto {av_projeto} — {len(st.session_state.av_itens)} itens — {av_obra} — {av_destino} — Pendente")
                        st.toast(f"Lote cadastrado como Pendente — {len(st.session_state.av_itens)} item(ns). Libere em 'Liberar OPs da Semana'.")
                        st.session_state.av_itens = []
                        time.sleep(0.5)
                        st.rerun()

            st.markdown("<br>", unsafe_allow_html=True)

            # ── CARD 1: LOTES PENDENTES ────────────────────────────
            df_pend_base = df_banco_micro[df_banco_micro['Status_Item'] == 'Pendente'].copy() if not df_banco_micro.empty else pd.DataFrame()

            def _liberar_um_lote(item_id: int, numero_projeto: str):
                conn = conectar_banco()
                try:
                    cursor = conn.cursor()
                    cursor.execute(
                        "SELECT COUNT(*) FROM itens_detalhado WHERE Numero_Projeto=%s AND Num_OP IS NOT NULL",
                        (numero_projeto,)
                    )
                    seq = cursor.fetchone()[0] + 1
                    num_op = f"OP-{numero_projeto}-{str(seq).zfill(3)}"
                    cursor.execute(
                        "UPDATE itens_detalhado SET Status_Item='Liberado para Fabrica', Num_OP=%s WHERE id=%s",
                        (num_op, item_id)
                    )
                    conn.commit()
                    _limpar_cache_geral()
                    registrar_auditoria(st.session_state.usuario_nome, "LIBERAR_OPS", f"OP {num_op} liberada")
                    st.toast(f"Liberado! Número da OP: {num_op}")
                except Exception as e:
                    conn.rollback()
                    st.error(f"Erro: {e}")
                finally:
                    liberar_conexao(conn)
                time.sleep(0.5)
                st.rerun()

            with st.expander(f"📋 Lotes Pendentes  ·  {len(df_pend_base)} aguardando liberação  (acompanhamento)", expanded=False):
                if df_pend_base.empty:
                    st.markdown('<div class="empty-state"><div class="empty-icon">✅</div><h4>Tudo em dia!</h4><p>Nenhum lote esperando liberação no momento.</p></div>', unsafe_allow_html=True)
                else:
                    st.caption("Escolha a obra (opcional) e clique em Liberar no lote que quiser mandar pra fábrica.")

                    obras_disp = sorted(df_pend_base['Obra_Vinculada'].dropna().unique().tolist())
                    obra_filtro_sec1 = st.selectbox("Obra:", ["Todas as obras"] + obras_disp, key="sec1_obra")
                    df_pend = df_pend_base if obra_filtro_sec1 == "Todas as obras" else df_pend_base[df_pend_base['Obra_Vinculada'] == obra_filtro_sec1]

                    st.markdown("<br>", unsafe_allow_html=True)
                    cols_card = st.columns(3)
                    for i, (_, row) in enumerate(df_pend.sort_values('Obra_Vinculada').iterrows()):
                        edt_row = row.get('EDT_Vinculado', '')
                        numero_projeto_row = str(row.get('Numero_Projeto') or '')
                        if edt_row and edt_row != 'AVULSO' and not df_banco_macro.empty:
                            fr_row = df_banco_macro[df_banco_macro['EDT'] == edt_row]
                            tag_frente = (
                                f"🔗 {fr_row.iloc[0].get('Tarefa','')} · {fr_row.iloc[0].get('Tipo_Escopo','—')}"
                                if not fr_row.empty else "🔗 Frente vinculada"
                            )
                        else:
                            tag_frente = "🧩 Sem frente — suporte/avulso"
                        with cols_card[i % 3]:
                            with st.container(border=True):
                                st.markdown(f"**{row['Obra_Vinculada']}**")
                                st.caption(f"{row.get('Tipo_Material','—')} · Projeto {numero_projeto_row or '—'}")
                                st.markdown(f"<span style='font-size:12px;color:#64748B;'>{tag_frente}</span>", unsafe_allow_html=True)
                                if not numero_projeto_row:
                                    st.warning("Sem nº de projeto — edite a frente/projeto antes de liberar.", icon="⚠️")
                                else:
                                    if st.button("✅ Liberar esta OP", key=f"liberar_{row['id']}", use_container_width=True):
                                        _liberar_um_lote(int(row['id']), numero_projeto_row)

            st.markdown("<br>", unsafe_allow_html=True)

            # ── CARD 2: LANÇAR PEÇAS ──────────────────────────────
            n_lib_card = len(df_banco_micro[df_banco_micro["Status_Item"].isin(["Liberado para Fabrica","Parcialmente Concluido"])]) if not df_banco_micro.empty else 0
            with st.expander(f"✏️ Lançar Peças da OP  ·  {n_lib_card} OP(s) liberada(s)", expanded=False):
                st.markdown("Selecione a OP liberada e lance as peças do romaneio.")
                st.markdown("---")

                obras_fat = sorted(df_banco_micro["Obra_Vinculada"].dropna().unique().tolist()) if not df_banco_micro.empty else []
                obra_fat_sel = st.selectbox("Filtrar por obra:", ["Todas"] + obras_fat, key="fat_obra_sel")

                if not df_banco_micro.empty:
                    mask_fat = df_banco_micro['Status_Item'].isin(["Liberado para Fabrica", "Parcialmente Concluido"])
                    if obra_fat_sel != "Todas":
                        mask_fat = mask_fat & (df_banco_micro['Obra_Vinculada'] == obra_fat_sel)
                    df_lib = df_banco_micro[mask_fat].copy()
                else:
                    df_lib = pd.DataFrame()

                if not df_lib.empty:
                    opcoes_lotes = [
                        f"{row['Num_OP']} — {row['Cod_Lote']} | {row['Tipo_Material']} | {row['M2_Item']:.2f}m²"
                        for _, row in df_lib.iterrows() if row.get('Num_OP')
                    ]
                    if opcoes_lotes:
                        lote_sel_str = st.selectbox("Selecione o Lote / OP:", opcoes_lotes, key="sel_lote_pecas")
                        num_op_sel   = lote_sel_str.split(" — ")[0].strip()
                        row_lote     = df_lib[df_lib['Num_OP'] == num_op_sel].iloc[0]
                        lote_id      = int(row_lote['id'])
                        edt_lote     = row_lote.get('EDT_Vinculado', '')

                        if not df_banco_macro.empty and edt_lote and edt_lote != 'AVULSO':
                            fr_edt = df_banco_macro[df_banco_macro['EDT'] == edt_lote]
                            if not fr_edt.empty:
                                macro_row_sel = fr_edt.iloc[0]
                                tipo_esc_edt  = str(macro_row_sel.get('Tipo_Escopo', 'ACM') or 'ACM')
                                num_proj_edt  = str(macro_row_sel.get('Numero_Projeto', '') or '')
                                se1, se2, se3, se4 = st.columns(4)
                                se1.metric("Etapa", edt_lote)
                                se2.metric("Projeto", num_proj_edt or "—")
                                if tipo_esc_edt == "Esquadria-Vidro":
                                    peso_total_raw = macro_row_sel.get('Peso_Total_Tarefa')
                                    peso_total_edt = float(peso_total_raw) if pd.notna(peso_total_raw) and float(peso_total_raw) > 0 else None
                                    peso_exec_edt  = float(macro_row_sel.get('Peso_Executado', 0) or 0)
                                    peso_saldo_edt = (peso_total_edt - peso_exec_edt) if peso_total_edt is not None else None
                                    se3.metric("Peso Total Etapa", f"{peso_total_edt:.2f} kg" if peso_total_edt is not None else "— (não definido)")
                                    se4.metric("Saldo Etapa", f"{peso_saldo_edt:.2f} kg" if peso_saldo_edt is not None else "—")
                                else:
                                    m2_total_raw = macro_row_sel.get('M2_Total_Tarefa')
                                    m2_total_edt = float(m2_total_raw) if pd.notna(m2_total_raw) and float(m2_total_raw) > 0 else None
                                    m2_exec_edt  = float(macro_row_sel.get('M2_Executado', 0) or 0)
                                    m2_saldo_edt = (m2_total_edt - m2_exec_edt) if m2_total_edt is not None else None
                                    se3.metric("m² Total Etapa", f"{m2_total_edt:.2f}" if m2_total_edt is not None else "— (não definido)")
                                    se4.metric("Saldo Etapa", f"{m2_saldo_edt:.2f} m²" if m2_saldo_edt is not None else "—")
                            else:
                                macro_row_sel = {}
                                tipo_esc_edt  = "ACM"
                                num_proj_edt  = ""
                        else:
                            macro_row_sel = {}
                            tipo_esc_edt  = str(row_lote.get('Escopo') or 'ACM')
                            num_proj_edt  = str(row_lote.get('Numero_Projeto', '') or '')
                            if edt_lote == 'AVULSO':
                                st.info("OP Avulsa — sem vínculo com etapa do cronograma.")

                        st.markdown("---")
                        df_pecas_existentes = carregar_pecas_lote(lote_id)

                        if not df_pecas_existentes.empty:
                            pc1, pc2 = st.columns([3, 1])
                            with pc1:
                                st.success(f"✅ {len(df_pecas_existentes)} peça(s) lançadas | "
                                           f"Total: {int(df_pecas_existentes['qtd_total'].sum())} un | "
                                           f"Saldo: {int(df_pecas_existentes['saldo'].sum())} un")
                            with pc2:
                                st.caption(f"Status comp.: **{df_pecas_existentes.iloc[0].get('componentes_status','—')}**")

                            with st.expander("📋 Ver peças lançadas", expanded=False):
                                cols_peca = [c for c in ['codigo', 'descricao', 'localizacao', 'medida', 'qtd_total', 'qtd_enviada', 'saldo']
                                             if c in df_pecas_existentes.columns]
                                st.dataframe(
                                    df_pecas_existentes[cols_peca],
                                    hide_index=True, use_container_width=True
                                )

                            st.markdown("#### 📄 Gerar Ordem de Produção")
                            with st.expander("Configurar e Gerar OP", expanded=False):
                                obs_op = st.text_area("Observações:", key="obs_op",
                                                       placeholder="Informações adicionais para a produção...")
                                campos_extras = {"observacoes": obs_op, "material": row_lote.get('Tipo_Material', '')}
                                if tipo_esc_edt.upper() == "ACM":
                                    gf1, gf2 = st.columns(2)
                                    with gf1:
                                        qtd_folhas = st.text_input("Qtd Folhas Projeto:", key="op_folhas", placeholder="Ex: 2")
                                    with gf2:
                                        area_real = st.number_input("Área Total (m²):", value=float(row_lote.get('M2_Item', 0)), key="op_area")
                                    campos_extras.update({
                                        "qtd_folhas": qtd_folhas, "area_total": area_real,
                                        "dificuldade": row_lote.get('Dificuldade', '—'),
                                        "material": row_lote.get('Tipo_Material', 'ACM')
                                    })
                                elif "ESQUADRIA" in tipo_esc_edt.upper() or "VIDRO" in tipo_esc_edt.upper():
                                    gf1, gf2 = st.columns(2)
                                    with gf1:
                                        peso_total = st.text_input("Peso Total (kg):", key="op_peso", placeholder="Ex: 67,08kg")
                                    with gf2:
                                        qtd_folhas = st.text_input("Qtd Folhas Projeto:", key="op_folhas2", placeholder="Ex: 2")
                                    campos_extras.update({
                                        "peso_total": peso_total, "qtd_folhas": qtd_folhas,
                                        "material": "PERFIL EM ALUMINIO"
                                    })
                                else:
                                    empresa  = st.text_input("Empresa Responsável:", key="op_empresa")
                                    mat_terc = st.text_input("Material:", key="op_mat_terc")
                                    campos_extras.update({"empresa": empresa, "material": mat_terc})
                                if st.button("🖨️ Gerar OP", key="btn_gerar_op", type="primary"):
                                    op_bytes = gerar_op_xlsx(row_lote, df_pecas_existentes, macro_row_sel, campos_extras)
                                    st.download_button(
                                        label="📥 Baixar Ordem de Produção",
                                        data=op_bytes,
                                        file_name=f"OP_{num_op_sel}_{row_lote['Cod_Lote']}.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        key="dl_op"
                                    )

                        with st.expander(
                            "✏️ Editar / Substituir Peças" if not df_pecas_existentes.empty else "➕ Lançar Peças",
                            expanded=df_pecas_existentes.empty
                        ):
                            st.caption("Cole os dados abaixo — um item por linha em cada campo.")
                            gm1, gm2 = st.columns([2, 3])
                            eh_esq_pecas = tipo_esc_edt == "Esquadria-Vidro"
                            with gm1:
                                if eh_esq_pecas:
                                    peso_op_real = st.number_input(
                                        "⚖️ Peso real desta OP (kg):",
                                        min_value=0.0, value=float(row_lote.get('Peso_Kg', 0)),
                                        step=0.1, format="%.2f", key="peso_op_real_input",
                                        help="Será abatido do saldo da etapa."
                                    )
                                    m2_op_real = 0.0
                                else:
                                    m2_op_real = st.number_input(
                                        "📐 m² real desta OP:",
                                        min_value=0.0, value=float(row_lote.get('M2_Item', 0)),
                                        step=0.1, format="%.2f", key="m2_op_real_input",
                                        help="Será abatido do saldo da etapa."
                                    )
                                    peso_op_real = 0.0
                            with gm2:
                                if not df_banco_macro.empty and edt_lote and edt_lote != 'AVULSO':
                                    fr2 = df_banco_macro[df_banco_macro['EDT'] == edt_lote]
                                    if not fr2.empty:
                                        if eh_esq_pecas:
                                            peso_exec2      = float(fr2.iloc[0].get('Peso_Executado', 0) or 0)
                                            peso_total2_raw = fr2.iloc[0].get('Peso_Total_Tarefa')
                                            if pd.notna(peso_total2_raw) and float(peso_total2_raw) > 0:
                                                saldo_apos = float(peso_total2_raw) - peso_exec2 - peso_op_real
                                                st.metric("Saldo após este lançamento", f"{saldo_apos:.2f} kg")
                                            else:
                                                st.metric("Saldo após este lançamento", "— (total não definido)")
                                        else:
                                            m2_exec2      = float(fr2.iloc[0].get('M2_Executado', 0) or 0)
                                            m2_total2_raw = fr2.iloc[0].get('M2_Total_Tarefa')
                                            if pd.notna(m2_total2_raw) and float(m2_total2_raw) > 0:
                                                saldo_apos = float(m2_total2_raw) - m2_exec2 - m2_op_real
                                                st.metric("Saldo após este lançamento", f"{saldo_apos:.2f} m²")
                                            else:
                                                st.metric("Saldo após este lançamento", "— (total não definido)")

                            tab_rom, tab_manual = st.tabs(["📋 Colar Romaneio", "✏️ Manual"])

                            with tab_rom:
                                st.caption("Cole o romaneio — uma peça por linha: `REF - COD - Descrição - Medida - Ângulos - Qtd`")
                                txt_rom = st.text_area(
                                    "Romaneio:",
                                    height=220,
                                    placeholder="JA19.AP402 - CT016 - Cantoneira 25x25 Preto Fosco - 3120mm - 90/90 - 1pç\nJA22 - MM035 - Luva da Coluna Central - 4450mm - 90/90 - 2pç",
                                    key="sec2_romaneio_txt"
                                )
                                pecas_rom = parse_romaneio(txt_rom) if txt_rom.strip() else []
                                if pecas_rom:
                                    st.success(f"**{len(pecas_rom)} peça(s) reconhecida(s):**")
                                    st.dataframe(
                                        pd.DataFrame(pecas_rom).rename(columns={
                                            "codigo": "Código", "descricao": "Descrição",
                                            "medida": "Medida", "localizacao": "Localização", "qtd": "Qtd"
                                        }),
                                        use_container_width=True, hide_index=True
                                    )
                                comp_status_rom = st.radio(
                                    "Status dos componentes:",
                                    ["Aguardando Projetista", "Componentes OK"],
                                    horizontal=True, key="comp_status_rom"
                                )
                                if st.button("💾 Salvar Romaneio", key="btn_salvar_rom", type="primary", disabled=not pecas_rom):
                                    salvar_pecas_lote(
                                        lote_id, row_lote['Obra_Vinculada'],
                                        row_lote['Cod_Lote'], row_lote['Num_OP'],
                                        pecas_rom, comp_status_rom, m2_op_real, peso_op_real
                                    )
                                    registrar_auditoria(st.session_state.usuario_nome, "LANCAMENTO_PECAS",
                                        f"OP {row_lote['Num_OP']} — {len(pecas_rom)} peça(s) via romaneio — Obra: {obra_selecionada}")
                                    st.toast(f"{len(pecas_rom)} peça(s) salvas!")
                                    time.sleep(0.3)
                                    st.rerun()

                            with tab_manual:
                                p1, p2, p3, p4 = st.columns(4)
                                with p1:
                                    codigos_txt = st.text_area("Códigos:", height=180, key="pecas_codigos",
                                                                placeholder="B24-01C\nB25-01C\nRF.JA.11/CNT")
                                with p2:
                                    qtds_txt = st.text_area("Quantidades:", height=180, key="pecas_qtds",
                                                             placeholder="3\n3\n1")
                                with p3:
                                    locs_txt = st.text_area("Localização:", height=180, key="pecas_locs",
                                                             placeholder="Pav 27-37\nPav 27-37\n(opcional)")
                                with p4:
                                    medidas_txt = st.text_area("Medidas:", height=180, key="pecas_medidas",
                                                                placeholder="550x390\n550x400\n(opcional)")
                                comp_status = st.radio(
                                    "Status dos componentes:",
                                    ["Aguardando Projetista", "Componentes OK"],
                                    horizontal=True, key="comp_status_radio"
                                )
                                if st.button("💾 Salvar Peças", key="btn_salvar_pecas", type="primary"):
                                    codigos = [l.strip() for l in codigos_txt.strip().split('\n') if l.strip()]
                                    qtds    = [l.strip() for l in qtds_txt.strip().split('\n') if l.strip()]
                                    locs    = [l.strip() for l in locs_txt.strip().split('\n')] if locs_txt.strip() else []
                                    medidas = [l.strip() for l in medidas_txt.strip().split('\n')] if medidas_txt.strip() else []
                                    if not codigos:
                                        st.error("Informe pelo menos um código.")
                                    else:
                                        pecas_list = []
                                        for i, cod in enumerate(codigos):
                                            pecas_list.append({
                                                "codigo":      cod,
                                                "descricao":   "",
                                                "qtd":         int(qtds[i]) if i < len(qtds) and qtds[i].isdigit() else 1,
                                                "localizacao": locs[i] if i < len(locs) else "",
                                                "medida":      medidas[i] if i < len(medidas) else "",
                                            })
                                        salvar_pecas_lote(
                                            lote_id, row_lote['Obra_Vinculada'],
                                            row_lote['Cod_Lote'], row_lote['Num_OP'],
                                            pecas_list, comp_status, m2_op_real, peso_op_real
                                        )
                                        registrar_auditoria(st.session_state.usuario_nome, "LANCAMENTO_PECAS",
                                            f"OP {row_lote['Num_OP']} — {len(pecas_list)} peça(s) — Obra: {obra_selecionada}")
                                        st.toast(f"{len(pecas_list)} peça(s) salvas!")
                                        time.sleep(0.3)
                                        st.rerun()

                        # ── SEÇÃO: ARQUIVOS DA OP ─────────────────────────────
                        st.markdown("---")
                        arqs_existentes = carregar_arquivos_op(lote_id)
                        label_arq = f"📎 Arquivos da OP ({len(arqs_existentes)})" if arqs_existentes else "📎 Arquivos da OP"
                        with st.expander(label_arq, expanded=False):
                            if setor in ["Master", "PCP", "Engenharia", "Producao"]:
                                uploaded = st.file_uploader(
                                    "Anexar arquivo (PDF, Excel, imagem):",
                                    type=["pdf", "xlsx", "xls", "png", "jpg", "jpeg", "dwg"],
                                    key=f"upload_arq_{lote_id}"
                                )
                                if uploaded is not None:
                                    if st.button("💾 Salvar arquivo", key=f"btn_salvar_arq_{lote_id}", type="primary"):
                                        ok = salvar_arquivo_op(
                                            lote_id,
                                            uploaded.name,
                                            uploaded.type or "",
                                            uploaded.read(),
                                            st.session_state.usuario_nome
                                        )
                                        if ok:
                                            st.toast(f"✅ {uploaded.name} salvo!")
                                            time.sleep(0.3)
                                            st.rerun()

                            if arqs_existentes:
                                st.markdown("**Arquivos anexados:**")
                                for arq in arqs_existentes:
                                    arq_id, arq_nome, arq_tipo, arq_enviado_por, arq_enviado_em = arq
                                    col_a, col_b, col_c = st.columns([4, 2, 1])
                                    col_a.markdown(f"📄 **{html_escape(arq_nome)}**")
                                    col_b.caption(f"{arq_enviado_por} — {pd.to_datetime(arq_enviado_em).strftime('%d/%m/%Y %H:%M')}")
                                    with col_c:
                                        conteudo_arq = carregar_conteudo_arquivo(arq_id)
                                        if conteudo_arq:
                                            _, _, bytes_arq = conteudo_arq
                                            st.download_button(
                                                "⬇️", data=bytes(bytes_arq),
                                                file_name=arq_nome, mime=arq_tipo or "application/octet-stream",
                                                key=f"dl_arq_{arq_id}"
                                            )
                                    if setor in ["Master", "PCP"]:
                                        if st.button("🗑️ Remover", key=f"del_arq_{arq_id}"):
                                            deletar_arquivo_op(arq_id)
                                            st.toast("Arquivo removido.")
                                            time.sleep(0.3)
                                            st.rerun()
                            else:
                                st.caption("Nenhum arquivo anexado ainda.")

                    else:
                        st.info("Nenhuma OP com número gerado ainda. Libere os lotes primeiro.")
                else:
                    st.info("Nenhuma OP liberada para esta obra ainda.")

            st.markdown("<br>", unsafe_allow_html=True)

            # ── CARD 4: COMPONENTES POR OP ────────────────────────
            with st.expander("📦 Componentes por OP", expanded=False):
                st.caption("Cadastre os componentes necessários para cada OP liberada.")

                df_lib_ops_todas = df_banco_micro[
                    df_banco_micro['Status_Item'].isin(["Liberado para Fabrica", "Parcialmente Concluido"])
                ].copy() if not df_banco_micro.empty else pd.DataFrame()

                obras_comp = sorted(df_lib_ops_todas['Obra_Vinculada'].dropna().unique().tolist()) if not df_lib_ops_todas.empty else []
                obra_comp_sel = st.selectbox("Obra:", ["Todas as obras"] + obras_comp, key="sel_obra_comp")
                df_lib_ops = df_lib_ops_todas if obra_comp_sel == "Todas as obras" else df_lib_ops_todas[df_lib_ops_todas['Obra_Vinculada'] == obra_comp_sel]

                if not df_lib_ops.empty:
                    opcoes_ops = [
                        f"{row['Num_OP']} — {row['Cod_Lote']} | {row['Tipo_Material']}"
                        for _, row in df_lib_ops.iterrows() if row.get('Num_OP')
                    ]
                    if opcoes_ops:
                        op_sel  = st.selectbox("OP:", opcoes_ops, key="sel_op_comp")
                        row_op  = df_lib_ops[df_lib_ops['Num_OP'] == op_sel.split(" — ")[0].strip()].iloc[0]

                        # Verificar status dos componentes via op_pecas
                        df_pecas_op = carregar_pecas_lote(int(row_op['id']))
                        comp_st = df_pecas_op.iloc[0]['componentes_status'] if not df_pecas_op.empty else "Aguardando Projetista"

                        if comp_st == "Aguardando Projetista":
                            st.warning("🟠 Componentes: **Aguardando Projetista** — cadastre as peças na Seção 2 primeiro e marque como 'Componentes OK'.")
                        else:
                            comp_existentes = carregar_componentes_op(int(row_op['id']))
                            if not comp_existentes.empty:
                                st.success(f"✅ {len(comp_existentes)} componente(s) já cadastrado(s) para esta OP.")

                            with st.expander(
                                "➕ Adicionar / Substituir lista de componentes",
                                expanded=comp_existentes.empty
                            ):
                                st.caption("⚠️ Salvar substitui a lista anterior desta OP.")
                                aba_manual, aba_romaneio = st.tabs(["✏️ Manual", "📋 Colar Romaneio"])

                                with aba_romaneio:
                                    st.caption("Cole o romaneio completo abaixo — uma peça por linha no formato: `Ref - Cód - Descrição - Medida - Ângulos - Qtd`")
                                    texto_romaneio = st.text_area(
                                        "Romaneio:",
                                        height=250,
                                        placeholder="JA19.AP402 - CT016 - Cantoneira 25x25 Preto Fosco - 3120mm - 90/90 - 1pç\nJA22 - MM035 - Luva da Coluna Central - 4450mm - 90/90 - 2pç",
                                        key="txt_romaneio_paste"
                                    )
                                    pecas_romaneio_comp = parse_romaneio(texto_romaneio) if texto_romaneio.strip() else []
                                    parsed = [
                                        {
                                            "nome": f"{p['codigo']} - {p['descricao']}".strip(' -'),
                                            "qtd": p['qtd'],
                                            "unidade": "pç",
                                        }
                                        for p in pecas_romaneio_comp
                                    ]
                                    if parsed:
                                        st.success(f"**{len(parsed)} linha(s) reconhecida(s):**")
                                        df_prev_romaneio = pd.DataFrame(parsed).rename(columns={"nome": "Componente", "qtd": "Qtd", "unidade": "Un"})
                                        st.dataframe(df_prev_romaneio, use_container_width=True, hide_index=True)
                                    if st.button("💾 Salvar romaneio", key="btn_salvar_romaneio", disabled=not parsed):
                                        salvar_componentes(int(row_op['id']), row_op['Obra_Vinculada'], row_op['Cod_Lote'], row_op['Num_OP'], parsed)
                                        st.toast(f"Romaneio salvo para {row_op['Num_OP']} ({len(parsed)} itens)!")
                                        time.sleep(0.3)
                                        st.rerun()

                                with aba_manual:
                                    num_itens = st.number_input("Quantos componentes?", min_value=1, max_value=30, value=3, key="num_comp")
                                    componentes_input = []
                                    for idx in range(int(num_itens)):
                                        c1, c2, c3 = st.columns([4, 2, 2])
                                        with c1:
                                            nome = st.text_input(f"Componente {idx+1}:", key=f"comp_nome_{idx}")
                                        with c2:
                                            qtd  = st.number_input(f"Qtd {idx+1}:", min_value=0.0, value=1.0, key=f"comp_qtd_{idx}")
                                        with c3:
                                            und  = st.selectbox(f"Un {idx+1}:", ["un", "kg", "m", "m²", "cx", "pç", "rolo"], key=f"comp_und_{idx}")
                                        if nome.strip():
                                            componentes_input.append({"nome": nome.strip(), "qtd": qtd, "unidade": und})
                                    if st.button("💾 Salvar lista de componentes", key="btn_salvar_comp"):
                                        if not componentes_input:
                                            st.error("Preencha pelo menos um componente.")
                                        else:
                                            salvar_componentes(int(row_op['id']), row_op['Obra_Vinculada'], row_op['Cod_Lote'], row_op['Num_OP'], componentes_input)
                                            st.toast(f"Lista salva para {row_op['Num_OP']}!")
                                            time.sleep(0.3)
                                            st.rerun()
                    else:
                        st.info("Nenhuma OP com número gerado ainda.")
                else:
                    st.info("Nenhuma OP liberada para esta obra ainda.")

    # ==================================================
    # VISAO MACRO
    # ==================================================
    elif nome_aba == "Visao Macro":
        with aba_objeto:
            st.markdown('<div class="page-header"><div class="page-header-left"><h2>Visão Macro</h2><p>Indicadores executivos e visão consolidada das obras</p></div><span class="page-icon">🗺️</span></div>', unsafe_allow_html=True)
            if obra_selecionada and not df_banco_micro.empty:
                df_dir = df_banco_micro[df_banco_micro['Obra_Vinculada'] == obra_selecionada].copy()
            else:
                df_dir = df_banco_micro.copy()

            if not df_dir.empty:
                data_max = df_dir['Data_Limite_Obra'].max()
                c1, c2, c3 = st.columns(3)
                c1.metric("Metragem Total", f"{df_dir['M2_Item'].sum():,.2f} m²")
                c2.metric("Subdivisoes", f"{df_dir['EDT_Vinculado'].nunique()} frentes")
                c3.metric("Prazo Mais Distante", data_max.strftime('%d/%m/%Y') if prazo_valido(data_max) else "N/A")

                st.markdown("---")
                st.subheader("Carga Semanal")
                obras_macro = ["Todas as obras"] + (
                    sorted(df_banco_micro['Obra_Vinculada'].dropna().unique().tolist())
                    if not df_banco_micro.empty else []
                )

                st.markdown("#### 📋 Previsão")
                col_f1, _ = st.columns([2, 3])
                with col_f1:
                    filtro_prev = st.selectbox("🔍 Obra:", obras_macro, key="filtro_prev")
                df_prev = df_banco_micro.copy() if not df_banco_micro.empty else pd.DataFrame()
                if not df_prev.empty:
                    if filtro_prev != "Todas as obras":
                        df_prev = df_prev[df_prev['Obra_Vinculada'] == filtro_prev]
                    df_prev = df_prev[df_prev['Status_Item'] == 'Pendente']
                blocos_semanais(df_prev)

                st.markdown("---")
                st.markdown("#### 🔧 Em Produção")
                col_f2, _ = st.columns([2, 3])
                with col_f2:
                    filtro_prod = st.selectbox("🔍 Obra:", obras_macro, key="filtro_prod")
                df_prod = df_banco_micro.copy() if not df_banco_micro.empty else pd.DataFrame()
                if not df_prod.empty:
                    if filtro_prod != "Todas as obras":
                        df_prod = df_prod[df_prod['Obra_Vinculada'] == filtro_prod]
                    df_prod = df_prod[df_prod['Status_Item'] == 'Liberado para Fabrica']
                blocos_semanais(df_prod)

                st.markdown("---")
                st.markdown("#### ✅ Concluídos")
                col_f3, _ = st.columns([2, 3])
                with col_f3:
                    filtro_conc = st.selectbox("🔍 Obra:", obras_macro, key="filtro_conc")
                df_conc = df_banco_micro.copy() if not df_banco_micro.empty else pd.DataFrame()
                if not df_conc.empty:
                    if filtro_conc != "Todas as obras":
                        df_conc = df_conc[df_conc['Obra_Vinculada'] == filtro_conc]
                    df_conc = df_conc[df_conc['Status_Item'] == 'Concluido']
                blocos_semanais(df_conc)

                st.markdown("---")
                st.subheader("Gantt — Ocupacao da Fabrica")
                col_fg, _ = st.columns([2, 3])
                with col_fg:
                    filtro_gantt = st.selectbox("🔍 Obra:", obras_macro, key="filtro_gantt")
                df_gantt_base = df_banco_micro.copy() if not df_banco_micro.empty else pd.DataFrame()
                if not df_gantt_base.empty and filtro_gantt != "Todas as obras":
                    df_gantt_base = df_gantt_base[df_gantt_base['Obra_Vinculada'] == filtro_gantt]
                df_gantt = df_gantt_base.groupby(['Obra_Vinculada', 'EDT_Vinculado', 'Romaneio_Chapas']).agg(
                    Inicio=('Data_Producao_Programada', 'min'),
                    Fim=('Data_Limite_Obra', 'max'),
                    M2=('M2_Item', 'sum')
                ).reset_index().dropna(subset=['Inicio', 'Fim'])
                if not df_gantt.empty:
                    fig = px.timeline(
                        df_gantt, x_start="Inicio", x_end="Fim", y="EDT_Vinculado",
                        color="Obra_Vinculada", hover_data=["Romaneio_Chapas", "M2"],
                        title="Ocupacao Fabrica vs Prazo Despacho",
                        color_discrete_sequence=["#1E3A8A", "#1A56DB", "#0891B2", "#15803D"]
                    )
                    fig.update_yaxes(autorange="reversed")
                    fig.update_layout(height=400, margin=dict(l=20, r=20, t=40, b=20))
                    st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("Nenhum dado encontrado.")

    # ==================================================
    # VINCULAR DATAS
    # ==================================================
    elif nome_aba == "Vincular Datas":
        with aba_objeto:
            if setor != "Master":
                st.error("⛔ Acesso negado.")
                st.stop()
            st.markdown('<div class="page-header"><div class="page-header-left"><h2>Vincular Datas</h2><p>Fatiamento e vinculação de datas aos lotes de produção</p></div><span class="page-icon">📅</span></div>', unsafe_allow_html=True)
            if st.session_state.get('lote_salvo_sucesso'):
                st.success("Lote gerado com sucesso!")
                st.session_state.lote_salvo_sucesso = False

            if obra_selecionada and not df_macro_filtrado.empty:
                opcoes_edt = []
                mapa_rows  = {}
                for _, row in df_macro_filtrado.iterrows():
                    sub   = f" [{row['Subdivisao']}]" if row.get('Subdivisao') else ""
                    label = f"{row['EDT']} - {row['Tarefa']}{sub}"
                    opcoes_edt.append(label)
                    mapa_rows[label] = row

                st.markdown("### Nova Entrega")
                with st.form("form_fatiamento"):
                    c1, c2 = st.columns(2)
                    with c1:
                        edt_sel  = st.selectbox("Frente (EDT):", opcoes_edt)
                        row_sel  = mapa_rows[edt_sel]
                        edt_puro = edt_sel.split(" - ")[0].strip()
                        escopo_frente = str(row_sel.get('Tipo_Escopo') or 'ACM')
                        cod_lote = st.text_input("Nome do Lote (ex: LOTE 1, LOTE 2):")
                        txt_pav  = st.text_area("Pavimentos / Destino:", value="Pav 39 ao 43")
                        espec    = st.text_input("Material:", value="ACM BRANCO")
                        dific    = st.selectbox("Complexidade:", [1, 2, 3, 4, 5], index=2)
                    with c2:
                        inicio_prev = row_sel['Inicio_Previsto']
                        default_dt  = (
                            pd.to_datetime(inicio_prev).date() if prazo_valido(inicio_prev)
                            else (datetime.now() + timedelta(days=30)).date()
                        )
                        data_alvo = st.date_input("Precisa estar na obra ate:", value=default_dt, format="DD/MM/YYYY")
                        dias_log  = st.number_input("Dias de transporte (corridos):", min_value=1, value=3)
                        dias_fab  = st.number_input("Dias uteis de producao:", min_value=1, value=10)
                        total_cx  = st.number_input("Quantidade (un):", min_value=1, value=31)
                        eh_esq = escopo_frente == "Esquadria-Vidro"
                        if eh_esq:
                            total_kg = st.number_input("Peso total (kg):", min_value=0.0, value=0.0, step=0.1)
                            total_m2 = 0.0
                        else:
                            total_m2  = st.number_input("Metragem (m²):", min_value=0.0, value=70.0)
                            total_kg  = 0.0
                        dt_alvo   = datetime.combine(data_alvo, datetime.min.time())
                        dt_desp   = dt_alvo - timedelta(days=int(dias_log))
                        dt_inicio = subtrair_dias_uteis(dt_desp, int(dias_fab))
                        st.success(
                            f"Cronograma calculado:\n\n"
                            f"Inicio producao: **{dt_inicio.strftime('%d/%m/%Y')}**\n\n"
                            f"Saida da fabrica: **{dt_desp.strftime('%d/%m/%Y')}**\n\n"
                            f"Chegada na obra: **{data_alvo.strftime('%d/%m/%Y')}**"
                        )
                    if row_sel.get('Detalhado', True) is not False and "Liberados" not in str(row_sel.get('Status_Engenharia', '')):
                        st.warning(f"Atencao — Engenharia: `{row_sel.get('Status_Engenharia', '—')}`")
                    # Mudanca 6: a frente e a fonte da verdade da janela — o lote e validado contra ela,
                    # nunca mais reescreve as datas da frente silenciosamente.
                    fora_da_janela = []
                    janela_primeiro = row_sel.get('Primeiro_Dia_Producao')
                    janela_desp     = row_sel.get('Data_Limite_Despacho')
                    janela_fim      = row_sel.get('Termino_Obra')
                    if prazo_valido(janela_primeiro) and dt_inicio.date() < pd.to_datetime(janela_primeiro).date():
                        fora_da_janela.append(f"início de produção ({dt_inicio.strftime('%d/%m/%Y')}) é antes do previsto pela frente ({pd.to_datetime(janela_primeiro).strftime('%d/%m/%Y')})")
                    if prazo_valido(janela_desp) and dt_desp.date() > pd.to_datetime(janela_desp).date():
                        fora_da_janela.append(f"despacho ({dt_desp.strftime('%d/%m/%Y')}) é depois do limite da frente ({pd.to_datetime(janela_desp).strftime('%d/%m/%Y')})")
                    if prazo_valido(janela_fim) and data_alvo > pd.to_datetime(janela_fim).date():
                        fora_da_janela.append(f"entrega ({data_alvo.strftime('%d/%m/%Y')}) é depois do término da frente ({pd.to_datetime(janela_fim).strftime('%d/%m/%Y')})")
                    confirmar_fora_janela = False
                    if fora_da_janela:
                        st.warning("Este lote fica fora da janela da frente: " + "; ".join(fora_da_janela) +
                                   ". Edite a frente se a janela mudou, ou confirme abaixo pra seguir mesmo assim.")
                        confirmar_fora_janela = st.checkbox("Confirmo e quero gerar o lote mesmo fora da janela da frente")
                    if st.form_submit_button("Gerar Lote"):
                        if not cod_lote.strip():
                            st.error("Digite o nome do lote.")
                        elif fora_da_janela and not confirmar_fora_janela:
                            st.error("Lote fora da janela da frente — marque a confirmação acima ou ajuste as datas/frente.")
                        else:
                            deletar_lotes_por_edt_lote(obra_selecionada, edt_puro, cod_lote.strip())
                            lote = gerar_lote_unico(
                                data_alvo, int(dias_log), int(dias_fab),
                                int(total_cx), float(total_m2),
                                obra_selecionada, edt_puro, cod_lote.strip(), espec, txt_pav, dific,
                                total_kg=float(total_kg),
                                escopo=escopo_frente,
                                numero_projeto=str(row_sel.get('Numero_Projeto') or '')
                            )
                            ok, msg = salvar_lotes_micro(lote)
                            if ok:
                                registrar_auditoria(st.session_state.usuario_nome, "GERAR_LOTE",
                                f"Lote {cod_lote} — EDT {edt_puro} — {total_m2}m² — {total_kg}kg — Obra: {obra_selecionada}")
                                st.session_state.lote_salvo_sucesso = True
                                st.rerun()
                            else:
                                st.error(msg)

                st.markdown("---")
                st.markdown("### Lotes Gerados")
                df_ed_raw = carregar_micro_por_obra(obra_selecionada)
                if not df_ed_raw.empty:
                    df_obra = df_ed_raw.copy()
                    df_obra['Data_Producao_Programada'] = df_obra['Data_Producao_Programada'].dt.strftime('%Y-%m-%d')
                    df_obra['Data_Limite_Obra']         = df_obra['Data_Limite_Obra'].dt.strftime('%Y-%m-%d')
                    df_str  = df_obra.copy()
                    df_edit = st.data_editor(df_str, key="editor_lotes", hide_index=True, use_container_width=True,
                                             disabled=["id", "Obra_Vinculada", "Num_OP"])
                    alteradas = []
                    for i in df_edit.index:
                        try:
                            if not df_edit.loc[i].equals(df_str.loc[i]):
                                alteradas.append(df_edit.loc[i])
                        except Exception:
                            pass
                    if alteradas:
                        conn = conectar_banco()
                        try:
                            cursor = conn.cursor()
                            for row in alteradas:
                                cursor.execute("""
                                    UPDATE itens_detalhado
                                    SET EDT_Vinculado=%s, Cod_Lote=%s, Tipo_Material=%s, Qtd_Caixas=%s, M2_Item=%s,
                                        Data_Producao_Programada=%s, Data_Limite_Obra=%s,
                                        Romaneio_Chapas=%s, Status_Item=%s, Dificuldade=%s, Fase_Produtiva=%s,
                                        updated_at=NOW()
                                    WHERE id=%s
                                """, (
                                    row['EDT_Vinculado'], row['Cod_Lote'], row['Tipo_Material'], int(row['Qtd_Caixas']),
                                    float(row['M2_Item']), row['Data_Producao_Programada'],
                                    row['Data_Limite_Obra'], row['Romaneio_Chapas'],
                                    row['Status_Item'], int(row['Dificuldade']),
                                    row['Fase_Produtiva'], int(row['id'])
                                ))
                            conn.commit()
                            _limpar_cache_geral()
                        except Exception as e:
                            conn.rollback()
                            st.error(f"Erro ao salvar: {e}")
                        finally:
                            liberar_conexao(conn)
                        st.toast("Salvo!")
                        time.sleep(0.3)
                        st.rerun()

                    st.markdown("#### Remover Lote")
                    lote_del = st.selectbox("Lote para excluir:", df_obra['Cod_Lote'].unique().tolist())
                    if st.button(f"Excluir {lote_del}"):
                        deletar_lotes_por_edt_lote(obra_selecionada, None, lote_del)
                        registrar_auditoria(st.session_state.usuario_nome, "EXCLUIR_LOTE",
                            f"Lote {lote_del} excluído — Obra: {obra_selecionada}")
                        st.toast(f"Lote {lote_del} removido!")
                        time.sleep(0.5)
                        st.rerun()
                else:
                    st.info("Nenhum lote fatiado ainda.")

    # ==================================================
    # CADASTRAR OBRA
    # ==================================================
    elif nome_aba == "Cadastrar Obra":
        with aba_objeto:
            st.markdown('<div class="page-header"><div class="page-header-left"><h2>Cadastrar Obra</h2><p>Registre obras, projetos e frentes no cronograma macro</p></div><span class="page-icon">🏗️</span></div>', unsafe_allow_html=True)
            for k, v in [
                ('mem_obra', ''), ('mem_frente', ''), ('mem_tarefa', ''),
                ('mem_dt_ini', datetime.now().date()),
                ('mem_dt_fim', (datetime.now() + timedelta(days=90)).date())
            ]:
                if k not in st.session_state:
                    st.session_state[k] = v

            # ===== PASSO 1 — Obra + Projeto ==========================
            st.markdown("### 1. Obra e Projeto")
            st.caption("Toda obra é organizada em projetos — ex: Dona Lola pode ter os projetos 1063, 1068, 1072.")
            obras_existentes = sorted(df_projetos['Obra'].dropna().unique().tolist()) if not df_projetos.empty else []
            with st.form("form_projeto"):
                fp1, fp2 = st.columns(2)
                with fp1:
                    obra_escolha = st.selectbox("Obra:", ["+ Nova obra"] + obras_existentes, key="obra_escolha_projeto")
                    nome_obra_novo = st.text_input("Nome da nova obra:", value=st.session_state.mem_obra) if obra_escolha == "+ Nova obra" else obra_escolha
                with fp2:
                    numero_projeto_novo = st.text_input("Número do Projeto:", placeholder="Ex: 1068")
                if st.form_submit_button("➕ Adicionar Projeto"):
                    nome_obra_final = (nome_obra_novo or "").strip().upper()
                    if not nome_obra_final:
                        st.error("Informe o nome da obra.")
                    elif not numero_projeto_novo.strip():
                        st.error("Informe o número do projeto.")
                    else:
                        st.session_state.mem_obra = nome_obra_final
                        conn = conectar_banco()
                        try:
                            cursor = conn.cursor()
                            cursor.execute(
                                "INSERT INTO projetos (Obra, Numero_Projeto) VALUES (%s,%s) ON CONFLICT (Obra, Numero_Projeto) DO NOTHING",
                                (nome_obra_final, numero_projeto_novo.strip())
                            )
                            criado = cursor.rowcount > 0
                            conn.commit()
                            _limpar_cache_geral()
                            registrar_auditoria(st.session_state.usuario_nome, "CADASTRAR_PROJETO",
                                f"Obra: {nome_obra_final} | Projeto: {numero_projeto_novo.strip()}")
                            st.toast("Projeto adicionado!" if criado else "Esse projeto já estava cadastrado — tudo certo.")
                            time.sleep(0.4)
                            st.rerun()
                        except Exception as e:
                            conn.rollback()
                            st.error(f"Erro: {e}")
                        finally:
                            liberar_conexao(conn)

            if not df_projetos.empty:
                st.dataframe(df_projetos[['Obra', 'Numero_Projeto']], hide_index=True, use_container_width=True)

            st.markdown("---")
            # ===== PASSO 2 — Frente (opcional) dentro de um Projeto ===
            st.markdown("### 2. Frente (opcional) dentro de um Projeto")
            st.caption("Preencha só quando já souber a subdivisão, a metragem/peso e as datas dessa etapa. "
                       "Sem isso, o projeto/escopo já fica pronto pra lançar OPs — o detalhamento pode ser adicionado depois.")
            if df_projetos.empty:
                st.info("Cadastre um projeto no passo 1 antes de detalhar uma frente.")
            else:
                with st.form("form_frente"):
                    ff1, ff2, ff3 = st.columns(3)
                    with ff1:
                        obra_frente = st.selectbox("Obra:", sorted(df_projetos['Obra'].unique().tolist()), key="obra_frente_sel")
                    with ff2:
                        projetos_da_obra = sorted(df_projetos[df_projetos['Obra'] == obra_frente]['Numero_Projeto'].unique().tolist())
                        projeto_frente = st.selectbox("Projeto:", projetos_da_obra, key="projeto_frente_sel")
                    with ff3:
                        escopo = st.selectbox("Escopo:", ["ACM", "Esquadria-Vidro", "Terceirizada"], key="escopo_frente_sel")

                    with st.expander("➕ Detalhar frente (opcional)", expanded=False):
                        frente = st.text_input("Frente Macro:", value=st.session_state.mem_frente)
                        tarefa = st.text_input("Nome da Tarefa:", value=st.session_state.mem_tarefa)
                        de1, de2 = st.columns(2)
                        with de1:
                            edt_cod = st.text_input("Codigo EDT (base, sem a subdivisao):")
                        with de2:
                            subdiv  = st.text_input("Subdivisao / Balancim:").upper()
                        st.caption("O EDT salvo será o código base + a subdivisão, garantindo que cada subdivisão tenha um identificador único mesmo dentro da mesma etapa.")
                        if escopo == "Esquadria-Vidro":
                            peso_tot = st.number_input("Peso total (kg):", min_value=0.0, value=0.0,
                                                        help="Deixe 0 se ainda não souber o total dessa etapa.")
                            m2_tot = 0.0
                        else:
                            m2_tot = st.number_input("Metragem (m²):", min_value=0.0, value=0.0,
                                                      help="Deixe 0 se ainda não souber o total dessa etapa.")
                            peso_tot = 0.0
                        dd1, dd2 = st.columns(2)
                        with dd1:
                            dt_ini = st.date_input("Inicio da Instalacao:", value=st.session_state.mem_dt_ini, format="DD/MM/YYYY")
                        with dd2:
                            dt_fim = st.date_input("Prazo Maximo Obra:", value=st.session_state.mem_dt_fim, format="DD/MM/YYYY")

                    if st.form_submit_button("Registrar"):
                        if bool(edt_cod.strip()) != bool(subdiv.strip()):
                            st.error("Preencha Código EDT e Subdivisão juntos, ou deixe os dois em branco para cadastrar só o projeto/escopo.")
                        elif edt_cod.strip() and subdiv.strip() and not tarefa.strip():
                            st.error("Informe o nome da tarefa para detalhar a frente.")
                        else:
                            st.session_state.mem_frente = frente
                            st.session_state.mem_tarefa = tarefa
                            st.session_state.mem_dt_ini = dt_ini
                            st.session_state.mem_dt_fim = dt_fim
                            detalhado = bool(edt_cod.strip() and subdiv.strip())
                            conn = conectar_banco()
                            try:
                                cursor = conn.cursor()
                                if detalhado:
                                    edt_final = f"{edt_cod.strip()} [{subdiv.strip()}]"
                                    cursor.execute("""
                                        INSERT INTO cronograma_macro
                                        (Obra, EDT, Tipo_Escopo, Etapa_Macro, Subdivisao, Tarefa,
                                         M2_Total_Tarefa, Peso_Total_Tarefa, Inicio_Previsto, Termino_Obra, Status,
                                         Status_Engenharia, Numero_Projeto, Detalhado)
                                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'Pendente','Aguardando Medicao In Loco',%s,TRUE)
                                    """, (obra_frente, edt_final, escopo, frente.strip(), subdiv.strip(), tarefa.strip(),
                                          float(m2_tot) if m2_tot > 0 else None, float(peso_tot) if peso_tot > 0 else None,
                                          dt_ini.strftime('%Y-%m-%d'),
                                          dt_fim.strftime('%Y-%m-%d'), projeto_frente))
                                    conn.commit()
                                    _limpar_cache_geral()
                                    registrar_auditoria(st.session_state.usuario_nome, "CADASTRAR_OBRA",
                                        f"Obra: {obra_frente} | Projeto: {projeto_frente} | EDT: {edt_final} | {m2_tot}m² | {peso_tot}kg")
                                    st.toast("Frente detalhada registrada!")
                                else:
                                    edt_final = f"{obra_frente}-{projeto_frente}-{escopo}".upper().replace(" ", "-")
                                    cursor.execute("""
                                        INSERT INTO cronograma_macro
                                        (Obra, EDT, Tipo_Escopo, Etapa_Macro, Subdivisao, Tarefa,
                                         M2_Total_Tarefa, Peso_Total_Tarefa, Inicio_Previsto, Termino_Obra, Status,
                                         Status_Engenharia, Numero_Projeto, Detalhado)
                                        VALUES (%s,%s,%s,'','',%s,NULL,NULL,NULL,NULL,'Pendente','Aguardando Medicao In Loco',%s,FALSE)
                                        ON CONFLICT (EDT) DO NOTHING
                                    """, (obra_frente, edt_final, escopo, "Geral (sem detalhamento)", projeto_frente))
                                    criado = cursor.rowcount > 0
                                    conn.commit()
                                    _limpar_cache_geral()
                                    registrar_auditoria(st.session_state.usuario_nome, "CADASTRAR_OBRA",
                                        f"Obra: {obra_frente} | Projeto: {projeto_frente} | Escopo: {escopo} | sem detalhamento")
                                    st.toast("Projeto/Escopo pronto para lançar OPs!" if criado else "Projeto/Escopo já estava cadastrado — tudo certo.")
                                time.sleep(0.4)
                                st.rerun()
                            except Exception as e:
                                conn.rollback()
                                st.error(f"EDT '{edt_final}' ja existe ou erro: {e}")
                            finally:
                                liberar_conexao(conn)

            if not df_banco_macro.empty:
                st.markdown("---")
                st.markdown("### Frentes Cadastradas")
                df_show = df_banco_macro.copy()
                for col in ['Inicio_Previsto', 'Termino_Obra']:
                    if col in df_show.columns:
                        df_show[col] = pd.to_datetime(df_show[col], errors='coerce').dt.strftime('%d/%m/%Y')
                if 'M2_Total_Tarefa' in df_show.columns:
                    df_show['M2_Total_Tarefa'] = df_show['M2_Total_Tarefa'].apply(
                        lambda v: f"{v:.2f}" if pd.notna(v) and v > 0 else "—")
                if 'Peso_Total_Tarefa' in df_show.columns:
                    df_show['Peso_Total_Tarefa'] = df_show['Peso_Total_Tarefa'].apply(
                        lambda v: f"{v:.2f}" if pd.notna(v) and v > 0 else "—")
                if 'Detalhado' in df_show.columns:
                    df_show['Detalhado'] = df_show['Detalhado'].apply(lambda v: "Sim" if v else "Não")
                cols_s = [c for c in ['Obra', 'EDT', 'Subdivisao', 'Tarefa', 'M2_Total_Tarefa', 'Peso_Total_Tarefa', 'Detalhado',
                                      'Inicio_Previsto', 'Termino_Obra', 'Status_Engenharia'] if c in df_show.columns]
                st.dataframe(df_show[cols_s], hide_index=True, use_container_width=True)

                st.markdown("#### ✏️ Editar Frente")
                st.caption("Corrige dados da frente sem precisar apagar e recriar. EDT, Obra e Subdivisão não mudam "
                           "(evita órfãos nos lotes já vinculados). Lotes já Liberados/Concluídos nunca são alterados; "
                           "só os ainda Pendentes recebem aviso se ficarem fora do novo saldo/janela.")
                opcoes_edit = [
                    f"{row['EDT']} — {row['Tarefa']} [{row.get('Subdivisao','')}]"
                    for _, row in df_banco_macro[df_banco_macro['Obra'] == obra_selecionada].iterrows()
                ] if not df_banco_macro.empty else []
                if opcoes_edit:
                    frente_edit_sel = st.selectbox("Frente para editar:", opcoes_edit, key="sel_edit_frente")
                    edt_edit = frente_edit_sel.split(" — ")[0].strip()
                    row_edit = df_banco_macro[df_banco_macro['EDT'] == edt_edit].iloc[0]
                    escopo_atual_edit = row_edit.get('Tipo_Escopo')
                    with st.form("form_editar_frente"):
                        fe1, fe2 = st.columns(2)
                        with fe1:
                            tarefa_edit = st.text_input("Tarefa:", value=str(row_edit.get('Tarefa') or ''))
                            escopo_edit = st.selectbox(
                                "Escopo:", ["ACM", "Esquadria-Vidro", "Terceirizada"],
                                index=["ACM", "Esquadria-Vidro", "Terceirizada"].index(escopo_atual_edit)
                                      if escopo_atual_edit in ["ACM", "Esquadria-Vidro", "Terceirizada"] else 0
                            )
                            projetos_obra_edit = sorted(df_projetos[df_projetos['Obra'] == obra_selecionada]['Numero_Projeto'].unique().tolist()) if not df_projetos.empty else []
                            projeto_atual_edit = str(row_edit.get('Numero_Projeto') or '')
                            if projeto_atual_edit and projeto_atual_edit not in projetos_obra_edit:
                                projetos_obra_edit.append(projeto_atual_edit)
                            projeto_edit = st.selectbox(
                                "Número do Projeto:", projetos_obra_edit,
                                index=projetos_obra_edit.index(projeto_atual_edit) if projeto_atual_edit in projetos_obra_edit else 0
                            ) if projetos_obra_edit else ""
                        with fe2:
                            if escopo_edit == "Esquadria-Vidro":
                                peso_edit = st.number_input(
                                    "Peso total (kg):", min_value=0.0,
                                    value=float(row_edit.get('Peso_Total_Tarefa') or 0.0)
                                )
                                m2_edit = 0.0
                            else:
                                m2_edit = st.number_input(
                                    "Metragem total (m²):", min_value=0.0,
                                    value=float(row_edit.get('M2_Total_Tarefa') or 0.0)
                                )
                                peso_edit = 0.0
                            dt_ini_edit = st.date_input(
                                "Início da Instalação:",
                                value=pd.to_datetime(row_edit.get('Inicio_Previsto')).date() if prazo_valido(row_edit.get('Inicio_Previsto')) else datetime.now().date(),
                                format="DD/MM/YYYY"
                            )
                            dt_fim_edit = st.date_input(
                                "Prazo Máximo Obra:",
                                value=pd.to_datetime(row_edit.get('Termino_Obra')).date() if prazo_valido(row_edit.get('Termino_Obra')) else (datetime.now() + timedelta(days=90)).date(),
                                format="DD/MM/YYYY"
                            )
                        if st.form_submit_button("💾 Salvar alterações da frente"):
                            conn = conectar_banco()
                            try:
                                cursor = conn.cursor()
                                cursor.execute("""
                                    UPDATE cronograma_macro
                                    SET Tarefa=%s, Tipo_Escopo=%s, Numero_Projeto=%s,
                                        M2_Total_Tarefa=%s, Peso_Total_Tarefa=%s, Inicio_Previsto=%s, Termino_Obra=%s
                                    WHERE EDT=%s
                                """, (
                                    tarefa_edit.strip(), escopo_edit, projeto_edit.strip(),
                                    m2_edit if m2_edit > 0 else None,
                                    peso_edit if peso_edit > 0 else None,
                                    dt_ini_edit.strftime('%Y-%m-%d'), dt_fim_edit.strftime('%Y-%m-%d'),
                                    edt_edit
                                ))
                                conn.commit()
                                _limpar_cache_geral()
                                registrar_auditoria(st.session_state.usuario_nome, "EDITAR_FRENTE",
                                    f"EDT {edt_edit} — Obra: {obra_selecionada} — Escopo: {escopo_edit}")

                                # Cascata (Mudanca 5): so avisa sobre lotes ainda Pendentes da mesma EDT —
                                # nunca toca em Liberado/Parcialmente Concluido/Concluido.
                                cursor.execute("""
                                    SELECT Cod_Lote, M2_Item, Peso_Kg, Data_Producao_Programada, Data_Limite_Obra
                                    FROM itens_detalhado WHERE EDT_Vinculado=%s AND Status_Item='Pendente'
                                """, (edt_edit,))
                                pendentes_edt = cursor.fetchall()
                                if pendentes_edt:
                                    if escopo_edit == "Esquadria-Vidro":
                                        soma_valor = sum(float(r[2] or 0) for r in pendentes_edt)
                                        total_novo, unidade = peso_edit, "kg"
                                    else:
                                        soma_valor = sum(float(r[1] or 0) for r in pendentes_edt)
                                        total_novo, unidade = m2_edit, "m²"
                                    if total_novo > 0 and soma_valor > total_novo:
                                        st.warning(
                                            f"⚠️ Os lotes Pendentes desta frente somam {soma_valor:.2f} {unidade}, "
                                            f"acima do novo total ({total_novo:.2f} {unidade}). Ajuste manualmente os lotes."
                                        )
                                    fora = [
                                        r[0] for r in pendentes_edt
                                        if (r[3] and pd.to_datetime(r[3]).date() < dt_ini_edit) or
                                           (r[4] and pd.to_datetime(r[4]).date() > dt_fim_edit)
                                    ]
                                    if fora:
                                        st.warning(
                                            "⚠️ Os lotes Pendentes a seguir ficaram fora da nova janela da frente "
                                            "(as datas deles NÃO foram alteradas automaticamente): " + ", ".join(fora)
                                        )
                                st.toast("Frente atualizada!")
                                time.sleep(0.5)
                                st.rerun()
                            except Exception as e:
                                conn.rollback()
                                st.error(f"Erro: {e}")
                            finally:
                                liberar_conexao(conn)

                st.markdown("#### Excluir Frente")
                st.warning("Isso remove a frente do cronograma macro.")
                opcoes_del = [
                    f"{row['EDT']} — {row['Tarefa']} [{row.get('Subdivisao','')}]"
                    for _, row in df_banco_macro[df_banco_macro['Obra'] == obra_selecionada].iterrows()
                ] if not df_banco_macro.empty else []
                if opcoes_del:
                    frente_del = st.selectbox("Frente para excluir:", opcoes_del, key="sel_del_frente")
                    edt_del    = frente_del.split(" — ")[0].strip()
                    if st.button(f"🗑️ Excluir frente {edt_del}", key="btn_del_frente", type="secondary"):
                        st.session_state["confirm_del_frente"] = edt_del
                    if st.session_state.get("confirm_del_frente") == edt_del:
                        st.error(f"⚠️ Confirma a exclusão da frente **{edt_del}**? Esta ação não pode ser desfeita.")
                        cc1, cc2 = st.columns(2)
                        with cc1:
                            if st.button("Sim, excluir", key="confirm_yes_frente", type="primary"):
                                conn = conectar_banco()
                                try:
                                    cursor = conn.cursor()
                                    cursor.execute("DELETE FROM cronograma_macro WHERE EDT=%s", (edt_del,))
                                    conn.commit()
                                    _limpar_cache_geral()
                                    registrar_auditoria(st.session_state.usuario_nome, "EXCLUIR_FRENTE",
                                        f"EDT {edt_del} excluída — Obra: {obra_selecionada}")
                                    st.toast(f"Frente {edt_del} removida!")
                                    del st.session_state["confirm_del_frente"]
                                    time.sleep(0.5)
                                    st.rerun()
                                except Exception as e:
                                    conn.rollback()
                                    st.error(f"Erro: {e}")
                                finally:
                                    liberar_conexao(conn)
                        with cc2:
                            if st.button("Cancelar", key="confirm_no_frente"):
                                del st.session_state["confirm_del_frente"]
                                st.rerun()

                st.markdown("---")
                st.markdown("#### 🗑️ Excluir Obra")
                st.warning("Isso remove **todas as frentes** da obra selecionada do cronograma macro.")
                obras_para_del = sorted(df_banco_macro['Obra'].unique().tolist()) if not df_banco_macro.empty else []
                if obras_para_del:
                    obra_del = st.selectbox("Obra para excluir:", obras_para_del, key="sel_del_obra")
                    n_frentes = len(df_banco_macro[df_banco_macro['Obra'] == obra_del])
                    st.caption(f"{n_frentes} frente(s) serão removidas.")
                    if st.button(f"🗑️ Excluir obra '{obra_del}'", key="btn_del_obra", type="secondary"):
                        st.session_state["confirm_del_obra"] = obra_del
                    if st.session_state.get("confirm_del_obra") == obra_del:
                        st.error(f"⚠️ Confirma a exclusão da obra **{obra_del}** e suas {n_frentes} frente(s)? Esta ação **não pode ser desfeita**.")
                        do1, do2 = st.columns(2)
                        with do1:
                            if st.button("Sim, excluir obra", key="confirm_yes_obra", type="primary"):
                                conn = conectar_banco()
                                try:
                                    cursor = conn.cursor()
                                    cursor.execute("DELETE FROM cronograma_macro WHERE Obra=%s", (obra_del,))
                                    conn.commit()
                                    _limpar_cache_geral()
                                    registrar_auditoria(st.session_state.usuario_nome, "EXCLUIR_OBRA",
                                        f"Obra '{obra_del}' removida — {n_frentes} frente(s)")
                                    st.toast(f"Obra '{obra_del}' removida com sucesso!")
                                    del st.session_state["confirm_del_obra"]
                                    time.sleep(0.5)
                                    st.rerun()
                                except Exception as e:
                                    conn.rollback()
                                    st.error(f"Erro: {e}")
                                finally:
                                    liberar_conexao(conn)
                        with do2:
                            if st.button("Cancelar", key="confirm_no_obra"):
                                del st.session_state["confirm_del_obra"]
                                st.rerun()
                else:
                    st.markdown('<div class="empty-state"><div class="empty-icon">🏗️</div><h4>Nenhuma obra cadastrada</h4><p>Cadastre uma obra na seção acima.</p></div>', unsafe_allow_html=True)

    # ==================================================
    # PAINEL DE ENGENHARIA
    # ==================================================
    elif nome_aba == "Painel de Engenharia":
        with aba_objeto:
            st.markdown('<div class="page-header"><div class="page-header-left"><h2>Painel de Engenharia</h2><p>Controle de status e alertas das frentes de engenharia</p></div><span class="page-icon">📐</span></div>', unsafe_allow_html=True)
            st.caption(f"Hoje: {hoje_projeto().strftime('%d/%m/%Y')} | Obra: **{obra_selecionada or 'Nenhuma'}**")
            df_eng = carregar_macro_por_obra(obra_selecionada) if obra_selecionada else pd.DataFrame()

            ESTADOS = [
                "Aguardando Medicao In Loco", "Medicao Realizada — Em Projetos",
                "Projetos em Revisao Interna", "Projetos Liberados para o PCP", "Arquivado / Concluido"
            ]
            STATUS_CORES = {
                "Aguardando Medicao In Loco":      "#EF4444",
                "Medicao Realizada — Em Projetos": "#EAB308",
                "Projetos em Revisao Interna":     "#3B82F6",
                "Projetos Liberados para o PCP":   "#22C55E",
                "Arquivado / Concluido":           "#94A3B8",
            }

            def classificar(dias_rest, status_tec):
                if status_tec == "Projetos Liberados para o PCP":
                    return "concluido", "Liberado para o PCP", None
                if dias_rest is None:
                    return "sem_prazo", "Aguardando programacao pelo PCP", None
                if dias_rest < 0:
                    return "vencido", f"VENCIDO ha {abs(int(dias_rest))} dias", abs(int(dias_rest))
                if dias_rest <= 7:
                    return "critico", f"Critico — faltam {int(dias_rest)} dias", int(dias_rest)
                return "ok", f"Dentro do prazo ({int(dias_rest)} dias)", int(dias_rest)

            frentes = []
            if not df_eng.empty:
                for _, row in df_eng.iterrows():
                    if row.get('Detalhado', True) is False:
                        continue  # obra/escopo sem frente detalhada — nada pra engenharia controlar aqui
                    prazo_raw = row.get('Prazo_Engenharia')
                    prazo_eng = prazo_raw if prazo_valido(prazo_raw) else None
                    dias_rest = (pd.to_datetime(prazo_eng) - hoje_projeto()).days if prazo_eng is not None else None
                    sk, situacao_txt, dias_num = classificar(dias_rest, row.get('Status_Engenharia', ESTADOS[0]))
                    m2_raw = row.get('M2_Total_Tarefa')
                    frentes.append({
                        "id": row['id'], "edt": row['EDT'], "tarefa": row['Tarefa'],
                        "subdivisao": row.get('Subdivisao', ''), "tipo_escopo": row.get('Tipo_Escopo', ''),
                        "inicio_previsto": row.get('Inicio_Previsto'), "despacho": row.get('Data_Limite_Despacho'),
                        "primeiro_prod": row.get('Primeiro_Dia_Producao'), "termino_obra": row.get('Termino_Obra'),
                        "m2": float(m2_raw) if pd.notna(m2_raw) and float(m2_raw) > 0 else None, "prazo_eng": prazo_eng,
                        "dias_restantes": dias_rest, "situacao_key": sk, "situacao_txt": situacao_txt,
                        "dias_num": dias_num, "status_tecnico": row.get('Status_Engenharia', ESTADOS[0]),
                    })

            criticas = [f for f in frentes if f['situacao_key'] in ('critico', 'vencido')]

            with st.expander(f"Frentes Criticas — {len(criticas)} alerta(s) · Obra: {obra_selecionada or 'Nenhuma'}", expanded=True):
                if not criticas:
                    st.success("Tudo dentro do prazo!")
                else:
                    for fr in sorted(criticas, key=lambda x: x['dias_restantes'] or 0):
                        with st.container(border=True):
                            ci, cc = st.columns([7, 3])
                            with ci:
                                sub = f" · {fr['subdivisao']}" if fr['subdivisao'] else ""
                                st.markdown(f"### {fr['tarefa']}{sub}")
                                cm1, cm2 = st.columns(2)
                                with cm1:
                                    st.write(f"EDT: `{fr['edt']}`")
                                    ini = pd.to_datetime(fr['inicio_previsto']).strftime('%d/%m/%Y') if prazo_valido(fr['inicio_previsto']) else "—"
                                    st.write(f"Inicio instalacao: {ini}")
                                    pp = pd.to_datetime(fr['primeiro_prod']).strftime('%d/%m/%Y') if prazo_valido(fr['primeiro_prod']) else "—"
                                    st.write(f"1º dia producao: {pp}")
                                with cm2:
                                    pe = pd.to_datetime(fr['prazo_eng']).strftime('%d/%m/%Y') if fr['prazo_eng'] else "—"
                                    st.write(f"Prazo engenharia: `{pe}`")
                                    dp = pd.to_datetime(fr['despacho']).strftime('%d/%m/%Y') if prazo_valido(fr['despacho']) else "—"
                                    st.write(f"Despacho: {dp}")
                                cor_status = STATUS_CORES.get(fr['status_tecnico'], "#94A3B8")
                                st.markdown(f"<span style='background:{cor_status}20;color:{cor_status};padding:3px 8px;border-radius:4px;font-size:12px;font-weight:600;'>{fr['status_tecnico']}</span>", unsafe_allow_html=True)
                            with cc:
                                if fr['situacao_key'] == 'vencido':
                                    st.error(f"VENCIDO\n\n**{fr['dias_num']} dias**")
                                else:
                                    st.warning(f"FALTAM\n\n**{fr['dias_num']} dias**")
                            st.markdown("---")
                            ca1, ca2 = st.columns(2)
                            with ca1:
                                idx = ESTADOS.index(fr['status_tecnico']) if fr['status_tecnico'] in ESTADOS else 0
                                ns  = st.selectbox("Atualizar status:", ESTADOS, index=idx, key=f"cs_{fr['id']}")
                            with ca2:
                                st.write("")
                                st.write("")
                                if st.button("Salvar", key=f"cb_{fr['id']}", use_container_width=True):
                                    atualizar_status_engenharia(fr['id'], ns)
                                    st.toast("Atualizado!")
                                    time.sleep(0.3)
                                    st.rerun()

            with st.expander(f"Todas as Frentes — {len(frentes)} · Obra: {obra_selecionada or 'Nenhuma'}", expanded=False):
                if not frentes:
                    st.info("Nenhuma frente cadastrada.")
                else:
                    cf1, cf2 = st.columns([3, 2])
                    with cf1:
                        filt_st  = st.selectbox("Status:", ["Todos"] + ESTADOS, key="eng_fst")
                    with cf2:
                        filt_sit = st.radio("Situacao:", ["Todas", "Criticas", "Liberadas"], horizontal=True, key="eng_fsi")
                    exibir = frentes.copy()
                    if filt_st != "Todos":
                        exibir = [f for f in exibir if f['status_tecnico'] == filt_st]
                    if filt_sit == "Criticas":
                        exibir = [f for f in exibir if f['situacao_key'] in ('critico', 'vencido')]
                    elif filt_sit == "Liberadas":
                        exibir = [f for f in exibir if f['situacao_key'] == 'concluido']
                    st.markdown(f"**{len(exibir)} frente(s)**")
                    st.markdown("---")
                    for fr in exibir:
                        with st.container(border=True):
                            ci, cd, ca = st.columns([5, 3, 2])
                            with ci:
                                sub = f" · *{fr['subdivisao']}*" if fr['subdivisao'] else ""
                                st.markdown(f"**{fr['tarefa']}**{sub}")
                                m2_txt = f"{fr['m2']:,.2f} m²" if fr['m2'] is not None else "m² não definido"
                                st.caption(f"EDT: {fr['edt']} | {fr['tipo_escopo']} | {m2_txt}")
                                cor_status = STATUS_CORES.get(fr['status_tecnico'], "#94A3B8")
                                st.markdown(f"<span style='background:{cor_status}20;color:{cor_status};padding:2px 8px;border-radius:4px;font-size:12px;font-weight:600;'>{fr['status_tecnico']}</span>", unsafe_allow_html=True)
                            with cd:
                                ini = pd.to_datetime(fr['inicio_previsto']).strftime('%d/%m/%Y') if prazo_valido(fr['inicio_previsto']) else "—"
                                st.caption("Inicio instalacao")
                                st.write(ini)
                                pe = pd.to_datetime(fr['prazo_eng']).strftime('%d/%m/%Y') if fr['prazo_eng'] else "—"
                                st.caption("Prazo PCP")
                                st.write(f"`{pe}`")
                                dp = pd.to_datetime(fr['despacho']).strftime('%d/%m/%Y') if prazo_valido(fr['despacho']) else "—"
                                st.caption("Despacho")
                                st.write(dp)
                            with ca:
                                sk = fr['situacao_key']
                                if sk == 'vencido':      st.error(fr['situacao_txt'])
                                elif sk == 'critico':    st.warning(fr['situacao_txt'])
                                elif sk in ('concluido', 'ok'): st.success(fr['situacao_txt'])
                                else:                    st.info(fr['situacao_txt'])
                            cs, cb = st.columns([4, 1])
                            with cs:
                                idx = ESTADOS.index(fr['status_tecnico']) if fr['status_tecnico'] in ESTADOS else 0
                                ns  = st.selectbox("", ESTADOS, index=idx, key=f"as_{fr['id']}", label_visibility="collapsed")
                            with cb:
                                if st.button("Salvar", key=f"ab_{fr['id']}", use_container_width=True):
                                    atualizar_status_engenharia(fr['id'], ns)
                                    st.toast("Atualizado!")
                                    time.sleep(0.3)
                                    st.rerun()

            df_sols = carregar_solicitacoes()
            n_pend  = len(df_sols[df_sols['status'] == 'Pendente de Aprovacao']) if not df_sols.empty else 0
            with st.expander(f"Solicitacoes de Prazo{f' — {n_pend} pendente(s)' if n_pend else ''}", expanded=False):
                nao_lib = [f for f in frentes if f['situacao_key'] != 'concluido']
                if setor in ["Engenharia", "Master"] and nao_lib:
                    st.markdown("#### Nova Solicitacao")
                    cs1, cs2 = st.columns(2)
                    with cs1:
                        opts = [f"{f['edt']} — {f['tarefa']}" for f in nao_lib]
                        sel  = st.selectbox("Frente:", opts, key="sol_fr")
                        fobj = nao_lib[opts.index(sel)]
                        pat  = pd.to_datetime(fobj['prazo_eng']).strftime('%d/%m/%Y') if fobj['prazo_eng'] else "Nao definido"
                        st.info(f"Prazo atual: **{pat}**")
                    with cs2:
                        nps = st.date_input("Novo prazo:", format="DD/MM/YYYY", key="sol_np",
                                            value=((pd.to_datetime(fobj['prazo_eng']) + timedelta(days=7)).date() if fobj['prazo_eng'] else hoje_projeto().date()))
                        jus = st.text_area("Justificativa:", key="sol_jus")
                    if st.button("Enviar solicitacao", key="sol_env"):
                        if not jus.strip():
                            st.error("Informe a justificativa.")
                        else:
                            salvar_solicitacao(fobj['edt'], fobj['tarefa'], pat, nps.strftime('%d/%m/%Y'), jus.strip(), st.session_state.usuario_nome)
                            st.success("Enviado!")
                            st.rerun()
                st.markdown("---")
                if setor == "Master" and not df_sols.empty:
                    pend = df_sols[df_sols['status'] == 'Pendente de Aprovacao']
                    if pend.empty:
                        st.info("Nenhuma pendente.")
                    else:
                        for _, sol in pend.iterrows():
                            with st.container(border=True):
                                sa, sb = st.columns([4, 2])
                                with sa:
                                    st.markdown(f"**{sol['tarefa']}** — `{sol['edt']}`")
                                    st.write(f"`{sol['prazo_atual']}` → `{sol['prazo_solicitado']}`")
                                    st.caption(f"*{sol['justificativa']}* | {sol['criado_por']} em {sol['criado_em']}")
                                with sb:
                                    cap, cre = st.columns(2)
                                    with cap:
                                        if st.button("Aprovar", key=f"ap_{sol['id']}", use_container_width=True):
                                            atualizar_status_solicitacao(sol['id'], "Aprovado")
                                            st.rerun()
                                    with cre:
                                        if st.button("Rejeitar", key=f"rj_{sol['id']}", use_container_width=True):
                                            atualizar_status_solicitacao(sol['id'], "Rejeitado")
                                            st.rerun()
                if not df_sols.empty:
                    hist = df_sols[df_sols['status'] != 'Pendente de Aprovacao']
                    if not hist.empty:
                        st.markdown("#### Historico")
                        for _, sol in hist.iterrows():
                            st.caption(f"{sol['status']} | **{sol['tarefa']}** | {sol['prazo_solicitado']} | {sol['criado_por']}")

            with st.expander("Carga da Fabrica por Semana", expanded=False):
                df_fab = (df_banco_micro[df_banco_micro['Status_Item'] == "Liberado para Fabrica"].copy()
                          if not df_banco_micro.empty else pd.DataFrame())
                if not df_fab.empty:
                    df_fab['Ano_Semana'] = df_fab['Data_Producao_Programada'].dt.isocalendar().year
                    df_fab['Num_Semana'] = df_fab['Data_Producao_Programada'].dt.isocalendar().week
                    def fmt_s(r):
                        try:
                            s = pd.to_datetime(f"{int(r['Ano_Semana'])}-W{int(r['Num_Semana'])}-1", format="%G-W%V-%u")
                            return f"Semana {int(r['Num_Semana']):02d} ({s.strftime('%d/%m')} – {(s + timedelta(days=6)).strftime('%d/%m')})"
                        except Exception:
                            return f"Semana {r['Num_Semana']}"
                    df_fab['Periodo'] = df_fab.apply(fmt_s, axis=1)
                    res_fab = df_fab.groupby(['Ano_Semana', 'Num_Semana', 'Periodo', 'Obra_Vinculada']).agg(
                        Caixas=('Qtd_Caixas', 'sum'), M2=('M2_Item', 'sum')
                    ).reset_index().sort_values(['Ano_Semana', 'Num_Semana'])
                    res_fab.columns = ['Ano', 'Sem', 'Periodo', 'Obra', 'Caixas (cx)', 'Metragem (m²)']
                    st.dataframe(res_fab[['Periodo', 'Obra', 'Caixas (cx)', 'Metragem (m²)']], hide_index=True, use_container_width=True)
                else:
                    st.success("Fabrica livre!")

    # ==================================================
    # LOGISTICA
    # ==================================================
    elif nome_aba == "Logistica":
        with aba_objeto:
            st.markdown('<div class="page-header"><div class="page-header-left"><h2>Logística</h2><p>Gestão de despachos, transportes e envios</p></div><span class="page-icon">🚚</span></div>', unsafe_allow_html=True)
            st.caption(f"Hoje: {hoje_projeto().strftime('%d/%m/%Y')}")
            df_log = carregar_fila_logistica()

            if not df_log.empty:
                n_aguard   = len(df_log[df_log['Status_Logistica'] == 'Aguardando Agendamento'])
                n_agendado = len(df_log[df_log['Status_Logistica'] == 'Envio Agendado'])
                n_despach  = len(df_log[df_log['Status_Logistica'] == 'Despachado'])
                df_ativos  = df_log[df_log['Status_Logistica'] != 'Despachado']
                n_atrasado = (
                    len(df_ativos[df_ativos['Data_Limite_Despacho'].apply(
                        lambda x: prazo_valido(x) and pd.to_datetime(x) < hoje_projeto()
                    )]) if not df_ativos.empty else 0
                )
                c1, c2, c3, c4 = st.columns(4)
                c1.metric("Aguardando Agendamento", n_aguard)
                c2.metric("Envios Agendados",       n_agendado)
                c3.metric("Despachados",             n_despach)
                if n_atrasado > 0:
                    c4.metric("Atrasados", n_atrasado, delta=f"-{n_atrasado}", delta_color="inverse")
                else:
                    c4.metric("Atrasos", "Nenhum")
            else:
                st.info("Nenhum lote na fila ainda.")

            st.markdown("---")

            # ── OPs FINALIZADAS — EMITIR ROMANEIO ─────────────────
            with st.expander("✅ OPs Finalizadas — Emitir Romaneio", expanded=True):
                df_micro_completo_log = carregar_micro_completo()
                df_conc_log = df_micro_completo_log[
                    df_micro_completo_log['Status_Item'] == 'Concluido'
                ].copy() if not df_micro_completo_log.empty else pd.DataFrame()

                if obra_selecionada and not df_conc_log.empty:
                    df_conc_log = df_conc_log[df_conc_log['Obra_Vinculada'] == obra_selecionada]

                if df_conc_log.empty:
                    st.info("Nenhuma OP finalizada ainda.")
                else:
                    for _, row_c in df_conc_log.iterrows():
                        df_pecas_c = carregar_pecas_lote(int(row_c['id']))
                        with st.container(border=True):
                            cc1, cc2, cc3 = st.columns([4, 2, 2])
                            with cc1:
                                st.markdown(
                                    f'<span class="badge-obra">{row_c["Obra_Vinculada"]}</span>&nbsp;'
                                    f'<span class="badge-edt">{row_c["EDT_Vinculado"]}</span>&nbsp;'
                                    f'<span class="badge-lote">{row_c["Cod_Lote"]}</span>',
                                    unsafe_allow_html=True
                                )
                                num_op_c = row_c.get('Num_OP') or '—'
                                st.markdown(f"**OP:** `{num_op_c}` &nbsp;|&nbsp; **{row_c['Tipo_Material']}**")
                                st.caption(f"{row_c['M2_Item']:.2f} m² &nbsp;|&nbsp; {int(row_c['Qtd_Caixas'])} cx &nbsp;|&nbsp; {row_c['Romaneio_Chapas']}")
                                if not df_pecas_c.empty:
                                    st.caption(f"🔩 {len(df_pecas_c)} peça(s) | Total: {int(df_pecas_c['qtd_total'].sum())} un")
                                else:
                                    st.caption("⚠️ Sem peças lançadas")
                            with cc2:
                                end_r = st.text_input("Endereço:", key=f"end_rom_{row_c['id']}",
                                                       placeholder="Endereço da obra")
                            with cc3:
                                st.write("")
                                st.write("")
                                if df_pecas_c.empty:
                                    st.warning("Sem peças para o romaneio")
                                else:
                                    rom_bytes = gerar_romaneio_xlsx(
                                        row_c, df_pecas_c, end_r,
                                        st.session_state.usuario_nome
                                    )
                                    st.download_button(
                                        label="🖨️ Emitir Romaneio",
                                        data=rom_bytes,
                                        file_name=f"Romaneio_{num_op_c}_{row_c['Cod_Lote']}.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        key=f"dl_rom_{row_c['id']}"
                                    )

            with st.expander("Fila Prioritaria — Aguardando Agendamento", expanded=True):
                if df_log.empty or df_log[df_log['Status_Logistica'] == 'Aguardando Agendamento'].empty:
                    st.success("Todos os lotes ja agendados!")
                else:
                    df_ag = df_log[df_log['Status_Logistica'] == 'Aguardando Agendamento'].copy().sort_values('Data_Limite_Despacho', na_position='last')
                    for _, row in df_ag.iterrows():
                        prazo_d = row['Data_Limite_Despacho']
                        if prazo_valido(prazo_d):
                            dias_r = (pd.to_datetime(prazo_d) - hoje_projeto()).days
                            if dias_r < 0:    css_bar = "bar-danger"; tag = f"ATRASADO {abs(dias_r)}d"
                            elif dias_r <= 3: css_bar = "bar-warn";   tag = f"URGENTE — {dias_r}d restantes"
                            else:             css_bar = "bar-ok";     tag = f"{dias_r} dias restantes"
                        else:
                            css_bar = "bar-neutral"; tag = "Sem prazo definido"

                        st.markdown(f"<div class='{css_bar}'>", unsafe_allow_html=True)
                        ci, cp, ca = st.columns([5, 3, 2])
                        with ci:
                            num_op_ag = row.get('Num_OP') or 'S/OP'
                            st.markdown(
                                f'<span class="badge-obra">{row["Obra_Vinculada"]}</span>&nbsp;'
                                f'<span class="badge-lote">Lote: {row["Cod_Lote"]}</span>&nbsp;'
                                f'<span class="badge-edt">OP: {num_op_ag}</span>',
                                unsafe_allow_html=True
                            )
                            st.markdown(f"**{row['Tipo_Material']}** | `{int(row['Qtd_Caixas'])} cx` — {row['M2_Item']:.2f} m²")
                            st.caption(f"Pavimentos: {row['Romaneio_Chapas']}")
                        with cp:
                            ptxt = pd.to_datetime(prazo_d).strftime('%d/%m/%Y') if prazo_valido(prazo_d) else "—"
                            st.caption("Prazo maximo despacho")
                            st.markdown(f"**{ptxt}**")
                            st.markdown(f"`{tag}`")
                        with ca:
                            if st.button("Agendar", key=f"ag_btn_{row['id']}", use_container_width=True):
                                st.session_state[f"ag_open_{row['id']}"] = True
                        st.markdown("</div>", unsafe_allow_html=True)

                        if st.session_state.get(f"ag_open_{row['id']}", False):
                            with st.container(border=True):
                                st.markdown(f"#### Agendar — `{num_op_ag}` | Lote `{row['Cod_Lote']}` | {row['Obra_Vinculada']}")
                                fa1, fa2 = st.columns(2)
                                with fa1:
                                    dt_env = st.date_input("Data envio:", format="DD/MM/YYYY", key=f"dt_env_{row['id']}",
                                                           value=(pd.to_datetime(prazo_d).date() if prazo_valido(prazo_d) else hoje_projeto().date()))
                                    transp = st.selectbox("Transporte:", ["Frota Propria (Passold)", "Transportadora Terceira", "Retirada pelo Cliente"], key=f"tr_{row['id']}")
                                with fa2:
                                    veic = st.text_input("Veiculo / Placa:", key=f"ve_{row['id']}")
                                    obs  = st.text_area("Observacoes:", key=f"ob_{row['id']}", height=80)
                                cb1, cb2 = st.columns(2)
                                with cb1:
                                    if st.button("Confirmar agendamento", key=f"conf_{row['id']}", use_container_width=True, type="primary"):
                                        agendar_envio(row['id'], dt_env, transp, veic, obs, st.session_state.usuario_nome)
                                        registrar_auditoria(st.session_state.usuario_nome, "AGENDAR_ENVIO",
                                            f"Lote {row['Cod_Lote']} — {transp} — {dt_env}")
                                        st.session_state[f"ag_open_{row['id']}"] = False
                                        st.toast(f"Agendado para {dt_env.strftime('%d/%m/%Y')}!")
                                        time.sleep(0.3)
                                        st.rerun()
                                with cb2:
                                    if st.button("Cancelar", key=f"can_{row['id']}", use_container_width=True):
                                        st.session_state[f"ag_open_{row['id']}"] = False
                                        st.rerun()

            with st.expander("Envios Agendados — Confirmar Saida", expanded=True):
                if df_log.empty or df_log[df_log['Status_Logistica'] == 'Envio Agendado'].empty:
                    st.info("Nenhum envio agendado.")
                else:
                    df_agend = df_log[df_log['Status_Logistica'] == 'Envio Agendado'].copy().sort_values('Data_Envio_Agendado', na_position='last')
                    for _, row in df_agend.iterrows():
                        pd_d = row['Data_Limite_Despacho']
                        de_d = row['Data_Envio_Agendado']
                        no_prazo = (pd.to_datetime(de_d) <= pd.to_datetime(pd_d)) if prazo_valido(pd_d) and prazo_valido(de_d) else True
                        css_bar  = "bar-ok" if no_prazo else "bar-danger"
                        st.markdown(f"<div class='{css_bar}'>", unsafe_allow_html=True)
                        ci2, cp2, ca2 = st.columns([5, 3, 2])
                        with ci2:
                            st.markdown(f'<span class="badge-obra">{row["Obra_Vinculada"]}</span>&nbsp;<span class="badge-lote">Lote: {row["Cod_Lote"]}</span>', unsafe_allow_html=True)
                            st.markdown(f"**{row['Tipo_Material']}** | `{int(row['Qtd_Caixas'])} cx` — {row['M2_Item']:.2f} m²")
                            st.caption(f"{row.get('Transportadora', '—')} | {row.get('Veiculo', '—')}")
                            if row.get('Observacoes'):
                                st.caption(f"Obs: {row['Observacoes']}")
                        with cp2:
                            et = pd.to_datetime(de_d).strftime('%d/%m/%Y') if prazo_valido(de_d) else "—"
                            pt = pd.to_datetime(pd_d).strftime('%d/%m/%Y') if prazo_valido(pd_d) else "—"
                            st.caption("Data envio agendada")
                            st.markdown(f"**{et}**")
                            st.caption("Prazo maximo")
                            st.write(pt)
                            if not no_prazo:
                                st.error("Fora do prazo!")
                        with ca2:
                            st.write("")
                            if st.button("Confirmar Despacho", key=f"des_{row['id']}", use_container_width=True, type="primary"):
                                confirmar_despacho(row['id'], st.session_state.usuario_nome)
                                registrar_auditoria(st.session_state.usuario_nome, "CONFIRMAR_DESPACHO",
                                    f"Lote {row['Cod_Lote']} — Obra {row.get('Obra_Vinculada','—')}")
                                st.toast("Despachado!")
                                time.sleep(0.3)
                                st.rerun()
                            if st.button("Reagendar", key=f"rag_{row['id']}", use_container_width=True):
                                conn = conectar_banco()
                                try:
                                    cursor = conn.cursor()
                                    cursor.execute("UPDATE logistica_envios SET Status_Logistica='Aguardando Agendamento' WHERE id=%s", (row['id'],))
                                    conn.commit()
                                    carregar_fila_logistica.clear()
                                except Exception as e:
                                    conn.rollback()
                                    st.error(f"Erro: {e}")
                                finally:
                                    liberar_conexao(conn)
                                st.rerun()
                        st.markdown("</div>", unsafe_allow_html=True)

            with st.expander("Historico de Despachos", expanded=False):
                if df_log.empty or df_log[df_log['Status_Logistica'] == 'Despachado'].empty:
                    st.info("Nenhum despacho realizado ainda.")
                else:
                    df_hist = df_log[df_log['Status_Logistica'] == 'Despachado'].copy()
                    for col in ['Data_Limite_Despacho', 'Data_Envio_Agendado']:
                        df_hist[col] = df_hist[col].apply(lambda x: pd.to_datetime(x).strftime('%d/%m/%Y') if prazo_valido(x) else "—")
                    cols_h = [c for c in ['Obra_Vinculada', 'Cod_Lote', 'Tipo_Material', 'Qtd_Caixas', 'M2_Item',
                                          'Transportadora', 'Veiculo', 'Data_Envio_Agendado', 'Data_Limite_Despacho',
                                          'Confirmado_Por', 'Confirmado_Em'] if c in df_hist.columns]
                    st.dataframe(df_hist[cols_h], hide_index=True, use_container_width=True)
                    df_pont = df_log[df_log['Status_Logistica'] == 'Despachado'].copy()
                    df_pont = df_pont[df_pont['Data_Limite_Despacho'].apply(prazo_valido) & df_pont['Data_Envio_Agendado'].apply(prazo_valido)]
                    if not df_pont.empty:
                        ok = (pd.to_datetime(df_pont['Data_Envio_Agendado']) <= pd.to_datetime(df_pont['Data_Limite_Despacho'])).sum()
                        st.metric("Pontualidade nos despachos", f"{ok / len(df_pont) * 100:.0f}%")

    # ==================================================
    # ALMOXARIFADO
    # ==================================================
    elif nome_aba == "Almoxarifado":
        with aba_objeto:
            st.markdown('<div class="page-header"><div class="page-header-left"><h2>Almoxarifado</h2><p>Conferência e controle de componentes recebidos</p></div><span class="page-icon">📦</span></div>', unsafe_allow_html=True)
            st.caption(f"Hoje: {hoje_projeto().strftime('%d/%m/%Y')} | Usuário: {st.session_state.usuario_nome}")
            df_ops_comp = carregar_todas_ops_com_componentes()

            if df_ops_comp.empty:
                st.info("Nenhuma OP com lista de componentes cadastrada ainda.")
            else:
                conn_alm = conectar_banco()
                try:
                    todos_comps = pd.read_sql_query("SELECT * FROM componentes_op ORDER BY item_id, id", conn_alm)
                finally:
                    liberar_conexao(conn_alm)

                n_aguard  = len(todos_comps[todos_comps['status_item'] == 'Aguardando Conferencia'])
                n_ok      = len(todos_comps[todos_comps['status_item'] == 'Disponivel'])
                n_falta   = len(todos_comps[todos_comps['status_item'] == 'Indisponivel'])
                c1, c2, c3 = st.columns(3)
                c1.metric("⏳ Aguardando", n_aguard)
                c2.metric("✅ Disponíveis", n_ok)
                c3.metric("❌ Indisponíveis", n_falta)
                st.markdown("---")

                for _, op_row in df_ops_comp.iterrows():
                    df_comp = carregar_componentes_op(int(op_row['item_id']))
                    if df_comp.empty:
                        continue
                    n_total   = len(df_comp)
                    n_conf    = len(df_comp[df_comp['status_item'] != 'Aguardando Conferencia'])
                    n_indisp  = len(df_comp[df_comp['status_item'] == 'Indisponivel'])
                    if n_indisp > 0:   css_bar = "bar-danger"; icone = "❌"
                    elif n_conf == n_total: css_bar = "bar-ok"; icone = "✅"
                    else:              css_bar = "bar-warn";  icone = "⏳"

                    with st.expander(
                        f"{icone} OP: {op_row['num_op']} — {op_row['cod_lote']} | {op_row['obra_vinculada']}  "
                        f"({n_conf}/{n_total} conferidos{f' — {n_indisp} FALTANDO' if n_indisp > 0 else ''})",
                        expanded=(n_indisp > 0 or n_conf < n_total)
                    ):
                        hc = st.columns([4, 2, 2, 3, 2])
                        for col_h, label in zip(hc, ["COMPONENTE", "QTD", "UN", "STATUS", "AÇÃO"]):
                            col_h.markdown(f"<div style='font-size:11px;font-weight:700;color:#94A3B8;text-transform:uppercase;letter-spacing:0.07em;'>{label}</div>", unsafe_allow_html=True)
                        st.markdown("<hr style='margin:4px 0 8px 0;border-color:#E2E8F0;'>", unsafe_allow_html=True)

                        for _, comp in df_comp.iterrows():
                            st_item = comp['status_item']
                            if st_item == 'Disponivel':     cor = "#15803D"; bg = "#F0FDF4"; emoji = "✅"
                            elif st_item == 'Indisponivel': cor = "#DC2626"; bg = "#FEF2F2"; emoji = "❌"
                            else:                           cor = "#D97706"; bg = "#FFFBEB"; emoji = "⏳"
                            rc = st.columns([4, 2, 2, 3, 2])
                            rc[0].markdown(f"**{comp['nome_componente']}**")
                            rc[1].markdown(f"`{comp['quantidade']}`")
                            rc[2].markdown(f"{comp['unidade']}")
                            rc[3].markdown(f"<span style='background:{bg};color:{cor};padding:3px 8px;border-radius:4px;font-size:12px;font-weight:600;'>{emoji} {st_item}</span>", unsafe_allow_html=True)
                            with rc[4]:
                                acao = st.selectbox(
                                    "", ["Aguardando Conferencia", "Disponivel", "Indisponivel"],
                                    index=["Aguardando Conferencia", "Disponivel", "Indisponivel"].index(st_item),
                                    key=f"alm_st_{comp['id']}", label_visibility="collapsed"
                                )
                                if acao != st_item:
                                    atualizar_componente(comp['id'], acao, comp.get('observacao') or '', st.session_state.usuario_nome)
                                    st.rerun()
                            obs_atual = comp.get('observacao') or ''
                            obs_nova  = st.text_input(f"Obs — {comp['nome_componente']}:", value=obs_atual,
                                                      key=f"alm_obs_{comp['id']}", placeholder="Ex: em falta, previsão 20/06...")
                            if obs_nova != obs_atual:
                                atualizar_componente(comp['id'], st_item, obs_nova, st.session_state.usuario_nome)
                            st.markdown("<hr style='margin:4px 0;border-color:#F1F5F9;'>", unsafe_allow_html=True)

                        if n_indisp > 0:
                            itens_falt = df_comp[df_comp['status_item'] == 'Indisponivel']['nome_componente'].tolist()
                            st.error(f"⚠️ Itens em falta: {', '.join(itens_falt)}")
                        elif n_conf == n_total:
                            st.success("✅ Todos os componentes conferidos e disponíveis!")

    # ==================================================
    # ROMANEIO MANUAL (sem OP vinculada)
    # ==================================================
    elif nome_aba == "Romaneio Manual":
        with aba_objeto:
            st.markdown('<div class="page-header"><div class="page-header-left"><h2>Romaneio Manual</h2><p>Registro de recebimento de material sem OP vinculada</p></div><span class="page-icon">📋</span></div>', unsafe_allow_html=True)
            st.caption(f"Hoje: {hoje_projeto().strftime('%d/%m/%Y')} | Usuário: {st.session_state.usuario_nome}")

            if "rom_itens" not in st.session_state:
                st.session_state.rom_itens = []

            obras_rom = sorted(df_banco_macro['Obra'].dropna().unique().tolist()) if not df_banco_macro.empty else []
            if not obras_rom:
                st.info("Nenhuma obra cadastrada ainda.")
            else:
                rm1, rm2 = st.columns(2)
                with rm1:
                    rom_obra = st.selectbox("Obra:", obras_rom, key="rom_obra")
                with rm2:
                    rom_data = st.date_input("Data de recebimento:", value=hoje_projeto().date(),
                                              format="DD/MM/YYYY", key="rom_data")

                st.markdown("**Adicionar item ao romaneio:**")
                ri1, ri2, ri3 = st.columns([5, 2, 1])
                with ri1:
                    rom_desc = st.text_input("Descrição do material:", key="rom_desc",
                                              placeholder="Ex: Perfil de alumínio natural 6m")
                with ri2:
                    rom_qtd = st.number_input("Quantidade:", min_value=0.0, value=1.0, step=1.0, key="rom_qtd")
                with ri3:
                    st.write("")
                    st.write("")
                    if st.button("➕ Adicionar", key="btn_add_rom_item"):
                        if not rom_desc.strip():
                            st.error("Informe a descrição do material antes de adicionar.")
                        else:
                            st.session_state.rom_itens.append({"descricao": rom_desc.strip(), "quantidade": rom_qtd})
                            st.rerun()

                if st.session_state.rom_itens:
                    st.markdown("**Itens do romaneio:**")
                    for i, item in enumerate(st.session_state.rom_itens):
                        col_desc, col_qtd, col_rem = st.columns([5, 2, 1])
                        col_desc.write(f"{i+1}. {item['descricao']}")
                        col_qtd.write(f"{item['quantidade']:g}")
                        with col_rem:
                            if st.button("🗑", key=f"rem_rom_item_{i}"):
                                st.session_state.rom_itens.pop(i)
                                st.rerun()

                    if st.button("💾 Salvar Romaneio", type="primary", key="btn_salvar_rom_manual"):
                        romaneio_id = salvar_romaneio_manual(rom_obra, rom_data, st.session_state.rom_itens, st.session_state.usuario_nome)
                        if romaneio_id:
                            registrar_auditoria(st.session_state.usuario_nome, "ROMANEIO_MANUAL",
                                f"Romaneio manual #{romaneio_id} — Obra: {rom_obra} — {len(st.session_state.rom_itens)} item(ns)")
                            st.session_state.rom_itens = []
                            st.toast("Romaneio salvo!")
                            time.sleep(0.3)
                            st.rerun()
                else:
                    st.caption("Nenhum item adicionado ainda.")

            st.markdown("---")
            st.markdown("### Histórico de Envios")
            df_rom_manual = carregar_romaneios_manuais()
            if df_rom_manual.empty:
                st.caption("Nenhum romaneio manual cadastrado ainda.")
            else:
                obras_hist = ["Todas"] + sorted(df_rom_manual['obra_vinculada'].dropna().unique().tolist())
                filtro_obra_hist = st.selectbox("Filtrar por obra:", obras_hist, key="filtro_obra_rom_hist")
                if filtro_obra_hist != "Todas":
                    df_rom_manual = df_rom_manual[df_rom_manual['obra_vinculada'] == filtro_obra_hist]

                if df_rom_manual.empty:
                    st.caption("Nenhum romaneio para essa obra ainda.")
                for _, rm_row in df_rom_manual.iterrows():
                    with st.expander(
                        f"{rm_row['obra_vinculada']} — {pd.to_datetime(rm_row['data_recebimento']).strftime('%d/%m/%Y')} "
                        f"({int(rm_row['qtd_itens'])} item(ns)) — por {rm_row['criado_por']}"
                    ):
                        df_itens_rom = carregar_itens_romaneio_manual(int(rm_row['id']))
                        st.dataframe(df_itens_rom[['descricao', 'quantidade']], hide_index=True, use_container_width=True)
                        rmb1, rmb2 = st.columns(2)
                        with rmb1:
                            rom_bytes = gerar_romaneio_manual_xlsx(
                                rm_row['obra_vinculada'], rm_row['data_recebimento'], df_itens_rom, rm_row['criado_por']
                            )
                            st.download_button(
                                "🖨️ Emitir Romaneio", data=rom_bytes,
                                file_name=f"Romaneio_Manual_{rm_row['obra_vinculada']}_{pd.to_datetime(rm_row['data_recebimento']).strftime('%Y%m%d')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"dl_rom_manual_{rm_row['id']}"
                            )
                        with rmb2:
                            if setor in ["Master", "Almoxarifado"]:
                                if st.button("🗑️ Excluir", key=f"del_rom_manual_{rm_row['id']}"):
                                    excluir_romaneio_manual(int(rm_row['id']))
                                    st.toast("Romaneio removido.")
                                    time.sleep(0.3)
                                    st.rerun()

    # ==================================================
    # SISTEMA DE MEDICAO — agora 100% no banco
    # ==================================================
    elif nome_aba == "Sistema de Medicao":
        with aba_objeto:
            st.header("Sistema de Medição — Gestão de Saldos & Escopos")
            st.caption(f"Usuário: {st.session_state.usuario_nome}")

            import json as _json

            tab_cad, tab_srv, tab_med, tab_dash = st.tabs([
                "1. Cadastrar Obra Base",
                "2. Serviços & Etapas",
                "3. Relatório de Saldos",
                "4. Dashboards"
            ])

            # ── ABA 1: CADASTRAR OBRA ─────────────────────────────
            with tab_cad:
                with st.form("form_med_obra"):
                    st.markdown("#### Nova Obra / Contrato")
                    mc1, mc2, mc3 = st.columns(3)
                    with mc1:
                        med_nome = st.text_input("Nome da Obra / Contrato")
                    with mc2:
                        med_valm2 = st.number_input("Valor m² Teto (R$/m²)", min_value=0.01, value=500.0)
                    with mc3:
                        med_metro = st.number_input("Metragem Geral (m²)", min_value=0.01, value=1500.0)
                    if st.form_submit_button("Salvar Obra"):
                        if not med_nome.strip():
                            st.error("Informe o nome da obra.")
                        else:
                            with st.spinner("Salvando..."):
                                new_id = salvar_medicao_obra(med_nome.strip(), med_valm2, med_metro)
                            if new_id:
                                st.success(f"Obra '{med_nome}' cadastrada!")
                                st.rerun()

                st.markdown("---")
                st.markdown("#### Obras Cadastradas")
                df_med_obras = carregar_medicao_obras()
                if df_med_obras.empty:
                    st.info("Nenhuma obra cadastrada ainda.")
                else:
                    cols_obra = st.columns(min(len(df_med_obras), 3))
                    for i, (_, obra_row) in enumerate(df_med_obras.iterrows()):
                        total = obra_row['valor_m2_global'] * obra_row['metragem_geral']
                        with cols_obra[i % 3]:
                            with st.container(border=True):
                                st.markdown(f"**{obra_row['nome']}**")
                                st.caption(f"Metragem: {obra_row['metragem_geral']:.2f} m²")
                                st.caption(f"Valor limite m²: R$ {obra_row['valor_m2_global']:.2f}")
                                st.metric("Total Contratado", f"R$ {total:,.2f}")
                                if st.button("Excluir", key=f"del_med_obra_{obra_row['id']}", type="primary"):
                                    excluir_medicao_obra(int(obra_row['id']))
                                    st.toast("Obra removida!")
                                    time.sleep(0.3)
                                    st.rerun()

            # ── ABA 2: SERVIÇOS & ETAPAS ──────────────────────────
            with tab_srv:
                df_med_obras2 = carregar_medicao_obras()
                if df_med_obras2.empty:
                    st.info("Cadastre uma obra primeiro.")
                else:
                    opcoes_med = {row['nome']: row['id'] for _, row in df_med_obras2.iterrows()}
                    obra_med_sel = st.selectbox("Obra:", list(opcoes_med.keys()), key="sel_obra_srv")
                    obra_med_id  = opcoes_med[obra_med_sel]
                    periodo_ref  = st.text_input("Período de referência (MM/AAAA):",
                                                  value=datetime.now(FUSO_BR).strftime('%m/%Y'),
                                                  placeholder="Ex: 06/2025")

                    obra_med_info = df_med_obras2[df_med_obras2['id'] == obra_med_id].iloc[0]
                    st.info(f"Teto da obra: **R$ {obra_med_info['valor_m2_global']:.2f}/m²**")

                    df_srvs = carregar_medicao_servicos(obra_med_id)
                    if st.session_state.get("med_obra_id_atual") != obra_med_id:
                        st.session_state.med_obra_id_atual = obra_med_id
                        st.session_state.med_num_servicos = max(1, len(df_srvs))
                        st.session_state.med_servicos_excluidos = set()
                    elif "med_num_servicos" not in st.session_state:
                        st.session_state.med_num_servicos = max(1, len(df_srvs))
                        st.session_state.med_servicos_excluidos = set()

                    ativos = st.session_state.med_num_servicos - len(st.session_state.med_servicos_excluidos)
                    st.markdown(f"#### Serviços ({ativos})")

                    servicos_form = []
                    for si in range(st.session_state.med_num_servicos):
                        if si in st.session_state.med_servicos_excluidos:
                            continue
                        srv_existente = df_srvs.iloc[si] if si < len(df_srvs) else None
                        with st.expander(f"Serviço {si+1}", expanded=(si == 0)):
                            s1, s2 = st.columns(2)
                            with s1:
                                srv_nome = st.text_input("Nome:", value=srv_existente['nome'] if srv_existente is not None else "", key=f"srv_nome_{si}")
                            with s2:
                                srv_val  = st.number_input("Valor m²:", min_value=0.0, value=float(srv_existente['valor_m2_servico']) if srv_existente is not None else 0.0, key=f"srv_val_{si}")

                            df_subs = carregar_medicao_subdivisoes(int(srv_existente['id'])) if srv_existente is not None else pd.DataFrame()
                            if "med_num_subs" not in st.session_state:
                                st.session_state[f"med_num_subs_{si}"] = max(1, len(df_subs))
                            n_subs = st.number_input(f"Qtd subdivisões:", min_value=1, max_value=20, value=st.session_state.get(f"med_num_subs_{si}", max(1, len(df_subs))), key=f"n_subs_{si}")

                            subdivisoes_form = []
                            for subi in range(int(n_subs)):
                                sub_existente = df_subs.iloc[subi] if subi < len(df_subs) else None
                                sb1, sb2, sb3, sb4 = st.columns([3, 2, 2, 2])
                                with sb1:
                                    sub_nome = st.text_input("Subdivisão:", value=sub_existente['nome'] if sub_existente is not None else "", key=f"sub_nome_{si}_{subi}")
                                with sb2:
                                    sub_m2   = st.number_input("m²:", min_value=0.0, value=float(sub_existente['m2']) if sub_existente is not None else 0.0, key=f"sub_m2_{si}_{subi}")
                                with sb3:
                                    sub_pct  = st.number_input("% serv:", min_value=0.0, max_value=100.0, value=float(sub_existente['percentual']) if sub_existente is not None else 0.0, key=f"sub_pct_{si}_{subi}")
                                with sb4:
                                    val_calc = srv_val * (sub_pct / 100)
                                    sub_total = sub_m2 * val_calc
                                    st.metric("Subtotal", f"R$ {sub_total:,.2f}")
                                if sub_nome.strip():
                                    subdivisoes_form.append({"nome": sub_nome, "m2": sub_m2, "percentual": sub_pct})

                            col_excluir = st.columns([1])[0]
                            with col_excluir:
                                if st.button(f"🗑️ Excluir Serviço {si+1}", key=f"excluir_srv_{si}", type="secondary"):
                                    st.session_state.med_servicos_excluidos.add(si)
                                    st.rerun()

                            if srv_nome.strip():
                                servicos_form.append({"nome": srv_nome, "valor_m2_servico": srv_val, "subdivisoes": subdivisoes_form})

                    col_add, col_save = st.columns([1, 2])
                    with col_add:
                        if st.button("+ Serviço"):
                            st.session_state.med_num_servicos += 1
                            st.rerun()
                    with col_save:
                        if st.button("💾 Salvar Lançamentos", type="primary"):
                            import re as _re
                            _per = periodo_ref.strip()
                            if not _per:
                                st.error("Informe o período.")
                            elif not _re.match(r"^\d{2}/\d{4}$", _per):
                                st.error("Formato inválido. Use MM/AAAA — ex: 06/2025")
                            elif not servicos_form:
                                st.error("Adicione pelo menos um serviço.")
                            else:
                                with st.spinner("Salvando..."):
                                    total = salvar_servicos_medicao(obra_med_id, periodo_ref.strip(), servicos_form)
                                if total is not None:
                                    st.success(f"Salvo! Total medido: R$ {total:,.2f}")
                                    st.rerun()

            # ── ABA 3: RELATÓRIO DE SALDOS ────────────────────────
            with tab_med:
                df_med_obras3 = carregar_medicao_obras()
                if df_med_obras3.empty:
                    st.info("Nenhuma obra cadastrada.")
                else:
                    opcoes_med3 = {row['nome']: row['id'] for _, row in df_med_obras3.iterrows()}
                    obra_med3   = st.selectbox("Obra:", list(opcoes_med3.keys()), key="sel_obra_med")
                    obra_id3    = opcoes_med3[obra_med3]
                    df_hist3    = carregar_medicao_historico(obra_id3)

                    if df_hist3.empty:
                        st.info("Nenhuma medição salva para esta obra.")
                    else:
                        periodos = df_hist3['periodo'].tolist()
                        per_sel  = st.selectbox("Período:", periodos, index=len(periodos)-1, key="sel_per_med")
                        hist_row = df_hist3[df_hist3['periodo'] == per_sel].iloc[0]
                        snapshot = hist_row['snapshot'] if isinstance(hist_row['snapshot'], list) else (_json.loads(hist_row['snapshot']) if isinstance(hist_row['snapshot'], str) else [])
                        obra_info3 = df_med_obras3[df_med_obras3['id'] == obra_id3].iloc[0]
                        total_contratado = obra_info3['valor_m2_global'] * obra_info3['metragem_geral']

                        if snapshot:
                            st.markdown("---")
                            # Tabela
                            rows_table = []
                            for srv in snapshot:
                                for sub in srv.get('subdivisoes', []):
                                    preco_m2 = srv['valor_m2_servico'] * (sub['percentual'] / 100)
                                    rows_table.append({
                                        "Serviço": srv['nome'],
                                        "Subdivisão": f"{sub['nome']} ({sub['percentual']:.0f}%)",
                                        "m²": f"{sub['m2']:.2f}",
                                        "R$/m²": f"R$ {preco_m2:.2f}",
                                        "Subtotal": f"R$ {sub.get('subtotal', sub['m2']*preco_m2):,.2f}"
                                    })
                            if rows_table:
                                st.dataframe(pd.DataFrame(rows_table), hide_index=True, use_container_width=True)

                        mc1, mc2 = st.columns(2)
                        mc1.metric("Total Contratado", f"R$ {total_contratado:,.2f}")
                        mc2.metric(f"Medição — {per_sel}", f"R$ {hist_row['total_medido']:,.2f}")

            # ── ABA 4: DASHBOARDS ─────────────────────────────────
            with tab_dash:
                df_med_obras4 = carregar_medicao_obras()
                if df_med_obras4.empty:
                    st.info("Nenhuma obra cadastrada.")
                else:
                    opcoes_med4 = {row['nome']: row['id'] for _, row in df_med_obras4.iterrows()}
                    obra_med4   = st.selectbox("Obra:", list(opcoes_med4.keys()), key="sel_obra_dash")
                    obra_id4    = opcoes_med4[obra_med4]
                    df_hist4    = carregar_medicao_historico(obra_id4)
                    obra_info4  = df_med_obras4[df_med_obras4['id'] == obra_id4].iloc[0]
                    total_contratado4 = obra_info4['valor_m2_global'] * obra_info4['metragem_geral']

                    if df_hist4.empty:
                        st.info("Nenhuma medição salva para esta obra.")
                    else:
                        # Gráfico de evolução
                        fig_evo = px.bar(
                            df_hist4, x='periodo', y='total_medido',
                            title="Evolução de Medições por Período",
                            labels={'periodo': 'Período', 'total_medido': 'R$ Medido'},
                            color_discrete_sequence=["#1E3A8A"]
                        )
                        fig_evo.add_hline(
                            y=total_contratado4, line_dash="dash",
                            line_color="#1A56DB", annotation_text="Total Contratado"
                        )
                        fig_evo.update_layout(height=350, margin=dict(l=20, r=20, t=40, b=20))
                        st.plotly_chart(fig_evo, use_container_width=True)

                        # Acumulado e saldo
                        total_medido_acum = df_hist4['total_medido'].sum()
                        saldo = total_contratado4 - total_medido_acum

                        d1, d2, d3 = st.columns(3)
                        d1.metric("Total Contratado", f"R$ {total_contratado4:,.2f}")
                        d2.metric("Total Medido (acumulado)", f"R$ {total_medido_acum:,.2f}")
                        d3.metric("Saldo Remanescente", f"R$ {saldo:,.2f}",
                                  delta=f"{(total_medido_acum/total_contratado4*100):.1f}% executado" if total_contratado4 > 0 else "—")

                        st.markdown("---")
                        st.markdown("#### Tabela Comparativa por Período")
                        rows_comp = []
                        for _, h in df_hist4.iterrows():
                            snap = h['snapshot'] if isinstance(h['snapshot'], list) else (_json.loads(h['snapshot']) if isinstance(h['snapshot'], str) else [])
                            for srv in snap:
                                for sub in srv.get('subdivisoes', []):
                                    preco_m2 = srv['valor_m2_servico'] * (sub['percentual'] / 100)
                                    rows_comp.append({
                                        "Período": h['periodo'],
                                        "Serviço": srv['nome'],
                                        "Subdivisão": sub['nome'],
                                        "m²": sub['m2'],
                                        "R$/m²": preco_m2,
                                        "Subtotal (R$)": sub.get('subtotal', sub['m2'] * preco_m2)
                                    })
                        if rows_comp:
                            df_comp_tab = pd.DataFrame(rows_comp)
                            st.dataframe(
                                df_comp_tab.style.format({"m²": "{:.2f}", "R$/m²": "R$ {:.2f}", "Subtotal (R$)": "R$ {:,.2f}"}),
                                hide_index=True, use_container_width=True
                            )

    # ==================================================
    # CONFIGURACOES
    # ==================================================
    elif nome_aba == "Configuracoes":
        with aba_objeto:
            if setor != "Master":
                st.error("⛔ Acesso negado. Apenas usuários Master podem acessar esta página.")
                st.stop()
            st.markdown('<div class="page-header"><div class="page-header-left"><h2>Configurações</h2><p>Painel de controle e administração do sistema</p></div><span class="page-icon">⚙️</span></div>', unsafe_allow_html=True)

            with st.expander("Cadastrar Novo Usuario"):
                with st.form("form_user"):
                    nu = st.text_input("Login:").lower().strip()
                    nn = st.text_input("Nome:")
                    ns = st.selectbox("Setor:", ["Producao", "Engenharia", "Diretoria", "Logistica", "Almoxarifado", "Medicao", "PCP", "Master"])
                    np = st.text_input("Senha:", type="password")
                    st.caption("Mínimo 8 caracteres, ao menos 1 número e 1 letra maiúscula.")
                    if st.form_submit_button("Salvar"):
                        _erros_senha = []
                        if not all([nu, nn, np.strip()]):
                            _erros_senha.append("Preencha todos os campos.")
                        if len(np) < 8:
                            _erros_senha.append("Senha deve ter no mínimo 8 caracteres.")
                        if not any(c.isupper() for c in np):
                            _erros_senha.append("Senha deve conter ao menos 1 letra maiúscula.")
                        if not any(c.isdigit() for c in np):
                            _erros_senha.append("Senha deve conter ao menos 1 número.")
                        if _erros_senha:
                            for _e in _erros_senha:
                                st.error(_e)
                        else:
                            conn = conectar_banco()
                            try:
                                cursor = conn.cursor()
                                cursor.execute(
                                    "INSERT INTO usuarios (usuario, nome, setor, senha) VALUES (%s,%s,%s,%s)",
                                    (nu, nn, ns, hash_senha(np))
                                )
                                conn.commit()
                                registrar_auditoria(st.session_state.usuario_nome, "CRIAR_USUARIO",
                                    f"Novo usuário: {nu} — Setor: {ns}")
                                st.success(f"{nn} criado!")
                                time.sleep(0.5)
                                st.rerun()
                            except Exception as e:
                                conn.rollback()
                                st.error(f"Erro: {e}")
                            finally:
                                liberar_conexao(conn)

            conn_cfg = conectar_banco()
            try:
                df_u = pd.read_sql_query("SELECT id, usuario, nome, setor FROM usuarios ORDER BY id", conn_cfg)
            finally:
                liberar_conexao(conn_cfg)
            st.dataframe(df_u, hide_index=True, use_container_width=True)

            if len(df_u) > 1:
                del_u = st.selectbox("Remover usuario:", df_u['usuario'].tolist())
                if del_u == 'master':
                    st.caption("Conta master nao pode ser removida.")
                else:
                    if st.button(f"Excluir {del_u}"):
                        conn = conectar_banco()
                        try:
                            cursor = conn.cursor()
                            cursor.execute("DELETE FROM usuarios WHERE usuario=%s", (del_u,))
                            conn.commit()
                            registrar_auditoria(st.session_state.usuario_nome, "REMOVER_USUARIO",
                                f"Usuário removido: {del_u}")
                            st.toast("Removido!")
                            time.sleep(0.5)
                            st.rerun()
                        except Exception as e:
                            conn.rollback()
                            st.error(f"Erro: {e}")
                        finally:
                            liberar_conexao(conn)

            # ── Redefinir senha de usuário ─────────────────────
            st.markdown("---")
            st.markdown("### 🔑 Redefinir senha de usuário")
            reset_u = st.selectbox("Usuário:", df_u['usuario'].tolist(), key="reset_senha_usuario")
            with st.form("form_reset_senha_usuario", clear_on_submit=True):
                nova_senha_reset  = st.text_input("Nova senha:", type="password", key="nova_senha_reset")
                conf_senha_reset  = st.text_input("Confirmar nova senha:", type="password", key="conf_senha_reset")
                if st.form_submit_button("Redefinir senha"):
                    if not nova_senha_reset:
                        st.error("Digite a nova senha.")
                    elif len(nova_senha_reset) < 8:
                        st.error("A nova senha precisa ter pelo menos 8 caracteres.")
                    elif nova_senha_reset != conf_senha_reset:
                        st.error("A confirmação não confere com a nova senha.")
                    else:
                        ok, msg = redefinir_senha_usuario(reset_u, nova_senha_reset)
                        if ok:
                            registrar_auditoria(st.session_state.usuario_nome, "REDEFINIR_SENHA",
                                f"Senha redefinida para o usuário: {reset_u}")
                            st.success(msg)
                        else:
                            st.error(msg)

            # ── Log de Auditoria ──────────────────────────────
            st.markdown("---")
            st.markdown("### 📋 Log de Auditoria")
            with st.expander("Ver registros de auditoria", expanded=False):
                conn_aud = conectar_banco()
                try:
                    df_aud = pd.read_sql_query(
                        "SELECT usuario, acao, detalhes, criado_em FROM auditoria_log ORDER BY id DESC LIMIT 200",
                        conn_aud
                    )
                finally:
                    liberar_conexao(conn_aud)
                if df_aud.empty:
                    st.info("Nenhum registro ainda.")
                else:
                    filtro_aud = st.selectbox(
                        "Filtrar por ação:",
                        ["Todas"] + sorted(df_aud["acao"].unique().tolist()),
                        key="filtro_aud"
                    )
                    if filtro_aud != "Todas":
                        df_aud = df_aud[df_aud["acao"] == filtro_aud]
                    st.dataframe(df_aud, hide_index=True, use_container_width=True)
                    st.caption(f"Exibindo últimos {len(df_aud)} registros.")

            st.markdown("---")
            st.markdown("### ⚠️ Reset Geral")
            st.error("Esta ação remove TODOS os dados permanentemente e não pode ser desfeita.")
            confirma_reset = st.text_input("Digite CONFIRMAR para habilitar o reset:")
            if confirma_reset == "CONFIRMAR":
                if st.button("🗑️ Executar limpeza total", type="primary"):
                    sucesso = resetar_banco_dados_completo(st.session_state.usuario_nome)
                    if sucesso:
                        st.toast("Sistema resetado!")
                        time.sleep(1)
                        st.rerun()
            else:
                st.caption("Digite CONFIRMAR no campo acima para liberar o botão de reset.")

    # ==================================================
    # RELATORIO GERAL
    # ==================================================
    if nome_aba == "Relatorio Geral":
        with aba_objeto:
            st.markdown("""
            <div style="display:flex;align-items:center;gap:16px;margin-bottom:8px;">
                <div style="background:linear-gradient(135deg,#0F172A,#334155);padding:14px 20px;border-radius:10px;flex:1;">
                    <span style="color:#1A56DB;font-weight:800;font-size:1.4rem;letter-spacing:-0.02em;">Relatório Geral de Produção</span>
                    <span style="color:#94A3B8;font-size:0.85rem;margin-left:16px;">Todas as OPs ativas, envios parciais e status em tempo real</span>
                </div>
            </div>
            """, unsafe_allow_html=True)

            df_banco_micro_rel = carregar_micro_completo()

            # ── Filtros ──────────────────────────────────────────
            rel_f1, rel_f2, rel_f3 = st.columns([2, 2, 2])
            obras_rel = ["Todas"] + sorted(df_banco_micro_rel['Obra_Vinculada'].dropna().unique().tolist()) if not df_banco_micro_rel.empty else ["Todas"]
            with rel_f1:
                filtro_obra_rel = st.selectbox("Obra:", obras_rel, key="rel_obra")
            with rel_f2:
                status_opcoes = ["Todos", "Pendente", "Liberado para Fabrica", "Parcialmente Concluido", "Concluido"]
                filtro_status_rel = st.selectbox("Status:", status_opcoes, key="rel_status")
            with rel_f3:
                escopo_opcoes = ["Todos"] + sorted(df_banco_micro_rel['Tipo_Material'].dropna().unique().tolist()) if not df_banco_micro_rel.empty else ["Todos"]
                filtro_escopo_rel = st.selectbox("Escopo / Material:", escopo_opcoes, key="rel_escopo")

            rel_f4, rel_f5, rel_f6 = st.columns([2, 2, 2])
            with rel_f4:
                rel_campo_data = st.selectbox(
                    "Filtrar data por:", ["Entrada em Produção", "Data Limite"],
                    key="rel_campo_data"
                )
            with rel_f5:
                rel_dt_ini = st.date_input("De:", value=None, format="DD/MM/YYYY", key="rel_dt_ini")
            with rel_f6:
                rel_dt_fim = st.date_input("Até:", value=None, format="DD/MM/YYYY", key="rel_dt_fim")

            mostrar_concluidos = st.toggle("Ver concluídos", value=False, key="rel_concl")

            # ── Montar dataframe filtrado ────────────────────────
            df_rel = df_banco_micro_rel.copy() if not df_banco_micro_rel.empty else pd.DataFrame()

            if not df_rel.empty:
                if not mostrar_concluidos:
                    df_rel = df_rel[df_rel['Status_Item'] != 'Concluido']
                if filtro_obra_rel != "Todas":
                    df_rel = df_rel[df_rel['Obra_Vinculada'] == filtro_obra_rel]
                if filtro_status_rel != "Todos":
                    df_rel = df_rel[df_rel['Status_Item'] == filtro_status_rel]
                if filtro_escopo_rel != "Todos":
                    df_rel = df_rel[df_rel['Tipo_Material'] == filtro_escopo_rel]
                col_data_map = {
                    "Entrada em Produção": "Data_Producao_Programada",
                    "Data Limite": "Data_Limite_Obra"
                }
                col_dt_filtro = col_data_map[rel_campo_data]
                df_rel[col_dt_filtro] = pd.to_datetime(df_rel[col_dt_filtro], errors='coerce')
                if rel_dt_ini:
                    df_rel = df_rel[df_rel[col_dt_filtro] >= pd.Timestamp(rel_dt_ini)]
                if rel_dt_fim:
                    df_rel = df_rel[df_rel[col_dt_filtro] <= pd.Timestamp(rel_dt_fim)]

            # ── KPIs ─────────────────────────────────────────────
            st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
            if not df_rel.empty:
                total_ops     = df_rel['Num_OP'].nunique()
                total_m2      = df_rel['M2_Item'].sum()
                total_cx      = df_rel['Qtd_Caixas'].sum()
                ops_atrasadas = df_rel[
                    (pd.to_datetime(df_rel['Data_Limite_Obra'], errors='coerce') < pd.Timestamp.now()) &
                    (~df_rel['Status_Item'].isin(['Concluido']))
                ]['Num_OP'].nunique()
                enviadas_parcial = df_rel[df_rel['Status_Item'].str.contains('Parcial|parcial', na=False)]['Num_OP'].nunique()

                k1, k2, k3, k4, k5 = st.columns(5)
                k1.metric("Total de OPs", total_ops)
                k2.metric("Total m²", f"{total_m2:,.1f}")
                k3.metric("Total Caixas", int(total_cx))
                k4.metric("OPs Atrasadas", ops_atrasadas, delta=f"-{ops_atrasadas}" if ops_atrasadas > 0 else None, delta_color="inverse")
                k5.metric("Envio Parcial", enviadas_parcial)
            else:
                st.info("Nenhum dado disponível para os filtros selecionados.")

            st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

            if not df_rel.empty:
                # ── Gráfico: m² por obra ──────────────────────────
                col_g1, col_g2 = st.columns(2)
                with col_g1:
                    df_m2_obra = df_rel.groupby('Obra_Vinculada')['M2_Item'].sum().reset_index()
                    df_m2_obra.columns = ['Obra', 'm²']
                    df_m2_obra = df_m2_obra.sort_values('m²', ascending=False)
                    fig_m2 = px.bar(
                        df_m2_obra, x='Obra', y='m²',
                        title='m² em Produção por Obra',
                        color='m²',
                        color_continuous_scale=[[0,'#334155'],[1,'#1A56DB']],
                        text_auto='.1f'
                    )
                    fig_m2.update_layout(
                        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                        font_family='Inter', title_font_size=14, showlegend=False,
                        coloraxis_showscale=False,
                        margin=dict(l=10,r=10,t=40,b=10),
                        xaxis=dict(tickfont=dict(size=11)),
                    )
                    fig_m2.update_traces(textfont_size=11, textposition='outside')
                    st.plotly_chart(fig_m2, use_container_width=True)

                with col_g2:
                    df_status_cnt = df_rel.groupby('Status_Item')['Num_OP'].nunique().reset_index()
                    df_status_cnt.columns = ['Status', 'OPs']
                    cores_status = {
                        'Liberado para Fabrica': '#334155',
                        'Aguardando Expedicao': '#D97706',
                        'Enviado Parcial': '#1A56DB',
                        'Concluido': '#059669',
                    }
                    fig_status = px.pie(
                        df_status_cnt, names='Status', values='OPs',
                        title='Distribuição por Status',
                        color='Status',
                        color_discrete_map=cores_status,
                        hole=0.45
                    )
                    fig_status.update_layout(
                        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                        font_family='Inter', title_font_size=14,
                        margin=dict(l=10,r=10,t=40,b=10),
                        legend=dict(font=dict(size=11))
                    )
                    fig_status.update_traces(textinfo='value+percent', textfont_size=11)
                    st.plotly_chart(fig_status, use_container_width=True)

                # ── Tabela detalhada ──────────────────────────────
                st.markdown("---")
                st.markdown("#### Detalhamento por OP")

                def badge_status(s):
                    cores = {
                        'Liberado para Fabrica': ('background:#EFF6FF;color:#1D4ED8', 'Lib. Fábrica'),
                        'Aguardando Expedicao':  ('background:#FEF3C7;color:#92400E', 'Ag. Expedição'),
                        'Enviado Parcial':       ('background:#EFF6FF;color:#1447C0', 'Env. Parcial'),
                        'Concluido':             ('background:#ECFDF5;color:#065F46', 'Concluído'),
                    }
                    estilo, label = cores.get(s, ('background:#F1F5F9;color:#334155', s))
                    return f'<span style="{estilo};padding:3px 9px;border-radius:5px;font-size:11px;font-weight:700">{label}</span>'

                hoje_ts = pd.Timestamp.now().normalize()
                cols_show = ['Obra_Vinculada','Num_OP','Tipo_Material','EDT_Vinculado',
                             'Qtd_Caixas','M2_Item','Data_Producao_Programada','Data_Limite_Obra','Status_Item',
                             'Em_Parada','Motivo_Parada']
                cols_show_exist = [c for c in cols_show if c in df_rel.columns]
                df_tabela = df_rel[cols_show_exist].copy()
                col_names = ['Obra','OP','Material','EDT/Lote','Caixas','m²','Ini Prod.','Limite','Status']
                if 'Em_Parada' in df_tabela.columns:
                    df_tabela['Situacao'] = df_tabela.apply(
                        lambda r: f"⛔ PARADA — {r.get('Motivo_Parada','')}" if r.get('Em_Parada') else '', axis=1
                    )
                    df_tabela = df_tabela.drop(columns=['Em_Parada','Motivo_Parada'], errors='ignore')
                    col_names = col_names + ['Situacao']
                df_tabela.columns = col_names

                # % de representação sobre o total filtrado
                total_m2_rel  = df_tabela['m²'].sum()
                total_cx_rel  = df_tabela['Caixas'].sum()
                if total_m2_rel > 0:
                    df_tabela['% m²'] = (df_tabela['m²'] / total_m2_rel * 100).round(1)
                    base_pct = 'm²'
                elif total_cx_rel > 0:
                    df_tabela['% m²'] = (df_tabela['Caixas'] / total_cx_rel * 100).round(1)
                    base_pct = 'Caixas'
                else:
                    df_tabela['% m²'] = 0.0
                    base_pct = 'm²'

                for col_dt in ['Ini Prod.','Limite']:
                    df_tabela[col_dt] = pd.to_datetime(df_tabela[col_dt], errors='coerce').dt.strftime('%d/%m/%Y')

                # Highlight atraso
                def highlight_row(row):
                    limite = pd.to_datetime(row['Limite'], format='%d/%m/%Y', errors='coerce')
                    if pd.notna(limite) and limite < hoje_ts and row['Status'] not in ['Concluído']:
                        return ['background-color:#FEF2F2'] * len(row)
                    return [''] * len(row)

                styled = df_tabela.style.apply(highlight_row, axis=1).format({'m²': '{:.2f}', '% m²': '{:.1f}%'})
                st.dataframe(styled, hide_index=True, use_container_width=True, height=420)

                st.caption(f"Total de {len(df_tabela)} registros | % calculada sobre {base_pct} do total filtrado | Linhas em vermelho = prazo vencido")

                # ── Exportar Excel formatado ───────────────────────
                import io
                from openpyxl import Workbook
                from openpyxl.styles import (PatternFill, Font, Alignment,
                                              Border, Side, GradientFill)
                from openpyxl.utils import get_column_letter

                def gerar_excel_relatorio(df_exp, titulo_filtro="Todas as Obras"):
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Relatório de Produção"

                    # Paleta
                    cor_header_dark = "0F172A"
                    cor_header_accent = "1A56DB"
                    cor_sub = "1E293B"
                    cor_linha_par = "F8FAFC"
                    cor_linha_impar = "FFFFFF"
                    cor_atrasado = "FEE2E2"
                    cor_texto_branco = "FFFFFF"
                    cor_texto_escuro = "1E293B"

                    thin = Side(style='thin', color="E2E8F0")
                    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

                    # ── Linha 1: título geral ────────────────────
                    ws.merge_cells("A1:J1")
                    ws["A1"] = "PASSOLD — SISTEMAS DE FACHADAS"
                    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color=cor_texto_branco)
                    ws["A1"].fill = PatternFill("solid", fgColor=cor_header_dark)
                    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
                    ws.row_dimensions[1].height = 32

                    # ── Linha 2: subtítulo ───────────────────────
                    ws.merge_cells("A2:J2")
                    ws["A2"] = f"Relatório Geral de Produção  |  Obra: {titulo_filtro}  |  Gerado em: {datetime.now(FUSO_BR).strftime('%d/%m/%Y %H:%M')}"
                    ws["A2"].font = Font(name="Calibri", size=10, color="94A3B8", italic=True)
                    ws["A2"].fill = PatternFill("solid", fgColor=cor_sub)
                    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
                    ws.row_dimensions[2].height = 20

                    # ── Linha 3: espaço ──────────────────────────
                    ws.row_dimensions[3].height = 6

                    # ── Linha 4: cabeçalho das colunas ──────────
                    cabecalhos = list(df_exp.columns)
                    for col_idx, cab in enumerate(cabecalhos, start=1):
                        cell = ws.cell(row=4, column=col_idx, value=cab.upper())
                        cell.font = Font(name="Calibri", bold=True, size=10, color=cor_texto_branco)
                        cell.fill = PatternFill("solid", fgColor=cor_header_accent)
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        cell.border = borda
                    ws.row_dimensions[4].height = 22

                    # ── Linhas de dados ──────────────────────────
                    hoje_str = datetime.now(FUSO_BR).strftime('%d/%m/%Y')
                    for row_idx, row_data in enumerate(df_exp.itertuples(index=False), start=5):
                        is_par = (row_idx % 2 == 0)
                        limite_val = str(row_data[7]) if len(row_data) > 7 else ""
                        status_val = str(row_data[8]) if len(row_data) > 8 else ""
                        atrasado = False
                        try:
                            lim_dt = datetime.strptime(limite_val, '%d/%m/%Y')
                            atrasado = lim_dt < datetime.now() and status_val not in ['Concluído', 'Concluido']
                        except Exception:
                            pass

                        bg = cor_atrasado if atrasado else (cor_linha_par if is_par else cor_linha_impar)
                        for col_idx, valor in enumerate(row_data, start=1):
                            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                            cell.font = Font(name="Calibri", size=10, color=cor_texto_escuro)
                            cell.fill = PatternFill("solid", fgColor=bg)
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.border = borda
                        ws.row_dimensions[row_idx].height = 18

                    # ── Larguras automáticas ──────────────────────
                    larguras = [14, 20, 18, 18, 10, 10, 14, 14, 20, 12]
                    for i, larg in enumerate(larguras, start=1):
                        ws.column_dimensions[get_column_letter(i)].width = larg

                    # ── Linha de rodapé ───────────────────────────
                    ultima = ws.max_row + 2
                    ws.merge_cells(f"A{ultima}:J{ultima}")
                    ws[f"A{ultima}"] = f"Total de {len(df_exp)} registros  |  Linhas em vermelho = prazo vencido"
                    ws[f"A{ultima}"].font = Font(name="Calibri", size=9, italic=True, color="64748B")
                    ws[f"A{ultima}"].alignment = Alignment(horizontal="right")

                    buf = io.BytesIO()
                    wb.save(buf)
                    buf.seek(0)
                    return buf.getvalue()

                titulo_filtro_excel = filtro_obra_rel if filtro_obra_rel != "Todas" else "Todas as Obras"
                excel_bytes = gerar_excel_relatorio(df_tabela, titulo_filtro_excel)
                st.download_button(
                    label="Baixar relatório em Excel",
                    data=excel_bytes,
                    file_name=f"relatorio_producao_{datetime.now(FUSO_BR).strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_rel_xlsx"
                )

            # ── Seção: OPs Avulsas ────────────────────────────────
            st.markdown("---")
            with st.expander("OPs Avulsas cadastradas", expanded=False):
                df_avulsas = df_banco_micro_rel[df_banco_micro_rel['EDT_Vinculado'].str.startswith('AVULSO', na=False)].copy() if not df_banco_micro_rel.empty else pd.DataFrame()
                if filtro_obra_rel != "Todas" and not df_avulsas.empty:
                    df_avulsas = df_avulsas[df_avulsas['Obra_Vinculada'] == filtro_obra_rel]
                if df_avulsas.empty:
                    st.info("Nenhuma OP avulsa encontrada.")
                else:
                    cols_av = ['Obra_Vinculada','Num_OP','Tipo_Material','Qtd_Caixas','M2_Item',
                               'Data_Producao_Programada','Data_Limite_Obra','Status_Item','Romaneio_Chapas']
                    df_avulsas = df_avulsas[cols_av].copy()
                    df_avulsas.columns = ['Obra','OP','Material','Caixas','m²','Ini Prod.','Limite','Status','Detalhes']
                    for col_dt in ['Ini Prod.','Limite']:
                        df_avulsas[col_dt] = pd.to_datetime(df_avulsas[col_dt], errors='coerce').dt.strftime('%d/%m/%Y')
                    st.dataframe(df_avulsas.style.format({'m²': '{:.2f}'}), hide_index=True, use_container_width=True)
                    st.caption(f"{len(df_avulsas)} OP(s) avulsa(s) encontrada(s).")

